#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
README / 実行方法
-----------------
- 前提: Python 3.x, openpyxl がインストール済み
- 実行例:
    python beam_optimizer.py input.xlsx output.xlsx

- 入力:
    Excelファイル（.xlsx）内のシート "INPUT" を参照
- 出力:
    指定した output.xlsx に、シート "RESULT" を生成して書き込み

設計モデル（固定仮定）
---------------------
- 小梁: 両端ピンの単純支持梁
- 荷重: 等分布荷重（q*支配幅 + 自重w_g） + 複数集中荷重（配分ルールで小梁へ割当）
- 断面照査:
    曲げ:   Mmax / Z <= fb
    せん断: Vmax / Av <= fv  （Av無い場合は近似。限界はコメント参照）
    たわみ: δmax <= L / deflection_limit

- たわみ算定: 曲率 k(x)=M(x)/(E*I) を数値積分（分割数 n_div）し、
  y(0)=0, y(L)=0 を満たすように初期回転 θ0 を調整する。

注意（自重反復枠）
-----------------
断面固定なら自重w_gは固定なので通常1回で収束します。
将来、断面更新ロジックを追加する拡張を想定して反復枠を残しています。
"""

from __future__ import annotations

import math
import re
import sys
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# -----------------------------
# Data models
# -----------------------------
@dataclass(frozen=True)
class PointLoad:
    load_id: str
    P: float     # kN
    x: float     # m (0..Lx)
    y: float     # m (0..Ly)


@dataclass(frozen=True)
class Section:
    rank: int
    name: str
    w_g: float      # kN/m
    h: Optional[float] = None   # mm
    b: Optional[float] = None   # mm
    tw: Optional[float] = None  # mm
    tf: Optional[float] = None  # mm
    A_mm2: Optional[float] = None
    Z_mm3: Optional[float] = None
    I_mm4: Optional[float] = None
    Av_mm2: Optional[float] = None


@dataclass(frozen=True)
class Material:
    E_kN_m2: float     # kN/m^2
    fb_kN_m2: float    # kN/m^2
    fv_kN_m2: float    # kN/m^2
    deflection_limit: float


@dataclass(frozen=True)
class SolverSettings:
    tol: float
    max_iter: int
    n_div: int


@dataclass(frozen=True)
class Config:
    Lx: float
    Ly: float
    q: float
    loads: List[PointLoad]

    enable_x: bool
    enable_y: bool
    edge_beams: bool
    load_rule: int  # 1 or 2

    pitch_start: float
    pitch_end: float
    pitch_step: float
    pitch_list: List[float]  # if not empty, overrides start/end/step

    short_pitch_limit: float  # fixed display value (3.0)


@dataclass
class Beam:
    beam_id: str
    direction: str  # 'X' or 'Y'
    pos: float      # position along pitch direction (m)
    span: float     # m
    trib_width: float  # m
    # concentrated loads on this beam: list of (P[kN], a[m] from left end)
    point_along: List[Tuple[float, float]]


@dataclass
class CandidateResult:
    direction: str
    pitch: float
    n_beams: int
    section_rank: int
    section_name: str

    w_total_max: float   # max(kN/m) among beams
    Mmax: float          # kN*m (max among beams)
    Vmax: float          # kN (max among beams)
    dmax: float          # m (max among beams)

    util_M: float
    util_V: float
    util_d: float
    util_max: float

    total_weight: float  # kN (sum w_g * span * n_beams)
    ok: bool
    ng_reason: str

    # For reporting (only for the candidate itself)
    allocation_rows: List[Tuple[str, str, float]]  # (LoadID, BeamID, AllocP[kN])


# -----------------------------
# Utility: unit conversions
# -----------------------------
def nmm2_to_kN_m2(val_nmm2: float) -> float:
    # 1 N/mm2 = 1 MPa = 1e6 N/m2 = 1e3 kN/m2
    return val_nmm2 * 1000.0


def ensure_positive(name: str, v: float) -> None:
    if v is None or not isinstance(v, (int, float)) or v <= 0:
        raise ValueError(f"{name} must be positive. got={v}")


# -----------------------------
# Section properties (approx)
# -----------------------------
def parse_h_section_dims(name: str) -> Optional[Tuple[float, float, float, float]]:
    """
    Parse "H-350x175x7x11" or similar (x, X, ×).
    Returns (h,b,tw,tf) in mm if parse succeeds.
    """
    s = name.replace("×", "x").replace("X", "x")
    m = re.search(r"(\d+(\.\d+)?)x(\d+(\.\d+)?)x(\d+(\.\d+)?)x(\d+(\.\d+)?)", s)
    if not m:
        return None
    h = float(m.group(1))
    b = float(m.group(3))
    tw = float(m.group(5))
    tf = float(m.group(7))
    return h, b, tw, tf


def approx_h_section_props_mm(h: float, b: float, tw: float, tf: float) -> Tuple[float, float, float, float]:
    """
    Fillet ignored. Strong-axis bending (Ixx, Zx).
    Returns (A_mm2, I_mm4, Z_mm3, Av_mm2).
    """
    A = 2.0 * b * tf + (h - 2.0 * tf) * tw

    # Ixx = 2*(b*tf^3/12 + b*tf*y^2) + tw*(h-2tf)^3/12
    I_flange = b * tf**3 / 12.0
    y = h / 2.0 - tf / 2.0
    I = 2.0 * (I_flange + b * tf * y**2) + tw * (h - 2.0 * tf)**3 / 12.0
    Z = I / (h / 2.0)

    Av = tw * (h - 2.0 * tf)  # web area approximation
    return A, I, Z, Av


def get_section_props_m(section: Section) -> Tuple[float, float, float, float]:
    """
    Returns (A[m2], Av[m2], Z[m3], I[m4]).
    If inputs missing, tries to parse from name and approximate.
    """
    h, b, tw, tf = section.h, section.b, section.tw, section.tf

    if any(v is None for v in (h, b, tw, tf)):
        dims = parse_h_section_dims(section.name)
        if dims:
            h, b, tw, tf = dims

    if any(v is None for v in (h, b, tw, tf)):
        raise ValueError(f"Section dims missing and cannot parse from name: {section.name}")

    A_mm2 = section.A_mm2
    Z_mm3 = section.Z_mm3
    I_mm4 = section.I_mm4
    Av_mm2 = section.Av_mm2

    if any(v is None for v in (A_mm2, Z_mm3, I_mm4, Av_mm2)):
        A2, I2, Z2, Av2 = approx_h_section_props_mm(float(h), float(b), float(tw), float(tf))
        if A_mm2 is None:
            A_mm2 = A2
        if I_mm4 is None:
            I_mm4 = I2
        if Z_mm3 is None:
            Z_mm3 = Z2
        if Av_mm2 is None:
            Av_mm2 = Av2

    # Convert: mm2->m2 (1e-6), mm3->m3 (1e-9), mm4->m4 (1e-12)
    A = float(A_mm2) * 1e-6
    Av = float(Av_mm2) * 1e-6
    Z = float(Z_mm3) * 1e-9
    I = float(I_mm4) * 1e-12

    # Safety check
    ensure_positive(f"A({section.name})", A)
    ensure_positive(f"Av({section.name})", Av)
    ensure_positive(f"Z({section.name})", Z)
    ensure_positive(f"I({section.name})", I)
    return A, Av, Z, I


# -----------------------------
# Geometry / beam layout
# -----------------------------
def make_pitch_candidates(cfg: Config) -> List[float]:
    if cfg.pitch_list:
        vals = [float(v) for v in cfg.pitch_list if v and float(v) > 0]
        return sorted(set(vals))
    ensure_positive("pitch_start", cfg.pitch_start)
    ensure_positive("pitch_end", cfg.pitch_end)
    ensure_positive("pitch_step", cfg.pitch_step)
    if cfg.pitch_end < cfg.pitch_start:
        raise ValueError("pitch_end must be >= pitch_start")

    out = []
    x = cfg.pitch_start
    # inclusive with tolerance
    while x <= cfg.pitch_end + 1e-12:
        out.append(round(x, 10))
        x += cfg.pitch_step
    return sorted(set(out))


def positions_along(width: float, pitch: float, edge_beams: bool) -> List[float]:
    """
    Generate beam centerlines along pitch direction from 0..width.
    edge_beams=True: includes 0 and width.
    edge_beams=False: starts at pitch/2, ends at width - pitch/2 (if possible).
    """
    ensure_positive("width", width)
    ensure_positive("pitch", pitch)

    pos: List[float] = []
    if edge_beams:
        pos = [0.0]
        k = 1
        while k * pitch < width - 1e-9:
            pos.append(k * pitch)
            k += 1
        if abs(pos[-1] - width) > 1e-9:
            pos.append(width)
    else:
        k = 0
        while True:
            p = pitch / 2.0 + k * pitch
            if p > width - 1e-9:
                break
            pos.append(p)
            k += 1
        if not pos:
            pos = [width / 2.0]

    # clean
    pos = sorted(set(round(p, 10) for p in pos))
    return pos


def tributary_widths(pos: List[float], width: float) -> List[float]:
    """
    Tributary width from midpoints between adjacent beams and boundaries (0,width).
    This handles last bay not equal pitch.
    """
    n = len(pos)
    trib: List[float] = []
    for i, p in enumerate(pos):
        left = 0.0 if i == 0 else 0.5 * (pos[i - 1] + p)
        right = width if i == n - 1 else 0.5 * (p + pos[i + 1])
        trib.append(right - left)
    return trib


def pitch_direction_of(direction: str) -> str:
    # X方向配置: beams span X, are arrayed along Y => pitch_dir='Y'
    # Y方向配置: beams span Y, are arrayed along X => pitch_dir='X'
    return "Y" if direction == "X" else "X"


def short_side_axis(Lx: float, Ly: float) -> str:
    return "X" if Lx <= Ly else "Y"


# -----------------------------
# Load distribution (Rule-1 / Rule-2)
# -----------------------------
def allocate_point_loads_to_beams(
    beams: List[Beam],
    direction: str,
    loads: List[PointLoad],
    Lx: float,
    Ly: float,
    rule: int
) -> List[Tuple[str, str, float]]:
    """
    Fill beam.point_along and return allocation table rows (LoadID, BeamID, AllocP).
    """
    if rule not in (1, 2):
        raise ValueError("load_rule must be 1 or 2")

    # coordinate used for pitch direction
    if direction == "X":
        # beams located by y, span coordinate is x
        coord_pitch = "y"
        coord_span = "x"
        span_L = Lx
        pitch_positions = [b.pos for b in beams]
    else:
        coord_pitch = "x"
        coord_span = "y"
        span_L = Ly
        pitch_positions = [b.pos for b in beams]

    alloc_rows: List[Tuple[str, str, float]] = []

    def add_alloc(load_id: str, beam: Beam, P: float, a: float) -> None:
        beam.point_along.append((P, a))
        alloc_rows.append((load_id, beam.beam_id, P))

    for pl in loads:
        # validate coords
        if not (0.0 <= pl.x <= Lx) or not (0.0 <= pl.y <= Ly):
            raise ValueError(f"Point load out of range: {pl}")

        c = pl.y if coord_pitch == "y" else pl.x
        a = pl.x if coord_span == "x" else pl.y
        if not (0.0 <= a <= span_L):
            raise ValueError(f"Point load along-span out of range: {pl}")

        if rule == 1:
            # nearest beam
            idx = min(range(len(beams)), key=lambda i: abs(pitch_positions[i] - c))
            add_alloc(pl.load_id, beams[idx], pl.P, a)
            continue

        # Rule-2: distance ratio between adjacent beams
        if c <= pitch_positions[0] + 1e-9:
            add_alloc(pl.load_id, beams[0], pl.P, a)
            continue
        if c >= pitch_positions[-1] - 1e-9:
            add_alloc(pl.load_id, beams[-1], pl.P, a)
            continue

        # find interval [i,i+1]
        for i in range(len(pitch_positions) - 1):
            p0 = pitch_positions[i]
            p1 = pitch_positions[i + 1]
            if p0 - 1e-9 <= c <= p1 + 1e-9:
                if abs(c - p0) < 1e-9:
                    add_alloc(pl.load_id, beams[i], pl.P, a)
                elif abs(c - p1) < 1e-9:
                    add_alloc(pl.load_id, beams[i + 1], pl.P, a)
                else:
                    Lseg = p1 - p0
                    if Lseg <= 0:
                        raise ValueError("Invalid beam positions (non-increasing).")
                    P0 = pl.P * (p1 - c) / Lseg
                    P1 = pl.P * (c - p0) / Lseg
                    add_alloc(pl.load_id, beams[i], P0, a)
                    add_alloc(pl.load_id, beams[i + 1], P1, a)
                break

    return alloc_rows


def build_beams(cfg: Config, direction: str, pitch: float) -> Tuple[List[Beam], List[Tuple[str, str, float]]]:
    """
    Build beam list with positions & tributary widths, and assign point loads according to rule.
    """
    if direction == "X":
        span = cfg.Lx
        width = cfg.Ly
    else:
        span = cfg.Ly
        width = cfg.Lx

    pos = positions_along(width, pitch, cfg.edge_beams)
    trib = tributary_widths(pos, width)

    beams: List[Beam] = []
    for i, (p, tw) in enumerate(zip(pos, trib), start=1):
        beam_id = f"{direction}{i:02d}"
        beams.append(Beam(beam_id=beam_id, direction=direction, pos=p, span=span, trib_width=tw, point_along=[]))

    alloc_rows = allocate_point_loads_to_beams(
        beams=beams,
        direction=direction,
        loads=cfg.loads,
        Lx=cfg.Lx,
        Ly=cfg.Ly,
        rule=cfg.load_rule
    )
    return beams, alloc_rows


# -----------------------------
# Beam analysis (numerical)
# -----------------------------
def analyze_simply_supported(
    L: float,
    w_udl: float,
    point_loads: List[Tuple[float, float]],
    E: float,
    I: float,
    n_div: int
) -> Tuple[float, float, float]:
    """
    Returns (Mmax[kN*m], Vmax[kN], dmax[m]) for a simply-supported beam.
    Loads:
      - UDL w_udl [kN/m] along full span
      - point_loads: list of (P[kN], a[m] from left)

    Method:
      - Reactions from statics
      - V(x), M(x) by superposition (step function)
      - Deflection from curvature integration with y(0)=0,y(L)=0
    """
    ensure_positive("L", L)
    if n_div < 50:
        raise ValueError("n_div too small (>=50 recommended)")

    # sort point loads
    pls = [(float(P), float(a)) for P, a in point_loads]
    for P, a in pls:
        if not (0.0 <= a <= L):
            raise ValueError(f"Point load position out of range: a={a}, L={L}")
        if P < 0:
            raise ValueError("Negative point load not supported in this template.")
    pls.sort(key=lambda x: x[1])

    # reactions
    Ra = w_udl * L / 2.0
    Rb = w_udl * L / 2.0
    for P, a in pls:
        Ra += P * (L - a) / L
        Rb += P * a / L

    # discretize
    n = n_div + 1
    dx = L / n_div
    xs = [i * dx for i in range(n)]

    V = [Ra - w_udl * x for x in xs]
    M = [Ra * x - w_udl * x * x / 2.0 for x in xs]

    # apply point loads with Heaviside step
    for P, a in pls:
        for i, x in enumerate(xs):
            if x >= a - 1e-12:
                V[i] -= P
                M[i] -= P * (x - a)

    Mmax = max(abs(v) for v in M)
    Vmax = max(abs(v) for v in V)

    # curvature k = M/(E*I)
    k = [m / (E * I) for m in M]

    # integrate curvature -> slope with trapezoid
    K1 = [0.0] * n
    for i in range(1, n):
        K1[i] = K1[i - 1] + 0.5 * (k[i - 1] + k[i]) * dx

    # integrate K1 -> helper integral
    intK1 = [0.0] * n
    for i in range(1, n):
        intK1[i] = intK1[i - 1] + 0.5 * (K1[i - 1] + K1[i]) * dx

    # enforce y(L)=0 => theta0 = -int_0^L K1(s) ds / L
    theta0 = -intK1[-1] / L
    theta = [theta0 + v for v in K1]

    # integrate slope -> deflection (y(0)=0)
    y = [0.0] * n
    for i in range(1, n):
        y[i] = y[i - 1] + 0.5 * (theta[i - 1] + theta[i]) * dx

    dmax = max(abs(v) for v in y)
    return Mmax, Vmax, dmax


# -----------------------------
# Candidate evaluation / optimizer
# -----------------------------
def evaluate_candidate(
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    direction: str,
    pitch: float,
    section: Section
) -> CandidateResult:
    beams, alloc_rows = build_beams(cfg, direction, pitch)

    # section props
    _, Av, Z, I = get_section_props_m(section)

    # self weight fixed per section; iteration frame kept for future extension
    prev_key = None
    ng_reason = ""

    # per-beam maxima
    Mmax_all = 0.0
    Vmax_all = 0.0
    dmax_all = 0.0
    w_total_max = 0.0

    utilM_all = 0.0
    utilV_all = 0.0
    utild_all = 0.0

    for it in range(1, setts.max_iter + 1):
        # reset maxima each iter (future-proof)
        Mmax_all = Vmax_all = dmax_all = w_total_max = 0.0
        utilM_all = utilV_all = utild_all = 0.0

        for b in beams:
            w_q = cfg.q * b.trib_width
            w_total = w_q + section.w_g
            w_total_max = max(w_total_max, w_total)

            Mmax, Vmax, dmax = analyze_simply_supported(
                L=b.span,
                w_udl=w_total,
                point_loads=b.point_along,
                E=mat.E_kN_m2,
                I=I,
                n_div=setts.n_div
            )

            Mmax_all = max(Mmax_all, Mmax)
            Vmax_all = max(Vmax_all, Vmax)
            dmax_all = max(dmax_all, dmax)

            utilM = (Mmax / Z) / mat.fb_kN_m2
            utilV = (Vmax / Av) / mat.fv_kN_m2
            utilD = dmax / (b.span / mat.deflection_limit)

            utilM_all = max(utilM_all, utilM)
            utilV_all = max(utilV_all, utilV)
            utild_all = max(utild_all, utilD)

        key = (Mmax_all, dmax_all)
        if prev_key is not None:
            # relative difference
            dM = abs(key[0] - prev_key[0]) / max(abs(prev_key[0]), 1e-9)
            dd = abs(key[1] - prev_key[1]) / max(abs(prev_key[1]), 1e-12)
            if max(dM, dd) < setts.tol:
                break
        prev_key = key

    # convergence note (in this version, usually 1 iter)
    if setts.max_iter > 0 and prev_key is None:
        ng_reason = "analysis not executed"

    util_max = max(utilM_all, utilV_all, utild_all)
    ok = util_max <= 1.0

    if not ok:
        reasons = []
        if utilM_all > 1.0:
            reasons.append(f"bending NG (util={utilM_all:.3f})")
        if utilV_all > 1.0:
            reasons.append(f"shear NG (util={utilV_all:.3f})")
        if utild_all > 1.0:
            reasons.append(f"deflection NG (util={utild_all:.3f})")
        ng_reason = "; ".join(reasons)

    total_weight = section.w_g * beams[0].span * len(beams) if beams else 0.0

    return CandidateResult(
        direction=direction,
        pitch=pitch,
        n_beams=len(beams),
        section_rank=section.rank,
        section_name=section.name,
        w_total_max=w_total_max,
        Mmax=Mmax_all,
        Vmax=Vmax_all,
        dmax=dmax_all,
        util_M=utilM_all,
        util_V=utilV_all,
        util_d=utild_all,
        util_max=util_max,
        total_weight=total_weight,
        ok=ok,
        ng_reason=ng_reason,
        allocation_rows=alloc_rows
    )


def optimize(cfg: Config, mat: Material, setts: SolverSettings, sections: List[Section]) -> Tuple[List[CandidateResult], Optional[CandidateResult], Optional[CandidateResult]]:
    # pitches
    pitches = make_pitch_candidates(cfg)

    # apply short-side pitch limit only when pitch direction is short side
    s_axis = short_side_axis(cfg.Lx, cfg.Ly)

    # build directions
    directions = []
    if cfg.enable_x:
        directions.append("X")
    if cfg.enable_y:
        directions.append("Y")
    if not directions:
        raise ValueError("Both X and Y directions are disabled.")

    results: List[CandidateResult] = []
    total_cases = len(directions) * len(pitches) * len(sections)
    done = 0
    ok_count = 0

    best_A: Optional[CandidateResult] = None
    best_B: Optional[CandidateResult] = None

    for d in directions:
        pdir = pitch_direction_of(d)
        for pitch in pitches:
            # pitch constraint check
            if pdir == s_axis and pitch > cfg.short_pitch_limit + 1e-9:
                continue

            for sec in sections:
                done += 1
                res = evaluate_candidate(cfg, mat, setts, d, pitch, sec)
                results.append(res)
                if res.ok:
                    ok_count += 1

                # Mode-A: min total_weight subject to OK
                if res.ok:
                    if best_A is None or (res.total_weight < best_A.total_weight - 1e-12) or (
                        abs(res.total_weight - best_A.total_weight) <= 1e-12 and res.util_max < best_A.util_max
                    ):
                        best_A = res
                        print(f"[Best-A updated] dir={res.direction} pitch={res.pitch} sec={res.section_name} W={res.total_weight:.3f} kN util={res.util_max:.3f}")

                # Mode-B: min max rank (here uniform section => rank), tie-break min total_weight
                if res.ok:
                    if best_B is None or (res.section_rank < best_B.section_rank) or (
                        res.section_rank == best_B.section_rank and res.total_weight < best_B.total_weight - 1e-12
                    ):
                        best_B = res
                        print(f"[Best-B updated] dir={res.direction} pitch={res.pitch} rank={res.section_rank} sec={res.section_name} W={res.total_weight:.3f} kN")

                if done % max(1, total_cases // 20) == 0:
                    ng_rate = 1.0 - (ok_count / max(done, 1))
                    print(f"[Progress] {done}/{total_cases} evaluated, OK={ok_count}, NG rate={ng_rate:.1%}")

    return results, best_A, best_B


# -----------------------------
# Excel I/O
# -----------------------------
def read_cell(ws, addr: str):
    return ws[addr].value


def read_bool(ws, addr: str, default: bool = False) -> bool:
    v = read_cell(ws, addr)
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "y", "on"):
        return True
    if s in ("false", "0", "no", "n", "off"):
        return False
    return default


def read_float(ws, addr: str, name: str, default: Optional[float] = None) -> float:
    v = read_cell(ws, addr)
    if v is None:
        if default is None:
            raise ValueError(f"Missing required value: {name} at {addr}")
        return float(default)
    try:
        return float(v)
    except Exception as e:
        raise ValueError(f"Invalid number for {name} at {addr}: {v}") from e


def read_input_xlsx(path: str) -> Tuple[Config, Material, SolverSettings, List[Section]]:
    wb = load_workbook(path, data_only=True)
    if "INPUT" not in wb.sheetnames:
        raise ValueError('Sheet "INPUT" not found.')
    ws = wb["INPUT"]

    Lx = read_float(ws, "B2", "Lx")
    Ly = read_float(ws, "B3", "Ly")
    q = read_float(ws, "B5", "q")

    # loads table A9:D...
    loads: List[PointLoad] = []
    r = 9
    while True:
        load_id = ws[f"A{r}"].value
        if load_id is None or str(load_id).strip() == "":
            break
        P = ws[f"B{r}"].value
        x = ws[f"C{r}"].value
        y = ws[f"D{r}"].value
        try:
            pl = PointLoad(load_id=str(load_id).strip(), P=float(P), x=float(x), y=float(y))
        except Exception as e:
            raise ValueError(f"Invalid point load row {r}: {load_id},{P},{x},{y}") from e
        loads.append(pl)
        r += 1

    enable_x = read_bool(ws, "B14", True)
    enable_y = read_bool(ws, "B15", True)
    edge_beams = read_bool(ws, "B16", True)
    load_rule = int(read_float(ws, "B17", "load_rule", 2))

    pitch_start = read_float(ws, "B20", "pitch_start", 1.5)
    pitch_end = read_float(ws, "B21", "pitch_end", 3.0)
    pitch_step = read_float(ws, "B22", "pitch_step", 0.5)
    short_pitch_limit = read_float(ws, "B23", "short_pitch_limit", 3.0)

    # optional pitch list at F20:F...
    pitch_list: List[float] = []
    rr = 20
    while True:
        v = ws[f"F{rr}"].value
        if v is None or str(v).strip() == "":
            break
        try:
            pitch_list.append(float(v))
        except Exception as e:
            raise ValueError(f"Invalid pitch_list at F{rr}: {v}") from e
        rr += 1

    # material
    E_val = read_float(ws, "B26", "E", 205000.0)
    E_unit = ws["C26"].value
    E_unit_s = str(E_unit).strip().lower() if E_unit is not None else "n/mm2"

    if E_unit_s in ("n/mm2", "n/mm^2", "mpa"):
        E_kN_m2 = nmm2_to_kN_m2(E_val)
    elif E_unit_s in ("kn/m2", "kn/m^2"):
        E_kN_m2 = float(E_val)
    else:
        raise ValueError('E_unit must be "N/mm2" or "kN/m2".')

    fb = read_float(ws, "B27", "fb", 165.0)
    fv = read_float(ws, "B28", "fv", 95.0)
    defl_lim = read_float(ws, "B29", "deflection_limit", 360.0)

    mat = Material(
        E_kN_m2=E_kN_m2,
        fb_kN_m2=nmm2_to_kN_m2(fb),  # N/mm2 -> kN/m2
        fv_kN_m2=nmm2_to_kN_m2(fv),
        deflection_limit=defl_lim
    )

    tol = read_float(ws, "B31", "tol", 1e-6)
    max_iter = int(read_float(ws, "B32", "max_iter", 5))
    n_div = int(read_float(ws, "B34", "n_div", 2000))
    setts = SolverSettings(tol=tol, max_iter=max_iter, n_div=n_div)

    # sections table A41:K...
    sections: List[Section] = []
    rr = 41
    while True:
        rk = ws[f"A{rr}"].value
        if rk is None or str(rk).strip() == "":
            break
        name = ws[f"B{rr}"].value
        w_g = ws[f"C{rr}"].value
        if name is None or str(name).strip() == "":
            raise ValueError(f"SectionName missing at row {rr}")
        if w_g is None:
            raise ValueError(f"w_g missing at row {rr}")

        def get_opt(col: str) -> Optional[float]:
            v = ws[f"{col}{rr}"].value
            if v is None or str(v).strip() == "":
                return None
            return float(v)

        sec = Section(
            rank=int(float(rk)),
            name=str(name).strip(),
            w_g=float(w_g),
            h=get_opt("D"),
            b=get_opt("E"),
            tw=get_opt("F"),
            tf=get_opt("G"),
            A_mm2=get_opt("H"),
            Z_mm3=get_opt("I"),
            I_mm4=get_opt("J"),
            Av_mm2=get_opt("K")
        )
        sections.append(sec)
        rr += 1

    sections = sorted(sections, key=lambda s: s.rank)

    cfg = Config(
        Lx=Lx, Ly=Ly, q=q, loads=loads,
        enable_x=enable_x, enable_y=enable_y,
        edge_beams=edge_beams, load_rule=load_rule,
        pitch_start=pitch_start, pitch_end=pitch_end, pitch_step=pitch_step,
        pitch_list=pitch_list,
        short_pitch_limit=short_pitch_limit
    )

    return cfg, mat, setts, sections


# -----------------------------
# Reporting to Excel "RESULT"
# -----------------------------
def set_col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def write_header(ws, row: int, col: int, text: str) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor="DDDDDD")
    c.alignment = Alignment(horizontal="center", vertical="center")


def draw_plan(ws, top: int, left: int, cfg: Config, best: CandidateResult) -> int:
    """
    Draw a simple plan (cell text) for best candidate.
    Returns bottom row index after drawing.
    """
    # build beams of best
    beams, _ = build_beams(cfg, best.direction, best.pitch)

    # decide grid resolution to fit ~ 50x35
    Lx, Ly = cfg.Lx, cfg.Ly
    max_cols = 50
    max_rows = 35

    nx = min(max_cols, max(10, int(Lx * 5) + 1))  # about 0.2m per cell if possible
    ny = min(max_rows, max(10, int(Ly * 5) + 1))

    dx = Lx / (nx - 1)
    dy = Ly / (ny - 1)

    # Title
    ws.cell(row=top, column=left, value=f"PLAN (best): dir={best.direction}, pitch={best.pitch}, sec={best.section_name}").font = Font(bold=True)

    # clear area
    for r in range(top + 1, top + 1 + ny):
        for c in range(left, left + nx):
            ws.cell(row=r, column=c, value="")

    # boundary
    for i in range(nx):
        ws.cell(row=top + 1, column=left + i, value="-")
        ws.cell(row=top + ny, column=left + i, value="-")
    for j in range(ny):
        ws.cell(row=top + 1 + j, column=left, value="|")
        ws.cell(row=top + 1 + j, column=left + nx - 1, value="|")
    ws.cell(row=top + 1, column=left, value="+")
    ws.cell(row=top + 1, column=left + nx - 1, value="+")
    ws.cell(row=top + ny, column=left, value="+")
    ws.cell(row=top + ny, column=left + nx - 1, value="+")

    # beams
    if best.direction == "X":
        # horizontal lines at y=pos
        for b in beams:
            y = b.pos
            rr = top + 1 + int(round(y / dy))
            rr = min(max(top + 1, rr), top + ny)
            for cc in range(left + 1, left + nx - 1):
                ws.cell(row=rr, column=cc, value="—")
    else:
        # vertical lines at x=pos
        for b in beams:
            x = b.pos
            cc = left + int(round(x / dx))
            cc = min(max(left, cc), left + nx - 1)
            for rr in range(top + 2, top + ny):
                ws.cell(row=rr, column=cc, value="|")

    return top + ny + 2


def write_result_xlsx(
    in_path: str,
    out_path: str,
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    results: List[CandidateResult],
    best_A: Optional[CandidateResult],
    best_B: Optional[CandidateResult]
) -> None:
    wb = load_workbook(in_path)
    if "RESULT" in wb.sheetnames:
        del wb["RESULT"]
    ws = wb.create_sheet("RESULT")

    # column widths (rough)
    for i, w in enumerate([10, 8, 8, 8, 22, 12, 12, 12, 12, 12, 10, 10, 10, 10, 35], start=1):
        set_col_width(ws, i, w)

    r = 1
    ws.cell(row=r, column=1, value="Secondary Beam Optimization RESULT").font = Font(bold=True, size=14)
    r += 2

    ws.cell(row=r, column=1, value="Notes: Units are kN, m. fb/fv/E are converted from N/mm2 to kN/m2.").font = Font(italic=True)
    r += 2

    # Summary blocks
    def write_summary(title: str, best: Optional[CandidateResult], row0: int) -> int:
        ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
        headers = ["Direction", "Pitch[m]", "Section(Rank)", "TotalWeight[kN]", "Mmax[kN*m]", "Vmax[kN]", "dmax[mm]", "util(M)", "util(V)", "util(d)", "OK/NG"]
        for j, h in enumerate(headers, start=1):
            write_header(ws, row0 + 1, j, h)
        if best is None:
            ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
            return row0 + 4
        vals = [
            best.direction,
            best.pitch,
            f"{best.section_name} (R{best.section_rank})",
            round(best.total_weight, 6),
            round(best.Mmax, 6),
            round(best.Vmax, 6),
            round(best.dmax * 1000.0, 6),
            round(best.util_M, 6),
            round(best.util_V, 6),
            round(best.util_d, 6),
            "OK" if best.ok else f"NG: {best.ng_reason}",
        ]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=row0 + 2, column=j, value=v)
        return row0 + 4

    r = write_summary("Mode-A: Total weight minimum", best_A, r)
    r = write_summary("Mode-B: Minimize maximum section rank (tie-break: weight)", best_B, r)

    # Candidate table
    ws.cell(row=r, column=1, value="All candidates").font = Font(bold=True)
    r += 1
    headers = ["Dir", "Pitch[m]", "Nbeams", "Rank", "Section", "w_total_max[kN/m]", "Mmax[kN*m]", "Vmax[kN]", "dmax[mm]",
               "utilM", "utilV", "utilD", "utilMax", "TotalWeight[kN]", "OK/NG (reason)"]
    for j, h in enumerate(headers, start=1):
        write_header(ws, r, j, h)
    r += 1

    # Sort: OK first, then Mode-A like
    results_sorted = sorted(results, key=lambda x: (not x.ok, x.total_weight, x.section_rank, x.direction, x.pitch))
    for res in results_sorted:
        row = [
            res.direction, res.pitch, res.n_beams, res.section_rank, res.section_name,
            round(res.w_total_max, 6),
            round(res.Mmax, 6),
            round(res.Vmax, 6),
            round(res.dmax * 1000.0, 6),
            round(res.util_M, 6),
            round(res.util_V, 6),
            round(res.util_d, 6),
            round(res.util_max, 6),
            round(res.total_weight, 6),
            "OK" if res.ok else f"NG: {res.ng_reason}",
        ]
        for j, v in enumerate(row, start=1):
            ws.cell(row=r, column=j, value=v)
        r += 1

    r += 2

    # Allocation results for best solutions
    def write_alloc(title: str, best: Optional[CandidateResult], row0: int) -> int:
        ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
        write_header(ws, row0 + 1, 1, "LoadID")
        write_header(ws, row0 + 1, 2, "BeamID")
        write_header(ws, row0 + 1, 3, "AllocatedP[kN]")
        if best is None:
            ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
            return row0 + 4
        rr = row0 + 2
        for lid, bid, p in best.allocation_rows:
            ws.cell(row=rr, column=1, value=lid)
            ws.cell(row=rr, column=2, value=bid)
            ws.cell(row=rr, column=3, value=round(p, 6))
            rr += 1
        return rr + 1

    r = write_alloc("Allocation (Mode-A best)", best_A, r)
    r = write_alloc("Allocation (Mode-B best)", best_B, r)

    # Member list (simple: since uniform section, count=beams)
    def write_member_list(title: str, best: Optional[CandidateResult], row0: int) -> int:
        ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
        write_header(ws, row0 + 1, 1, "Section")
        write_header(ws, row0 + 1, 2, "Count")
        write_header(ws, row0 + 1, 3, "Length each [m]")
        write_header(ws, row0 + 1, 4, "w_g [kN/m]")
        write_header(ws, row0 + 1, 5, "Weight [kN]")
        if best is None:
            ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
            return row0 + 4
        beams, _ = build_beams(cfg, best.direction, best.pitch)
        # find w_g from candidate list
        # (store w_g in results as total_weight/(span*nbeams))
        w_g = best.total_weight / (beams[0].span * len(beams)) if beams else 0.0
        ws.cell(row=row0 + 2, column=1, value=best.section_name)
        ws.cell(row=row0 + 2, column=2, value=len(beams))
        ws.cell(row=row0 + 2, column=3, value=beams[0].span if beams else 0.0)
        ws.cell(row=row0 + 2, column=4, value=round(w_g, 6))
        ws.cell(row=row0 + 2, column=5, value=round(best.total_weight, 6))
        return row0 + 5

    r = write_member_list("Member list (Mode-A best)", best_A, r)
    r = write_member_list("Member list (Mode-B best)", best_B, r)

    # Plans
    r += 1
    if best_A is not None:
        r = draw_plan(ws, r, 1, cfg, best_A)
    if best_B is not None:
        r = draw_plan(ws, r, 1, cfg, best_B)

    wb.save(out_path)


# -----------------------------
# Main
# -----------------------------
def main(argv: List[str]) -> int:
    if len(argv) < 2:
        print("Usage: python beam_optimizer.py input.xlsx output.xlsx")
        return 2

    in_path = argv[1]
    out_path = argv[2] if len(argv) >= 3 else (in_path.replace(".xlsx", "") + "_out.xlsx")

    cfg, mat, setts, sections = read_input_xlsx(in_path)

    # basic validation
    ensure_positive("Lx", cfg.Lx)
    ensure_positive("Ly", cfg.Ly)
    ensure_positive("q", cfg.q)
    ensure_positive("E", mat.E_kN_m2)
    ensure_positive("fb", mat.fb_kN_m2)
    ensure_positive("fv", mat.fv_kN_m2)
    ensure_positive("deflection_limit", mat.deflection_limit)

    if not sections:
        raise ValueError("No sections provided.")

    # run
    results, best_A, best_B = optimize(cfg, mat, setts, sections)

    # write excel
    write_result_xlsx(in_path, out_path, cfg, mat, setts, results, best_A, best_B)

    print(f"Done. Output written to: {out_path}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main(sys.argv))
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)
