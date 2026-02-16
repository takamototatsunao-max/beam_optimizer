#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
beam_optimizer_b6.py

README / 実行方法
-----------------
- 前提: Python 3.x, openpyxl
- 実行例:
    python beam_optimizer_b6.py input.xlsx output.xlsx

入出力
------
- 入力:  Excelファイル（.xlsx）のシート "INPUT"
- 出力:  output.xlsx にシート "RESULT" と "LAYOUT" を生成して書き込み
         ※Excel図形は不使用。LAYOUTはセルのみの簡易伏図。

今回の修正点（ユーザー要望対応）
------------------------------
1) LAYOUTを正方形セル（見やすいグリッド）にするため、表（Beam checks / allocation 等）は RESULT シートへ移動。
   LAYOUT は「伏図のみ」に限定（文字潰れ防止）。
2) MAIN小梁は「1本ごとに断面が違ってOK」とし、各部材ごとに断面選定を行う。
3) 集中荷重位置にMAIN小梁が無い場合（= 配置ライン上に無い場合）、その位置(yまたはx)に TRANS小梁（黄色）を追加し、
   隣接する支持（MAIN小梁または外周大梁=PERIM）へ反力として荷重を伝達する。
   - TRANS小梁に作用する荷重: 集中荷重P + TRANS自重(w_g)
   - 支持へ伝達する荷重: 反力（静定梁の反力）＝「集中荷重＋自重」を作用位置で案分した値
   - PERIM（外周大梁）は存在すると仮定し、PERIM上には小梁を配置しない（ただし反力集計には登場）。

設計モデル（固定仮定）
----------------------
- 梁: 両端ピンの単純支持梁
- 荷重:
    MAIN: 等分布(q*支配幅 + 自重w_g) + (直接集中荷重) + (TRANS反力)
    TRANS: 等分布(自重w_g) + (集中荷重)
- 断面照査:
    曲げ:   (Mmax / Z) <= fb
    せん断: (Vmax / Av) <= fv   ※Av未入力時はweb面積近似（H形）
    たわみ: δmax <= L / deflection_limit
- たわみ算定: 曲率 k(x)=M(x)/(E*I) を数値積分（n_div分割）。
  精度と計算時間のバランスは n_div で調整（>=200推奨）。

最適化モード
------------
(Mode-A) 総重量最小:
    - 各部材ごとに「重量(w_g*長さ)が最小」になる断面を選定（OK断面の中で最軽量）。
    - 方向×ピッチの組合せを全探索し、総重量最小の案を採用。

(Mode-B) 最大断面ランク最小（ミニマックス）:
    - 方向×ピッチの組合せごとに、必要最大ランクをR=1..Maxで探索。
    - そのRで全員がOKとなる最小Rを採用（同R内は総重量最小でタイブレーク）。

注意（自重反復枠）
-----------------
断面固定なら自重は固定なので通常1回で十分です。本実装では将来の拡張に備え反復枠は残していますが、
現状は「断面選定後の解析」は1回で安定します。
"""

from __future__ import annotations

import math
import re
import sys
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# -----------------------------
# Data models
# -----------------------------
@dataclass(frozen=True)
class PointLoad:
    load_id: str
    P: float  # kN
    x: float  # m (0..Lx)
    y: float  # m (0..Ly)


@dataclass(frozen=True)
class Section:
    rank: int
    name: str
    w_g: float      # kN/m
    h: Optional[float] = None   # mm
    b: Optional[float] = None   # mm
    tw: Optional[float] = None  # mm
    tf: Optional[float] = None  # mm
    A_mm2: Optional[float] = None
    Z_mm3: Optional[float] = None
    I_mm4: Optional[float] = None
    Av_mm2: Optional[float] = None


@dataclass(frozen=True)
class Material:
    E_kN_m2: float     # kN/m^2
    fb_kN_m2: float    # kN/m^2
    fv_kN_m2: float    # kN/m^2
    deflection_limit: float


@dataclass(frozen=True)
class SolverSettings:
    tol: float
    max_iter: int
    n_div: int


@dataclass(frozen=True)
class Config:
    Lx: float
    Ly: float
    q: float
    loads: List[PointLoad]

    enable_x: bool
    enable_y: bool

    # edge_beams=True: ピッチ起点を外周(0)に取り、0, pitch, 2pitch...を生成（ただし0とLはPERIM扱いでMAINには置かない）
    # edge_beams=False: ピッチ/2起点で内側に配置
    edge_beams: bool

    # 旧仕様互換（現状は TRANS追加ロジック優先）。残しておくが、load_ruleは直接使わない。
    load_rule: int

    pitch_start: float
    pitch_end: float
    pitch_step: float
    pitch_list: List[float]

    short_pitch_limit: float  # 3.0
    snap_tol: float           # "on-beam" 判定の許容[m]


@dataclass
class BeamGeom:
    beam_id: str
    direction: str  # 'X' or 'Y'  (MAIN span direction)
    pos: float      # position along pitch axis (m)
    span: float     # m
    trib_width: float  # m
    # point loads list for analysis: (P[kN], a[m] from left)
    point_along: List[Tuple[float, float]]


@dataclass(frozen=True)
class TransferDef:
    tb_id: str
    span_dir: str         # 'X' or 'Y' (TRANS span direction)
    fixed_coord: float    # m (coordinate perpendicular to span_dir)
    left_support: str     # BeamID or 'PERIM'
    right_support: str    # BeamID or 'PERIM'
    left_pos: float       # m coordinate on span_dir
    right_pos: float      # m coordinate on span_dir
    a_tb: float           # m from left support along TRANS
    a_on_support: float   # m along support span direction (to place reaction)
    load_id: str
    P: float              # kN (original point load)


@dataclass(frozen=True)
class MemberCheck:
    beam_id: str
    member_type: str      # 'MAIN' or 'TRANS'
    direction: str        # span direction ('X'/'Y')
    pos_or_fixed: Optional[float]  # MAIN: pos along pitch, TRANS: fixed_coord
    left: Optional[float]          # TRANS: left_pos, else None
    right: Optional[float]         # TRANS: right_pos, else None
    span: float
    section_rank: int
    section_name: str
    w_g: float
    w_udl: float
    n_point: int
    Mmax: float
    Vmax: float
    dmax: float
    util_M: float
    util_V: float
    util_d: float
    util_max: float
    ok: bool


@dataclass(frozen=True)
class Solution:
    mode: str  # 'A' or 'B'
    direction: str
    pitch: float

    total_weight: float
    max_rank_used: int

    Mmax: float
    Vmax: float
    dmax: float
    util_max: float

    ok: bool
    ng_reason: str

    member_checks: List[MemberCheck]
    # For plan view and reporting
    main_geoms: List[BeamGeom]
    transfer_defs: List[TransferDef]
    allocation_rows: List[Tuple[str, str, float]]  # (LoadID, SupportID, AllocatedP[kN])
    worst_member_id: str


# -----------------------------
# Utility: unit conversions
# -----------------------------
def nmm2_to_kN_m2(val_nmm2: float) -> float:
    # 1 N/mm2 = 1 MPa = 1e6 N/m2 = 1e3 kN/m2
    return float(val_nmm2) * 1000.0


def ensure_positive(name: str, v: float) -> None:
    if v is None or not isinstance(v, (int, float)) or float(v) <= 0.0:
        raise ValueError(f"{name} must be positive. got={v}")


# -----------------------------
# Section properties (approx)
# -----------------------------
def parse_h_section_dims(name: str) -> Optional[Tuple[float, float, float, float]]:
    """
    Parse "H-350x175x7x11" or similar (x, X, ×).
    Returns (h,b,tw,tf) in mm if parse succeeds.
    """
    s = str(name).replace("×", "x").replace("X", "x")
    m = re.search(r"(\d+(\.\d+)?)x(\d+(\.\d+)?)x(\d+(\.\d+)?)x(\d+(\.\d+)?)", s)
    if not m:
        return None
    h = float(m.group(1))
    b = float(m.group(3))
    tw = float(m.group(5))
    tf = float(m.group(7))
    return h, b, tw, tf


def approx_h_section_props_mm(h: float, b: float, tw: float, tf: float) -> Tuple[float, float, float, float]:
    """
    Fillet ignored. Strong-axis bending (Ixx, Zx).
    Returns (A_mm2, I_mm4, Z_mm3, Av_mm2).
    """
    A = 2.0 * b * tf + (h - 2.0 * tf) * tw

    # Ixx = 2*(b*tf^3/12 + b*tf*y^2) + tw*(h-2tf)^3/12
    I_flange = b * tf**3 / 12.0
    y = h / 2.0 - tf / 2.0
    I = 2.0 * (I_flange + b * tf * y**2) + tw * (h - 2.0 * tf)**3 / 12.0
    Z = I / (h / 2.0)

    Av = tw * (h - 2.0 * tf)  # web area approximation
    return A, I, Z, Av


def get_section_props_m(section: Section) -> Tuple[float, float, float, float]:
    """
    Returns (A[m2], Av[m2], Z[m3], I[m4]).
    If inputs missing, tries to parse from name and approximate.
    """
    h, b, tw, tf = section.h, section.b, section.tw, section.tf

    if any(v is None for v in (h, b, tw, tf)):
        dims = parse_h_section_dims(section.name)
        if dims:
            h, b, tw, tf = dims

    if any(v is None for v in (h, b, tw, tf)):
        raise ValueError(f"Section dims missing and cannot parse from name: {section.name}")

    A_mm2 = section.A_mm2
    Z_mm3 = section.Z_mm3
    I_mm4 = section.I_mm4
    Av_mm2 = section.Av_mm2

    if any(v is None for v in (A_mm2, Z_mm3, I_mm4, Av_mm2)):
        A2, I2, Z2, Av2 = approx_h_section_props_mm(float(h), float(b), float(tw), float(tf))
        if A_mm2 is None:
            A_mm2 = A2
        if I_mm4 is None:
            I_mm4 = I2
        if Z_mm3 is None:
            Z_mm3 = Z2
        if Av_mm2 is None:
            Av_mm2 = Av2

    # Convert: mm2->m2 (1e-6), mm3->m3 (1e-9), mm4->m4 (1e-12)
    A = float(A_mm2) * 1e-6
    Av = float(Av_mm2) * 1e-6
    Z = float(Z_mm3) * 1e-9
    I = float(I_mm4) * 1e-12

    ensure_positive(f"A({section.name})", A)
    ensure_positive(f"Av({section.name})", Av)
    ensure_positive(f"Z({section.name})", Z)
    ensure_positive(f"I({section.name})", I)
    return A, Av, Z, I


# -----------------------------
# Geometry / beam layout
# -----------------------------
def pitch_direction_of(direction: str) -> str:
    # X方向配置: beams span X, are arrayed along Y => pitch_dir='Y'
    # Y方向配置: beams span Y, are arrayed along X => pitch_dir='X'
    return "Y" if direction == "X" else "X"


def short_side_axis(Lx: float, Ly: float) -> str:
    return "X" if Lx <= Ly else "Y"


def make_pitch_candidates(cfg: Config) -> List[float]:
    if cfg.pitch_list:
        vals = [float(v) for v in cfg.pitch_list if v is not None and str(v).strip() != "" and float(v) > 0]
        return sorted(set(vals))
    ensure_positive("pitch_start", cfg.pitch_start)
    ensure_positive("pitch_end", cfg.pitch_end)
    ensure_positive("pitch_step", cfg.pitch_step)
    if cfg.pitch_end < cfg.pitch_start:
        raise ValueError("pitch_end must be >= pitch_start")

    out = []
    x = cfg.pitch_start
    while x <= cfg.pitch_end + 1e-12:
        out.append(round(x, 10))
        x += cfg.pitch_step
    return sorted(set(out))


def positions_along(width: float, pitch: float, edge_beams: bool) -> List[float]:
    """
    Generate beam centerlines along pitch direction from 0..width.
    edge_beams=True: includes 0 and width (later excluded as PERIM).
    edge_beams=False: starts at pitch/2.
    """
    ensure_positive("width", width)
    ensure_positive("pitch", pitch)

    pos: List[float] = []
    if edge_beams:
        pos = [0.0]
        k = 1
        while k * pitch < width - 1e-9:
            pos.append(k * pitch)
            k += 1
        if abs(pos[-1] - width) > 1e-9:
            pos.append(width)
    else:
        k = 0
        while True:
            p = pitch / 2.0 + k * pitch
            if p > width - 1e-9:
                break
            pos.append(p)
            k += 1
        if not pos:
            pos = [width / 2.0]

    pos = sorted(set(round(p, 10) for p in pos))
    return pos


def tributary_widths(pos: List[float], width: float) -> List[float]:
    """
    Tributary width from midpoints between adjacent beams and boundaries (0,width).
    Note: PERIM strips are excluded automatically (handled by boundaries).
    """
    n = len(pos)
    trib: List[float] = []
    for i, p in enumerate(pos):
        left = 0.0 if i == 0 else 0.5 * (pos[i - 1] + p)
        right = width if i == n - 1 else 0.5 * (p + pos[i + 1])
        trib.append(right - left)
    return trib


def build_main_beams(cfg: Config, direction: str, pitch: float) -> List[BeamGeom]:
    """
    Build MAIN secondary beams (excluding PERIM at 0 and width).
    """
    if direction == "X":
        span = cfg.Lx
        width = cfg.Ly
    else:
        span = cfg.Ly
        width = cfg.Lx

    pos_all = positions_along(width, pitch, cfg.edge_beams)

    # Exclude PERIM lines (0 and width) from MAIN beams
    pos = [p for p in pos_all if p > cfg.snap_tol and p < width - cfg.snap_tol]
    if not pos:
        # still allow no MAIN beam (then point loads go to PERIM or TRANS between PERIMs)
        pos = []

    trib = tributary_widths(pos, width) if pos else []
    beams: List[BeamGeom] = []
    for i, (p, tw) in enumerate(zip(pos, trib), start=1):
        beam_id = f"{direction}{i:02d}"
        beams.append(BeamGeom(beam_id=beam_id, direction=direction, pos=p, span=span, trib_width=tw, point_along=[]))
    return beams


# -----------------------------
# Beam analysis (numerical)
# -----------------------------
def analyze_simply_supported(
    L: float,
    w_udl: float,
    point_loads: List[Tuple[float, float]],
    E: float,
    I: float,
    n_div: int
) -> Tuple[float, float, float]:
    """
    Returns (Mmax[kN*m], Vmax[kN], dmax[m]) for a simply-supported beam.
    Loads:
      - UDL w_udl [kN/m] along full span
      - point_loads: list of (P[kN], a[m] from left)
    """
    ensure_positive("L", L)
    if n_div < 50:
        raise ValueError("n_div too small (>=50 recommended)")

    pls = [(float(P), float(a)) for P, a in point_loads]
    for P, a in pls:
        if not (0.0 <= a <= L):
            raise ValueError(f"Point load position out of range: a={a}, L={L}")
        if P < 0:
            raise ValueError("Negative point load not supported.")
    pls.sort(key=lambda x: x[1])

    # reactions
    Ra = w_udl * L / 2.0
    Rb = w_udl * L / 2.0
    for P, a in pls:
        Ra += P * (L - a) / L
        Rb += P * a / L

    # discretize
    n = n_div + 1
    dx = L / n_div
    xs = [i * dx for i in range(n)]

    V = [Ra - w_udl * x for x in xs]
    M = [Ra * x - w_udl * x * x / 2.0 for x in xs]

    for P, a in pls:
        for i, x in enumerate(xs):
            if x >= a - 1e-12:
                V[i] -= P
                M[i] -= P * (x - a)

    Mmax = max(abs(v) for v in M)
    Vmax = max(abs(v) for v in V)

    # curvature k = M/(E*I)
    k = [m / (E * I) for m in M]

    # integrate curvature -> slope with trapezoid
    K1 = [0.0] * n
    for i in range(1, n):
        K1[i] = K1[i - 1] + 0.5 * (k[i - 1] + k[i]) * dx

    # integrate K1 -> helper integral
    intK1 = [0.0] * n
    for i in range(1, n):
        intK1[i] = intK1[i - 1] + 0.5 * (K1[i - 1] + K1[i]) * dx

    # enforce y(L)=0 => theta0 = -int_0^L K1(s) ds / L
    theta0 = -intK1[-1] / L
    theta = [theta0 + v for v in K1]

    # integrate slope -> deflection (y(0)=0)
    y = [0.0] * n
    for i in range(1, n):
        y[i] = y[i - 1] + 0.5 * (theta[i - 1] + theta[i]) * dx

    dmax = max(abs(v) for v in y)
    return Mmax, Vmax, dmax


def check_member(
    member_type: str,
    span_dir: str,
    span: float,
    pos_or_fixed: Optional[float],
    left: Optional[float],
    right: Optional[float],
    trib_width: float,
    q: float,
    point_loads: List[Tuple[float, float]],
    section: Section,
    mat: Material,
    setts: SolverSettings,
) -> MemberCheck:
    """
    member_type='MAIN': w_udl = q*trib + w_g
    member_type='TRANS': w_udl = w_g
    """
    _, Av, Z, I = get_section_props_m(section)
    if member_type == "MAIN":
        w_udl = q * trib_width + section.w_g
    else:
        w_udl = section.w_g

    Mmax, Vmax, dmax = analyze_simply_supported(
        L=span, w_udl=w_udl, point_loads=point_loads, E=mat.E_kN_m2, I=I, n_div=setts.n_div
    )

    utilM = (Mmax / Z) / mat.fb_kN_m2
    utilV = (Vmax / Av) / mat.fv_kN_m2
    utilD = dmax / (span / mat.deflection_limit)
    utilMax = max(utilM, utilV, utilD)
    ok = utilMax <= 1.0

    return MemberCheck(
        beam_id="",
        member_type=member_type,
        direction=span_dir,
        pos_or_fixed=pos_or_fixed,
        left=left,
        right=right,
        span=span,
        section_rank=section.rank,
        section_name=section.name,
        w_g=section.w_g,
        w_udl=w_udl,
        n_point=len(point_loads),
        Mmax=Mmax,
        Vmax=Vmax,
        dmax=dmax,
        util_M=utilM,
        util_V=utilV,
        util_d=utilD,
        util_max=utilMax,
        ok=ok
    )


# -----------------------------
# Load distribution with TRANS beams
# -----------------------------
def build_system_defs(
    cfg: Config,
    direction: str,
    pitch: float,
) -> Tuple[List[BeamGeom], List[TransferDef], List[Tuple[str, str, float]]]:
    """
    Build MAIN geometries and TRANS definitions.
    Returns:
      main_beams: BeamGeom[] (point_along filled for on-beam loads only)
      transfer_defs: TransferDef[] for off-beam loads
      base_alloc_rows: allocations for on-beam and PERIM-direct loads only
                       (TRANS reactions added later after selecting TRANS section)
    """
    main_beams = build_main_beams(cfg, direction, pitch)

    # axis setup
    if direction == "X":
        pitch_axis = "Y"
        span_axis = "X"
        width = cfg.Ly
        span_L = cfg.Lx
        main_pos = [b.pos for b in main_beams]  # y
    else:
        pitch_axis = "X"
        span_axis = "Y"
        width = cfg.Lx
        span_L = cfg.Ly
        main_pos = [b.pos for b in main_beams]  # x

    # snap tol
    snap = max(cfg.snap_tol, 1e-6)

    # helper
    def add_point_to_main(beam_id: str, P: float, a: float) -> None:
        for b in main_beams:
            if b.beam_id == beam_id:
                b.point_along.append((P, a))
                return
        raise ValueError(f"Internal: main beam not found: {beam_id}")

    base_alloc: List[Tuple[str, str, float]] = []
    transfer_defs: List[TransferDef] = []
    tb_index = 0

    # support list (including PERIM as virtual supports at 0 and width)
    support_positions: List[Tuple[str, float]] = [("PERIM_L", 0.0)] + [(b.beam_id, b.pos) for b in main_beams] + [("PERIM_R", width)]

    for pl in cfg.loads:
        if not (0.0 <= pl.x <= cfg.Lx) or not (0.0 <= pl.y <= cfg.Ly):
            raise ValueError(f"Point load out of range: {pl}")

        c_pitch = pl.y if pitch_axis == "Y" else pl.x
        a_span = pl.x if span_axis == "X" else pl.y

        if not (0.0 <= a_span <= span_L):
            raise ValueError(f"Point load along-span out of range: {pl}")

        # PERIM direct if on perimeter line
        if c_pitch <= snap or c_pitch >= width - snap:
            base_alloc.append((pl.load_id, "PERIM", pl.P))
            continue

        # on-beam if matches a MAIN beam position
        on_beam_id: Optional[str] = None
        for b in main_beams:
            if abs(b.pos - c_pitch) <= snap:
                on_beam_id = b.beam_id
                break

        if on_beam_id is not None:
            add_point_to_main(on_beam_id, pl.P, a_span)
            base_alloc.append((pl.load_id, on_beam_id, pl.P))
            continue

        # off-beam => create TRANS beam at fixed coord = a_span? no: fixed coord is along span axis (y for Y-dir system, x for X-dir system)
        # TRANS spans along pitch axis between adjacent supports around c_pitch.
        # Find adjacent supports
        left_sup, left_pos = "PERIM", 0.0
        right_sup, right_pos = "PERIM", width

        # scan support_positions
        for (sid0, p0), (sid1, p1) in zip(support_positions[:-1], support_positions[1:]):
            if p0 - 1e-12 <= c_pitch <= p1 + 1e-12:
                # map PERIM_L/ PERIM_R to PERIM for reporting
                left_sup = "PERIM" if sid0.startswith("PERIM") else sid0
                right_sup = "PERIM" if sid1.startswith("PERIM") else sid1
                left_pos, right_pos = p0, p1
                break

        Ltb = right_pos - left_pos
        if Ltb <= 1e-9:
            # should not happen
            base_alloc.append((pl.load_id, "PERIM", pl.P))
            continue

        a_tb = c_pitch - left_pos

        tb_index += 1
        tb_id = f"TB_{pl.load_id}"
        # ensure uniqueness
        if any(t.tb_id == tb_id for t in transfer_defs):
            tb_id = f"TB_{pl.load_id}_{tb_index}"

        # TRANS span direction is pitch axis
        span_dir = pitch_axis

        transfer_defs.append(TransferDef(
            tb_id=tb_id,
            span_dir=span_dir,
            fixed_coord=a_span,  # fixed at span-axis coordinate
            left_support=left_sup,
            right_support=right_sup,
            left_pos=left_pos,
            right_pos=right_pos,
            a_tb=a_tb,
            a_on_support=a_span,
            load_id=pl.load_id,
            P=pl.P,
        ))

    return main_beams, transfer_defs, base_alloc


# -----------------------------
# Design / selection per mode
# -----------------------------
def choose_section_for_member(
    member_type: str,
    span_dir: str,
    span: float,
    pos_or_fixed: Optional[float],
    left: Optional[float],
    right: Optional[float],
    trib_width: float,
    q: float,
    point_loads: List[Tuple[float, float]],
    sections: List[Section],
    mat: Material,
    setts: SolverSettings,
    mode: str,
    rank_cap: Optional[int],
) -> Optional[MemberCheck]:
    """
    Return chosen MemberCheck with beam_id empty (caller fills),
    or None if no feasible section.
    Selection:
      - Mode A: among feasible sections, minimize weight = w_g * span
      - Mode B: among feasible sections with rank<=rank_cap, minimize weight = w_g * span
    """
    best: Optional[MemberCheck] = None
    best_weight = None

    for sec in sections:
        if rank_cap is not None and sec.rank > rank_cap:
            continue
        try:
            bc = check_member(
                member_type=member_type,
                span_dir=span_dir,
                span=span,
                pos_or_fixed=pos_or_fixed,
                left=left,
                right=right,
                trib_width=trib_width,
                q=q,
                point_loads=point_loads,
                section=sec,
                mat=mat,
                setts=setts,
            )
        except Exception:
            continue

        if not bc.ok:
            continue

        w = sec.w_g * span
        if best is None:
            best = bc
            best_weight = w
        else:
            # Mode-A & Mode-B both use weight tie-break at member level
            if w < float(best_weight) - 1e-12:
                best = bc
                best_weight = w
            elif abs(w - float(best_weight)) <= 1e-12 and sec.rank < best.section_rank:
                # stabilize
                best = bc
                best_weight = w

    return best


def reactions_simply_supported(L: float, w_udl: float, P: float, a: float) -> Tuple[float, float]:
    """Return (Ra,Rb) for UDL + one point load at a from left."""
    Ra = w_udl * L / 2.0 + P * (L - a) / L
    Rb = w_udl * L / 2.0 + P * a / L
    return Ra, Rb


def solve_layout_one_mode(
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    sections: List[Section],
    direction: str,
    pitch: float,
    mode: str,
) -> Solution:
    """
    Solve a given (direction,pitch) for one mode.
    mode='A' or 'B'
    """
    if mode not in ("A", "B"):
        raise ValueError("mode must be 'A' or 'B'")

    max_rank = max(s.rank for s in sections)

    def attempt(rank_cap: Optional[int]) -> Optional[Solution]:
        # fresh system
        main_beams, transfer_defs, base_alloc = build_system_defs(cfg, direction, pitch)

        # 1) Design TRANS beams first (choose section), then compute reactions to supports and add to MAIN beams
        member_checks: List[MemberCheck] = []
        alloc_rows: List[Tuple[str, str, float]] = list(base_alloc)

        # support reactions to MAIN beams: map beam_id -> list[(P,a)]
        add_to_main: Dict[str, List[Tuple[float, float]]] = {}

        for td in transfer_defs:
            Ltb = td.right_pos - td.left_pos
            if Ltb <= 1e-9:
                continue

            # TRANS beam point load at a_tb
            pts = [(td.P, td.a_tb)]
            chosen = choose_section_for_member(
                member_type="TRANS",
                span_dir=td.span_dir,
                span=Ltb,
                pos_or_fixed=td.fixed_coord,
                left=td.left_pos,
                right=td.right_pos,
                trib_width=0.0,
                q=cfg.q,
                point_loads=pts,
                sections=sections,
                mat=mat,
                setts=setts,
                mode=mode,
                rank_cap=rank_cap,
            )
            if chosen is None:
                return None

            # reactions include self-weight (w_udl = chosen.w_udl = w_g)
            Ra, Rb = reactions_simply_supported(Ltb, chosen.w_udl, td.P, td.a_tb)

            alloc_rows.append((td.load_id, td.left_support, Ra))
            alloc_rows.append((td.load_id, td.right_support, Rb))

            if td.left_support != "PERIM":
                add_to_main.setdefault(td.left_support, []).append((Ra, td.a_on_support))
            if td.right_support != "PERIM":
                add_to_main.setdefault(td.right_support, []).append((Rb, td.a_on_support))

            chosen = MemberCheck(
                beam_id=td.tb_id,
                member_type=chosen.member_type,
                direction=chosen.direction,
                pos_or_fixed=chosen.pos_or_fixed,
                left=chosen.left,
                right=chosen.right,
                span=chosen.span,
                section_rank=chosen.section_rank,
                section_name=chosen.section_name,
                w_g=chosen.w_g,
                w_udl=chosen.w_udl,
                n_point=chosen.n_point,
                Mmax=chosen.Mmax,
                Vmax=chosen.Vmax,
                dmax=chosen.dmax,
                util_M=chosen.util_M,
                util_V=chosen.util_V,
                util_d=chosen.util_d,
                util_max=chosen.util_max,
                ok=chosen.ok
            )
            member_checks.append(chosen)

        # 2) Add reactions to MAIN beams as point loads
        for b in main_beams:
            extra = add_to_main.get(b.beam_id, [])
            for P, a in extra:
                b.point_along.append((P, a))

        # 3) Design MAIN beams (per beam)
        for b in main_beams:
            chosen = choose_section_for_member(
                member_type="MAIN",
                span_dir=b.direction,
                span=b.span,
                pos_or_fixed=b.pos,
                left=None,
                right=None,
                trib_width=b.trib_width,
                q=cfg.q,
                point_loads=b.point_along,
                sections=sections,
                mat=mat,
                setts=setts,
                mode=mode,
                rank_cap=rank_cap,
            )
            if chosen is None:
                return None

            chosen = MemberCheck(
                beam_id=b.beam_id,
                member_type=chosen.member_type,
                direction=chosen.direction,
                pos_or_fixed=chosen.pos_or_fixed,
                left=chosen.left,
                right=chosen.right,
                span=chosen.span,
                section_rank=chosen.section_rank,
                section_name=chosen.section_name,
                w_g=chosen.w_g,
                w_udl=chosen.w_udl,
                n_point=chosen.n_point,
                Mmax=chosen.Mmax,
                Vmax=chosen.Vmax,
                dmax=chosen.dmax,
                util_M=chosen.util_M,
                util_V=chosen.util_V,
                util_d=chosen.util_d,
                util_max=chosen.util_max,
                ok=chosen.ok
            )
            member_checks.append(chosen)

        # aggregate
        total_weight = 0.0
        max_rank_used = 0
        Mmax = Vmax = dmax = 0.0
        util_max = 0.0
        worst_id = ""
        worst_u = -1.0

        # sum weights by member id => need member span; for MAIN it is in chosen.span, for TRANS too.
        for mc in member_checks:
            total_weight += mc.w_g * mc.span
            max_rank_used = max(max_rank_used, mc.section_rank)
            Mmax = max(Mmax, mc.Mmax)
            Vmax = max(Vmax, mc.Vmax)
            dmax = max(dmax, mc.dmax)
            util_max = max(util_max, mc.util_max)
            if mc.util_max > worst_u:
                worst_u = mc.util_max
                worst_id = mc.beam_id

        ok = util_max <= 1.0
        ng_reason = "" if ok else "Some member failed (should not happen if selection is OK-only)."

        return Solution(
            mode=mode,
            direction=direction,
            pitch=pitch,
            total_weight=total_weight,
            max_rank_used=max_rank_used,
            Mmax=Mmax,
            Vmax=Vmax,
            dmax=dmax,
            util_max=util_max,
            ok=ok,
            ng_reason=ng_reason,
            member_checks=sorted(member_checks, key=lambda x: (x.member_type, x.beam_id)),
            main_geoms=main_beams,
            transfer_defs=transfer_defs,
            allocation_rows=alloc_rows,
            worst_member_id=worst_id
        )

    if mode == "A":
        sol = attempt(rank_cap=None)
        if sol is None:
            return Solution(mode=mode, direction=direction, pitch=pitch, total_weight=math.inf, max_rank_used=10**9,
                            Mmax=0.0, Vmax=0.0, dmax=0.0, util_max=math.inf, ok=False, ng_reason="No feasible section set.",
                            member_checks=[], main_geoms=[], transfer_defs=[], allocation_rows=[], worst_member_id="")
        return sol

    # Mode B: scan rank cap
    for R in range(1, max_rank + 1):
        sol = attempt(rank_cap=R)
        if sol is not None:
            return sol

    return Solution(mode=mode, direction=direction, pitch=pitch, total_weight=math.inf, max_rank_used=10**9,
                    Mmax=0.0, Vmax=0.0, dmax=0.0, util_max=math.inf, ok=False, ng_reason="No feasible rank cap.",
                    member_checks=[], main_geoms=[], transfer_defs=[], allocation_rows=[], worst_member_id="")


# -----------------------------
# Optimizer (direction × pitch)
# -----------------------------
@dataclass(frozen=True)
class CandidateRow:
    mode: str
    direction: str
    pitch: float
    n_main: int
    n_trans: int
    max_rank_used: int
    total_weight: float
    Mmax: float
    Vmax: float
    dmax_mm: float
    util_max: float
    ok: bool
    ng_reason: str


def optimize(
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    sections: List[Section],
) -> Tuple[List[CandidateRow], Optional[Solution], Optional[Solution]]:
    pitches = make_pitch_candidates(cfg)

    directions: List[str] = []
    if cfg.enable_x:
        directions.append("X")
    if cfg.enable_y:
        directions.append("Y")
    if not directions:
        raise ValueError("Both X and Y directions are disabled.")

    s_axis = short_side_axis(cfg.Lx, cfg.Ly)

    rows: List[CandidateRow] = []
    best_A: Optional[Solution] = None
    best_B: Optional[Solution] = None

    # progress
    total_cases = len(directions) * len(pitches) * 2  # A/B
    done = 0
    ok_cnt = 0

    for d in directions:
        pdir = pitch_direction_of(d)
        for pitch in pitches:
            if pdir == s_axis and pitch > cfg.short_pitch_limit + 1e-9:
                continue

            for mode in ("A", "B"):
                done += 1
                sol = solve_layout_one_mode(cfg, mat, setts, sections, d, pitch, mode)

                n_main = len([m for m in sol.member_checks if m.member_type == "MAIN"])
                n_trans = len([m for m in sol.member_checks if m.member_type == "TRANS"])

                row = CandidateRow(
                    mode=mode,
                    direction=d,
                    pitch=pitch,
                    n_main=n_main,
                    n_trans=n_trans,
                    max_rank_used=sol.max_rank_used if sol.ok else 0,
                    total_weight=sol.total_weight,
                    Mmax=sol.Mmax,
                    Vmax=sol.Vmax,
                    dmax_mm=sol.dmax * 1000.0,
                    util_max=sol.util_max,
                    ok=sol.ok,
                    ng_reason=sol.ng_reason
                )
                rows.append(row)

                if sol.ok:
                    ok_cnt += 1

                # Mode-A best: minimum total weight
                if mode == "A" and sol.ok:
                    if best_A is None or sol.total_weight < best_A.total_weight - 1e-12 or (
                        abs(sol.total_weight - best_A.total_weight) <= 1e-12 and sol.util_max < best_A.util_max
                    ):
                        best_A = sol
                        print(f"[Best-A updated] dir={d} pitch={pitch} W={sol.total_weight:.3f} kN maxR={sol.max_rank_used} util={sol.util_max:.3f}")

                # Mode-B best: minimize max rank, tie-break weight
                if mode == "B" and sol.ok:
                    if best_B is None or sol.max_rank_used < best_B.max_rank_used or (
                        sol.max_rank_used == best_B.max_rank_used and sol.total_weight < best_B.total_weight - 1e-12
                    ):
                        best_B = sol
                        print(f"[Best-B updated] dir={d} pitch={pitch} maxR={sol.max_rank_used} W={sol.total_weight:.3f} kN util={sol.util_max:.3f}")

                if done % max(1, total_cases // 12) == 0:
                    ng_rate = 1.0 - ok_cnt / max(done, 1)
                    print(f"[Progress] {done}/{total_cases} evaluated, OK={ok_cnt}, NG rate={ng_rate:.1%}")

    return rows, best_A, best_B


# -----------------------------
# Excel I/O (INPUT)
# -----------------------------
def read_cell(ws, addr: str):
    return ws[addr].value


def read_bool(ws, addr: str, default: bool = False) -> bool:
    v = read_cell(ws, addr)
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "y", "on"):
        return True
    if s in ("false", "0", "no", "n", "off"):
        return False
    return default


def read_float(ws, addr: str, name: str, default: Optional[float] = None) -> float:
    v = read_cell(ws, addr)
    if v is None or str(v).strip() == "":
        if default is None:
            raise ValueError(f"Missing required value: {name} at {addr}")
        return float(default)
    try:
        return float(v)
    except Exception as e:
        raise ValueError(f"Invalid number for {name} at {addr}: {v}") from e


def read_input_xlsx(path: str) -> Tuple[Config, Material, SolverSettings, List[Section]]:
    wb = load_workbook(path, data_only=True)
    if "INPUT" not in wb.sheetnames:
        raise ValueError('Sheet "INPUT" not found.')
    ws = wb["INPUT"]

    Lx = read_float(ws, "B2", "Lx")
    Ly = read_float(ws, "B3", "Ly")
    q = read_float(ws, "B5", "q")

    # point loads table A9:D... blank row terminates
    loads: List[PointLoad] = []
    r = 9
    while True:
        load_id = ws[f"A{r}"].value
        if load_id is None or str(load_id).strip() == "":
            break
        P = ws[f"B{r}"].value
        x = ws[f"C{r}"].value
        y = ws[f"D{r}"].value
        try:
            loads.append(PointLoad(load_id=str(load_id).strip(), P=float(P), x=float(x), y=float(y)))
        except Exception as e:
            raise ValueError(f"Invalid point load row {r}: {load_id},{P},{x},{y}") from e
        r += 1

    enable_x = read_bool(ws, "B14", True)
    enable_y = read_bool(ws, "B15", True)
    edge_beams = read_bool(ws, "B16", True)
    load_rule = int(read_float(ws, "B17", "load_rule", 2))

    pitch_start = read_float(ws, "B20", "pitch_start", 1.5)
    pitch_end = read_float(ws, "B21", "pitch_end", 3.0)
    pitch_step = read_float(ws, "B22", "pitch_step", 0.5)
    short_pitch_limit = read_float(ws, "B23", "short_pitch_limit", 3.0)

    # optional pitch list at F20:F...
    pitch_list: List[float] = []
    rr = 20
    while True:
        v = ws[f"F{rr}"].value
        if v is None or str(v).strip() == "":
            break
        try:
            pitch_list.append(float(v))
        except Exception as e:
            raise ValueError(f"Invalid pitch_list at F{rr}: {v}") from e
        rr += 1

    # snap tolerance (optional) at B24, default 1e-3
    snap_tol = float(ws["B24"].value) if ws["B24"].value not in (None, "") else 1e-3

    # material
    E_val = read_float(ws, "B26", "E", 205000.0)
    E_unit = ws["C26"].value
    E_unit_s = str(E_unit).strip().lower() if E_unit is not None else "n/mm2"

    if E_unit_s in ("n/mm2", "n/mm^2", "mpa"):
        E_kN_m2 = nmm2_to_kN_m2(E_val)
    elif E_unit_s in ("kn/m2", "kn/m^2"):
        E_kN_m2 = float(E_val)
    else:
        raise ValueError('E_unit must be "N/mm2" or "kN/m2".')

    fb = read_float(ws, "B27", "fb", 165.0)
    fv = read_float(ws, "B28", "fv", 95.0)
    defl_lim = read_float(ws, "B29", "deflection_limit", 360.0)

    mat = Material(
        E_kN_m2=E_kN_m2,
        fb_kN_m2=nmm2_to_kN_m2(fb),
        fv_kN_m2=nmm2_to_kN_m2(fv),
        deflection_limit=defl_lim
    )

    tol = read_float(ws, "B31", "tol", 1e-6)
    max_iter = int(read_float(ws, "B32", "max_iter", 5))
    n_div = int(read_float(ws, "B34", "n_div", 2000))
    setts = SolverSettings(tol=tol, max_iter=max_iter, n_div=n_div)

    # sections table (header allowed)
    sections: List[Section] = []
    rr = 41
    while True:
        rk = ws[f"A{rr}"].value
        name = ws[f"B{rr}"].value
        w_g = ws[f"C{rr}"].value

        # blank row terminates
        if rk is None and (name is None or str(name).strip() == ""):
            break

        # skip header rows like "Rank"
        try:
            rk_f = float(rk)
        except Exception:
            rr += 1
            continue

        if name is None or str(name).strip() == "":
            raise ValueError(f"SectionName missing at row {rr}")
        if w_g is None or str(w_g).strip() == "":
            raise ValueError(f"w_g missing at row {rr}")

        def get_opt(col: str) -> Optional[float]:
            v = ws[f"{col}{rr}"].value
            if v is None or str(v).strip() == "":
                return None
            return float(v)

        sections.append(Section(
            rank=int(rk_f),
            name=str(name).strip(),
            w_g=float(w_g),
            h=get_opt("D"),
            b=get_opt("E"),
            tw=get_opt("F"),
            tf=get_opt("G"),
            A_mm2=get_opt("H"),
            Z_mm3=get_opt("I"),
            I_mm4=get_opt("J"),
            Av_mm2=get_opt("K"),
        ))
        rr += 1

    sections = sorted(sections, key=lambda s: s.rank)
    if not sections:
        raise ValueError("No sections provided.")

    cfg = Config(
        Lx=float(Lx),
        Ly=float(Ly),
        q=float(q),
        loads=loads,
        enable_x=enable_x,
        enable_y=enable_y,
        edge_beams=edge_beams,
        load_rule=load_rule,
        pitch_start=pitch_start,
        pitch_end=pitch_end,
        pitch_step=pitch_step,
        pitch_list=pitch_list,
        short_pitch_limit=float(short_pitch_limit),
        snap_tol=float(snap_tol),
    )
    return cfg, mat, setts, sections


# -----------------------------
# Excel output helpers
# -----------------------------
def set_col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def write_header(ws, row: int, col: int, text: str) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor="DDDDDD")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# -----------------------------
# RESULT sheet
# -----------------------------
def write_solution_summary(ws, row0: int, title: str, sol: Optional[Solution]) -> int:
    ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
    headers = ["Dir", "Pitch[m]", "MaxRankUsed", "TotalWeight[kN]", "Mmax[kN*m]", "Vmax[kN]", "dmax[mm]", "utilMax", "OK/NG"]
    for j, h in enumerate(headers, start=1):
        write_header(ws, row0 + 1, j, h)
    if sol is None or not sol.ok:
        ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
        return row0 + 4
    vals = [
        sol.direction,
        sol.pitch,
        sol.max_rank_used,
        round(sol.total_weight, 6),
        round(sol.Mmax, 6),
        round(sol.Vmax, 6),
        round(sol.dmax * 1000.0, 6),
        round(sol.util_max, 6),
        "OK",
    ]
    for j, v in enumerate(vals, start=1):
        ws.cell(row=row0 + 2, column=j, value=v)
    return row0 + 4


def write_candidates_table(ws, row0: int, rows: List[CandidateRow]) -> int:
    ws.cell(row=row0, column=1, value="All candidates (Mode-A and Mode-B)").font = Font(bold=True)
    row0 += 1
    headers = ["Mode", "Dir", "Pitch[m]", "N_MAIN", "N_TRANS", "MaxRankUsed", "TotalWeight[kN]", "Mmax[kN*m]", "Vmax[kN]", "dmax[mm]", "utilMax", "OK/NG"]
    for j, h in enumerate(headers, start=1):
        write_header(ws, row0, j, h)
    row0 += 1

    rows_sorted = sorted(rows, key=lambda r: (r.mode, not r.ok, r.max_rank_used if r.ok else 10**9, r.total_weight))
    for r in rows_sorted:
        out = [
            r.mode, r.direction, r.pitch, r.n_main, r.n_trans,
            r.max_rank_used if r.ok else "",
            round(r.total_weight, 6) if math.isfinite(r.total_weight) else "",
            round(r.Mmax, 6),
            round(r.Vmax, 6),
            round(r.dmax_mm, 6),
            round(r.util_max, 6) if math.isfinite(r.util_max) else "",
            "OK" if r.ok else f"NG: {r.ng_reason}",
        ]
        for j, v in enumerate(out, start=1):
            ws.cell(row=row0, column=j, value=v)
        row0 += 1

    return row0 + 2


def write_member_checks(ws, row0: int, title: str, sol: Optional[Solution]) -> int:
    ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
    headers = ["BeamID", "Type", "Dir", "Pos/Fixed[m]", "Left[m]", "Right[m]", "Span[m]",
               "Rank", "Section", "w_g[kN/m]", "w_udl[kN/m]", "Npoint",
               "Mmax[kN*m]", "Vmax[kN]", "dmax[mm]",
               "util_M", "util_V", "util_d", "util_max", "OK/NG"]
    for j, h in enumerate(headers, start=1):
        write_header(ws, row0 + 1, j, h)
    if sol is None or not sol.ok:
        ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
        return row0 + 4

    r = row0 + 2
    # sort: MAIN then TRANS
    checks = sorted(sol.member_checks, key=lambda x: (0 if x.member_type == "MAIN" else 1, x.beam_id))
    for mc in checks:
        out = [
            mc.beam_id,
            mc.member_type,
            mc.direction,
            mc.pos_or_fixed if mc.pos_or_fixed is not None else "",
            mc.left if mc.left is not None else "",
            mc.right if mc.right is not None else "",
            round(mc.span, 6),
            mc.section_rank,
            mc.section_name,
            round(mc.w_g, 6),
            round(mc.w_udl, 6),
            mc.n_point,
            round(mc.Mmax, 6),
            round(mc.Vmax, 6),
            round(mc.dmax * 1000.0, 6),
            round(mc.util_M, 6),
            round(mc.util_V, 6),
            round(mc.util_d, 6),
            round(mc.util_max, 6),
            "OK" if mc.ok else "NG",
        ]
        for j, v in enumerate(out, start=1):
            ws.cell(row=r, column=j, value=v)
        r += 1
    return r + 1


def write_allocations(ws, row0: int, title: str, sol: Optional[Solution]) -> int:
    ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
    write_header(ws, row0 + 1, 1, "LoadID")
    write_header(ws, row0 + 1, 2, "SupportID")
    write_header(ws, row0 + 1, 3, "AllocatedP[kN]")
    if sol is None or not sol.ok:
        ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
        return row0 + 4

    r = row0 + 2
    for lid, sid, p in sol.allocation_rows:
        ws.cell(row=r, column=1, value=lid)
        ws.cell(row=r, column=2, value=sid)
        ws.cell(row=r, column=3, value=round(p, 6))
        r += 1
    return r + 1


def write_member_list(ws, row0: int, title: str, sol: Optional[Solution]) -> int:
    ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
    headers = ["MemberType", "Rank", "Section", "Count", "TotalLen[m]", "w_g[kN/m]", "Weight[kN]"]
    for j, h in enumerate(headers, start=1):
        write_header(ws, row0 + 1, j, h)
    if sol is None or not sol.ok:
        ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
        return row0 + 4

    # aggregate by (type, section)
    agg: Dict[Tuple[str, int, str, float], Tuple[int, float, float]] = {}
    for mc in sol.member_checks:
        key = (mc.member_type, mc.section_rank, mc.section_name, mc.w_g)
        cnt, Ltot, W = agg.get(key, (0, 0.0, 0.0))
        agg[key] = (cnt + 1, Ltot + mc.span, W + mc.w_g * mc.span)

    r = row0 + 2
    total_w = 0.0
    for (mtype, rk, name, wg), (cnt, Ltot, W) in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1], x[0][2])):
        ws.cell(row=r, column=1, value=mtype)
        ws.cell(row=r, column=2, value=rk)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=cnt)
        ws.cell(row=r, column=5, value=round(Ltot, 6))
        ws.cell(row=r, column=6, value=round(wg, 6))
        ws.cell(row=r, column=7, value=round(W, 6))
        total_w += W
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=7, value=round(total_w, 6)).font = Font(bold=True)
    return r + 2


def write_result_xlsx(
    in_path: str,
    out_path: str,
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    cand_rows: List[CandidateRow],
    best_A: Optional[Solution],
    best_B: Optional[Solution],
) -> None:
    wb = load_workbook(in_path)
    if "RESULT" in wb.sheetnames:
        del wb["RESULT"]
    ws = wb.create_sheet("RESULT")

    # column widths
    widths = [8, 8, 10, 10, 10, 18, 12, 12, 12, 12, 10, 28, 12, 12, 12, 10, 10, 10, 10, 10]
    for i, w in enumerate(widths, start=1):
        set_col_width(ws, i, w)

    r = 1
    ws.cell(row=r, column=1, value="Secondary Beam Optimization RESULT").font = Font(bold=True, size=14)
    r += 2
    ws.cell(row=r, column=1, value="Notes: Units are kN, m. E/fb/fv are converted from N/mm2 to kN/m2.").font = Font(italic=True)
    r += 2

    r = write_solution_summary(ws, r, "Mode-A best (total weight minimum)", best_A)
    r = write_solution_summary(ws, r, "Mode-B best (minimize maximum section rank, tie-break: weight)", best_B)

    r = write_candidates_table(ws, r, cand_rows)

    r = write_member_checks(ws, r, "Beam checks (per member) - Mode-A best", best_A)
    r = write_allocations(ws, r, "Load allocation to supports (includes TRANS reactions) - Mode-A best", best_A)
    r = write_member_list(ws, r, "Member list - Mode-A best", best_A)

    r += 1
    r = write_member_checks(ws, r, "Beam checks (per member) - Mode-B best", best_B)
    r = write_allocations(ws, r, "Load allocation to supports (includes TRANS reactions) - Mode-B best", best_B)
    r = write_member_list(ws, r, "Member list - Mode-B best", best_B)

    r += 1
    ws.cell(row=r, column=1, value="PLAN view is written to sheet 'LAYOUT' (cells only).").font = Font(italic=True)

    # layout sheet
    write_layout_sheet(wb, cfg, best_A, best_B)

    wb.save(out_path)


# -----------------------------
# LAYOUT (plan only)
# -----------------------------
def write_layout_sheet(wb, cfg: Config, best_A: Optional[Solution], best_B: Optional[Solution]) -> None:
    if "LAYOUT" in wb.sheetnames:
        del wb["LAYOUT"]
    ws = wb.create_sheet("LAYOUT")

    top = 1
    left = 1
    ws.cell(row=top, column=left, value="LAYOUT / PLAN VIEW (cells only)").font = Font(bold=True, size=14)

    top += 2
    if best_A and best_A.ok:
        top = draw_plan(ws, top, left, cfg, best_A, title="Mode-A best (total weight minimum)")
        top += 2
    if best_B and best_B.ok:
        top = draw_plan(ws, top, left, cfg, best_B, title="Mode-B best (minimize maximum section rank)")
        top += 2

    if (best_A is None or not best_A.ok) and (best_B is None or not best_B.ok):
        ws.cell(row=top, column=left, value="NO FEASIBLE SOLUTION").font = Font(bold=True)


def draw_plan(ws, top: int, left: int, cfg: Config, sol: Solution, title: str) -> int:
    """
    Plan-only sheet (no tables). Square cells are applied only to the grid area.
    - MAIN beams: light blue
    - TRANS beams: yellow
    - Worst member: red fill
    - Point load: red ●ID at actual position
    - Allocation marker: blue ▲ID on receiving support line
    """
    # grid size (m per cell)
    grid = 0.25
    nx = int(round(cfg.Lx / grid)) + 1
    ny = int(round(cfg.Ly / grid)) + 1

    # styles
    fill_main = PatternFill("solid", fgColor="DDEBF7")
    fill_trans = PatternFill("solid", fgColor="FFF2CC")
    fill_worst = PatternFill("solid", fgColor="FF0000")  # red
    fill_blank = PatternFill("solid", fgColor="FFFFFF")

    font_small = Font(size=9)
    font_bold = Font(bold=True)

    border_thick = Border(left=Side(style="thick"), right=Side(style="thick"),
                          top=Side(style="thick"), bottom=Side(style="thick"))
    border_grid = Border(left=Side(style="thin", color="D9D9D9"),
                         right=Side(style="thin", color="D9D9D9"),
                         top=Side(style="thin", color="D9D9D9"),
                         bottom=Side(style="thin", color="D9D9D9"))

    # coordinate mapping
    grid_left = left + 2
    grid_top = top + 2

    def x_to_c(xm: float) -> int:
        return grid_left + int(round(xm / grid))

    def y_to_r(ym: float) -> int:
        return grid_top + int(round(ym / grid))

    # Title lines
    ws.cell(row=top, column=left, value=title).font = Font(bold=True)
    ws.cell(row=top + 1, column=left, value=f"PLAN: {sol.direction}-dir, pitch={sol.pitch} m, totalW={sol.total_weight:.3f} kN, maxR={sol.max_rank_used}").font = font_small

    # Format square cells only for grid area (+ axis label margins)
    col_w = 2.0   # about square with default row height
    row_h = 12.0
    for c in range(grid_left, grid_left + nx):
        ws.column_dimensions[get_column_letter(c)].width = col_w
    for r in range(grid_top, grid_top + ny):
        ws.row_dimensions[r].height = row_h

    # clear and grid
    for r in range(grid_top, grid_top + ny):
        for c in range(grid_left, grid_left + nx):
            cell = ws.cell(row=r, column=c, value="")
            cell.fill = fill_blank
            cell.border = border_grid
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # thick boundary
    for c in range(grid_left, grid_left + nx):
        ws.cell(row=grid_top, column=c).border = border_thick
        ws.cell(row=grid_top + ny - 1, column=c).border = border_thick
    for r in range(grid_top, grid_top + ny):
        ws.cell(row=r, column=grid_left).border = border_thick
        ws.cell(row=r, column=grid_left + nx - 1).border = border_thick

    # axis labels (integer meters only)
    ws.cell(row=grid_top - 1, column=grid_left, value="y[m]→x[m]").font = font_small
    for xm in range(0, int(math.floor(cfg.Lx)) + 1):
        c = x_to_c(float(xm))
        if grid_left <= c < grid_left + nx:
            ws.cell(row=grid_top - 1, column=c, value=str(xm)).font = font_small

    for ym in range(0, int(math.floor(cfg.Ly)) + 1):
        r0 = y_to_r(float(ym))
        if grid_top <= r0 < grid_top + ny:
            ws.cell(row=r0, column=grid_left - 1, value=str(ym)).font = font_small

    # build lookup of checks by id
    chk = {m.beam_id: m for m in sol.member_checks}

    # Draw MAIN beams
    # MAIN beams stored as member_checks with member_type=MAIN => need their direction and pos
    for mc in sol.member_checks:
        if mc.member_type != "MAIN":
            continue
        pos = float(mc.pos_or_fixed) if mc.pos_or_fixed is not None else 0.0
        is_worst = (mc.beam_id == sol.worst_member_id)

        if sol.direction == "Y":
            # MAIN beams are vertical (span Y), placed at x=pos
            c = x_to_c(pos)
            for r in range(grid_top, grid_top + ny):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_worst if is_worst else fill_main
        else:
            # MAIN beams are horizontal (span X), placed at y=pos
            r0 = y_to_r(pos)
            for c in range(grid_left, grid_left + nx):
                cell = ws.cell(row=r0, column=c)
                cell.fill = fill_worst if is_worst else fill_main

        # label (merged cell to avoid vertical text)
        label = f"{mc.beam_id} u={mc.util_max:.2f}"
        if sol.direction == "Y":
            c = x_to_c(pos)
            c0 = max(grid_left, c - 1)
            c1 = min(grid_left + nx - 1, c + 1)
            ws.merge_cells(start_row=grid_top - 2, start_column=c0, end_row=grid_top - 2, end_column=c1)
            lc = ws.cell(row=grid_top - 2, column=c0, value=label)
        else:
            r0 = y_to_r(pos)
            # label to the left of the beam row
            lc = ws.cell(row=r0, column=grid_left - 2, value=label)
        lc.font = font_small
        lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Draw TRANS beams (span_dir is pitch axis; fixed_coord is perpendicular)
    for td in sol.transfer_defs:
        mc = chk.get(td.tb_id)
        if mc is None:
            continue
        is_worst = (td.tb_id == sol.worst_member_id)

        if td.span_dir == "X":
            # horizontal trans at y=fixed_coord, x from left_pos to right_pos
            r0 = y_to_r(td.fixed_coord)
            c0 = x_to_c(td.left_pos)
            c1 = x_to_c(td.right_pos)
            for c in range(min(c0, c1), max(c0, c1) + 1):
                cell = ws.cell(row=r0, column=c)
                cell.fill = fill_worst if is_worst else fill_trans
        else:
            # vertical trans at x=fixed_coord
            c0 = x_to_c(td.fixed_coord)
            r0 = y_to_r(td.left_pos)
            r1 = y_to_r(td.right_pos)
            for r in range(min(r0, r1), max(r0, r1) + 1):
                cell = ws.cell(row=r, column=c0)
                cell.fill = fill_worst if is_worst else fill_trans

        # label near mid span
        midx = 0.5 * (td.left_pos + td.right_pos) if td.span_dir == "X" else td.fixed_coord
        midy = td.fixed_coord if td.span_dir == "X" else 0.5 * (td.left_pos + td.right_pos)
        lc = ws.cell(row=y_to_r(midy), column=x_to_c(midx), value=f"{td.tb_id}")
        lc.font = Font(size=8, bold=True)
        lc.alignment = Alignment(horizontal="center", vertical="center")

    # Point loads and allocation markers
    # Point load marker ● at actual coordinate
    # Allocation marker ▲ on receiving supports (MAIN beam or PERIM)
    # We'll use allocation_rows + original load coordinates
    load_map = {pl.load_id: pl for pl in cfg.loads}

    # mark actual load
    for pl in cfg.loads:
        c = x_to_c(pl.x)
        r0 = y_to_r(pl.y)
        if grid_left <= c < grid_left + nx and grid_top <= r0 < grid_top + ny:
            ws.cell(row=r0, column=c, value="●").font = Font(color="FF0000", bold=True)
            ws.cell(row=r0, column=c + 1, value=pl.load_id).font = font_small

    # allocation markers
    for lid, sid, _p in sol.allocation_rows:
        if sid == "PERIM":
            continue
        pl = load_map.get(lid)
        if pl is None:
            continue
        # location on support: depends on sol.direction
        if sol.direction == "Y":
            # supports are vertical MAIN beams (at x=pos), marker at y=pl.y
            # find support pos
            sup_mc = chk.get(sid)
            if sup_mc and sup_mc.pos_or_fixed is not None:
                c = x_to_c(float(sup_mc.pos_or_fixed))
                r0 = y_to_r(pl.y)
                ws.cell(row=r0, column=c - 1, value="▲").font = Font(color="1F4E79", bold=True)
                ws.cell(row=r0, column=c, value=lid).font = font_small
        else:
            # supports are horizontal MAIN beams at y=pos, marker at x=pl.x
            sup_mc = chk.get(sid)
            if sup_mc and sup_mc.pos_or_fixed is not None:
                r0 = y_to_r(float(sup_mc.pos_or_fixed))
                c = x_to_c(pl.x)
                ws.cell(row=r0, column=c, value="▲").font = Font(color="1F4E79", bold=True)
                ws.cell(row=r0, column=c + 1, value=lid).font = font_small

    # Legend (below grid)
    legend_r = grid_top + ny + 1
    ws.cell(row=legend_r, column=left, value="Legend:").font = Font(bold=True)
    ws.cell(row=legend_r + 1, column=left, value="MAIN beam: light-blue fill").font = font_small
    ws.cell(row=legend_r + 2, column=left, value="TRANS beam: yellow fill").font = font_small
    ws.cell(row=legend_r + 3, column=left, value="Worst utilization member: RED fill").font = font_small
    ws.cell(row=legend_r + 4, column=left, value="Point load: red ●ID, Allocation marker: blue ▲ID on receiving support").font = font_small

    return legend_r + 6


# -----------------------------
# Main
# -----------------------------
def main(argv: List[str]) -> int:
    if len(argv) < 2:
        print("Usage: python beam_optimizer_b6.py input.xlsx output.xlsx")
        return 2

    in_path = argv[1]
    out_path = argv[2] if len(argv) >= 3 else (in_path.replace(".xlsx", "") + "_out.xlsx")

    cfg, mat, setts, sections = read_input_xlsx(in_path)

    # basic validation
    ensure_positive("Lx", cfg.Lx)
    ensure_positive("Ly", cfg.Ly)
    ensure_positive("q", cfg.q)
    ensure_positive("E", mat.E_kN_m2)
    ensure_positive("fb", mat.fb_kN_m2)
    ensure_positive("fv", mat.fv_kN_m2)
    ensure_positive("deflection_limit", mat.deflection_limit)

    # run
    cand_rows, best_A, best_B = optimize(cfg, mat, setts, sections)

    # write excel
    write_result_xlsx(in_path, out_path, cfg, mat, setts, cand_rows, best_A, best_B)

    print(f"Done. Output written to: {out_path}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main(sys.argv))
    except PermissionError as e:
        print(f"[ERROR] Permission denied. Close the target Excel file and retry: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)
