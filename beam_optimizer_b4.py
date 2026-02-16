#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
README / 実行方法
-----------------
- 前提: Python 3.x, openpyxl がインストール済み
- 実行例:
    python beam_optimizer_b1_flex.py input.xlsx output.xlsx

- 入力:
    Excelファイル（.xlsx）内のシート "INPUT" を参照
- 出力:
    指定した output.xlsx に、シート "RESULT" を生成して書き込み

設計モデル（固定仮定）
---------------------
- 小梁: 両端ピンの単純支持梁（単純支持梁）
- 荷重: 等分布荷重（q*支配幅 + 自重w_g） + 複数集中荷重（配分ルールで小梁へ割当）
- 断面照査:
    曲げ:   Mmax / Z <= fb
    せん断: Vmax / Av <= fv  （Av無い場合は近似。限界はコメント参照）
    たわみ: δmax <= L / deflection_limit

- たわみ算定: 曲率 k(x)=M(x)/(E*I) を数値積分（分割数 n_div）し、
  y(0)=0, y(L)=0 を満たすように初期回転 θ0 を調整する。

注意（自重反復枠）
-----------------
断面固定なら自重w_gは固定なので通常1回で収束します。
将来、断面更新ロジックを追加する拡張を想定して反復枠を残しています。

Excel INPUTの“行増減”対応（今回の修正点）
---------------------------------------
旧版は固定セル番地（例: B2,B3,A41...）で読んでいたため、
行挿入や表の位置変更で「Rank」「LoadID」などのヘッダ行を数値として読みに行き、例外になりました。

本版は次を採用して、行が増減しても動くようにしています（テンプレのレイアウトは自由度UP）:
- 単一値: 「ラベルセル（例: 'Lx [m]'）の右隣」を読む（ラベル検索）
- テーブル: ヘッダ行（例: 'Rank','SectionName','w_g [kN/m]'）を検索して、その直下を読む
- テーブル終端: “連続空行”で終端（途中に空行があっても継続可能）

※ラベル文字列／ヘッダ名は、テンプレに合わせて“なるべくそのまま”入れてください。
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from typing import List, Optional, Tuple, Dict

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# -----------------------------
# Data models
# -----------------------------
@dataclass(frozen=True)
class PointLoad:
    load_id: str
    P: float     # kN
    x: float     # m (0..Lx)
    y: float     # m (0..Ly)


@dataclass(frozen=True)
class Section:
    rank: int
    name: str
    w_g: float      # kN/m
    h: Optional[float] = None   # mm
    b: Optional[float] = None   # mm
    tw: Optional[float] = None  # mm
    tf: Optional[float] = None  # mm
    A_mm2: Optional[float] = None
    Z_mm3: Optional[float] = None
    I_mm4: Optional[float] = None
    Av_mm2: Optional[float] = None


@dataclass(frozen=True)
class Material:
    E_kN_m2: float     # kN/m^2
    fb_kN_m2: float    # kN/m^2
    fv_kN_m2: float    # kN/m^2
    deflection_limit: float


@dataclass(frozen=True)
class SolverSettings:
    tol: float
    max_iter: int
    n_div: int


@dataclass(frozen=True)
class Config:
    Lx: float
    Ly: float
    q: float
    loads: List[PointLoad]

    enable_x: bool
    enable_y: bool
    edge_beams: bool
    perimeter_primary: bool  # True: perimeter is primary beams (no secondary at edges)
    load_rule: int  # 1 or 2

    pitch_start: float
    pitch_end: float
    pitch_step: float
    pitch_list: List[float]  # if not empty, overrides start/end/step

    short_pitch_limit: float  # fixed display value (3.0)


@dataclass
class Beam:
    beam_id: str
    direction: str  # 'X' or 'Y'
    pos: float      # position along pitch direction (m)
    span: float     # m
    trib_width: float  # m
    point_along: List[Tuple[float, float]]  # (P[kN], a[m] from left)


@dataclass
class CandidateResult:
    direction: str
    pitch: float
    n_beams: int
    section_rank: int
    section_name: str

    w_total_max: float   # max(kN/m) among beams
    Mmax: float          # kN*m (max among beams)
    Vmax: float          # kN (max among beams)
    dmax: float          # m (max among beams)

    util_M: float
    util_V: float
    util_d: float
    util_max: float

    total_weight: float  # kN (sum w_g * span * n_beams)
    ok: bool
    ng_reason: str

    allocation_rows: List[Tuple[str, str, float]]  # (LoadID, BeamID, AllocP[kN])


# -----------------------------
# Utility: unit conversions
# -----------------------------
def nmm2_to_kN_m2(val_nmm2: float) -> float:
    # 1 N/mm2 = 1 MPa = 1e6 N/m2 = 1e3 kN/m2
    return float(val_nmm2) * 1000.0


def ensure_positive(name: str, v: float) -> None:
    if v is None or not isinstance(v, (int, float)) or v <= 0:
        raise ValueError(f"{name} must be positive. got={v}")


def to_float(v, *, name: str, where: str) -> float:
    try:
        return float(v)
    except Exception as e:
        raise ValueError(f"Invalid number for {name} at {where}: {v!r}") from e


def to_int(v, *, name: str, where: str) -> int:
    try:
        return int(float(v))
    except Exception as e:
        raise ValueError(f"Invalid integer for {name} at {where}: {v!r}") from e


def to_bool(v, default: bool = False) -> bool:
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "y", "on", "ok"):
        return True
    if s in ("false", "0", "no", "n", "off", "ng"):
        return False
    return default


# -----------------------------
# Excel helpers: label/table scanning (row insert tolerant)
# -----------------------------
def norm_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def find_first_cell(ws, patterns: List[str], *, max_row: int = 500, max_col: int = 30) -> Optional[Tuple[int, int, str]]:
    """
    Find first cell (row,col,text) whose normalized text matches ANY regex in patterns (case-insensitive).
    """
    regs = [re.compile(p, re.IGNORECASE) for p in patterns]
    for r in range(1, min(max_row, ws.max_row) + 1):
        for c in range(1, min(max_col, ws.max_column) + 1):
            t = norm_text(ws.cell(r, c).value)
            if not t:
                continue
            for rg in regs:
                if rg.search(t):
                    return r, c, t
    return None


def read_value_right_of_label(
    ws,
    label_patterns: List[str],
    *,
    value_offset_cols: int = 1,
    default=None,
    required: bool = True,
    cast=None,
    name: str = "value",
) :
    """
    Read the cell located at (label_cell.row, label_cell.col + value_offset_cols).
    """
    hit = find_first_cell(ws, label_patterns)
    if not hit:
        if required:
            raise ValueError(f"Label for {name} not found. patterns={label_patterns}")
        return default
    r, c, label = hit
    v = ws.cell(r, c + value_offset_cols).value
    if v is None or norm_text(v) == "":
        if required and default is None:
            raise ValueError(f"Missing value for {name} at row={r} (label={label!r})")
        v = default
    return cast(v) if (cast and v is not None) else v


def find_header_row(ws, required_headers: List[str], *, max_row: int = 600, max_col: int = 50) -> Optional[Tuple[int, Dict[str, int]]]:
    """
    Find a row that contains all required headers (case-insensitive, whitespace-insensitive).
    Returns (row_index, {header->col_index}).
    """
    req = [h.strip().lower() for h in required_headers]
    for r in range(1, min(max_row, ws.max_row) + 1):
        row_map: Dict[str, int] = {}
        for c in range(1, min(max_col, ws.max_column) + 1):
            t = norm_text(ws.cell(r, c).value).lower()
            if not t:
                continue
            # exact-ish match (allow extra unit text)
            for h in req:
                if h not in row_map and (t == h or t.startswith(h)):
                    row_map[h] = c
        if all(h in row_map for h in req):
            # map back using original header strings
            out = {required_headers[i]: row_map[req[i]] for i in range(len(req))}
            return r, out
    return None


def read_table_rows(
    ws,
    header_row: int,
    col_map: Dict[str, int],
    *,
    start_row_offset: int = 1,
    max_blank_rows: int = 3,
    max_rows: int = 2000
) -> List[Dict[str, object]]:
    """
    Read rows under header_row. Terminates after max_blank_rows consecutive blank key rows.
    The "key" is the first header in col_map order.
    """
    headers = list(col_map.keys())
    key_h = headers[0]
    key_col = col_map[key_h]
    rows: List[Dict[str, object]] = []

    blank = 0
    r = header_row + start_row_offset
    end_r = min(ws.max_row, r + max_rows)
    while r <= end_r:
        key_v = ws.cell(r, key_col).value
        if key_v is None or norm_text(key_v) == "":
            blank += 1
            if blank >= max_blank_rows:
                break
            r += 1
            continue
        blank = 0

        # skip if this is a repeated header row
        if isinstance(key_v, str) and key_v.strip().lower() == key_h.strip().lower():
            r += 1
            continue

        d: Dict[str, object] = {}
        for h, c in col_map.items():
            d[h] = ws.cell(r, c).value
        rows.append(d)
        r += 1
    return rows


# -----------------------------
# Section properties (approx)
# -----------------------------
def parse_h_section_dims(name: str) -> Optional[Tuple[float, float, float, float]]:
    """
    Parse "H-350x175x7x11" or similar (x, X, ×).
    Returns (h,b,tw,tf) in mm if parse succeeds.
    """
    s = name.replace("×", "x").replace("X", "x")
    m = re.search(r"(\d+(\.\d+)?)x(\d+(\.\d+)?)x(\d+(\.\d+)?)x(\d+(\.\d+)?)", s)
    if not m:
        return None
    h = float(m.group(1))
    b = float(m.group(3))
    tw = float(m.group(5))
    tf = float(m.group(7))
    return h, b, tw, tf


def approx_h_section_props_mm(h: float, b: float, tw: float, tf: float) -> Tuple[float, float, float, float]:
    """
    Fillet ignored. Strong-axis bending (Ixx, Zx).
    Returns (A_mm2, I_mm4, Z_mm3, Av_mm2).
    """
    A = 2.0 * b * tf + (h - 2.0 * tf) * tw

    # Ixx = 2*(b*tf^3/12 + b*tf*y^2) + tw*(h-2tf)^3/12
    I_flange = b * tf**3 / 12.0
    y = h / 2.0 - tf / 2.0
    I = 2.0 * (I_flange + b * tf * y**2) + tw * (h - 2.0 * tf)**3 / 12.0
    Z = I / (h / 2.0)

    Av = tw * (h - 2.0 * tf)  # web area approximation
    return A, I, Z, Av


def get_section_props_m(section: Section) -> Tuple[float, float, float, float]:
    """
    Returns (A[m2], Av[m2], Z[m3], I[m4]).
    If inputs missing, tries to parse from name and approximate.
    """
    h, b, tw, tf = section.h, section.b, section.tw, section.tf

    if any(v is None for v in (h, b, tw, tf)):
        dims = parse_h_section_dims(section.name)
        if dims:
            h, b, tw, tf = dims

    if any(v is None for v in (h, b, tw, tf)):
        raise ValueError(f"Section dims missing and cannot parse from name: {section.name}")

    A_mm2 = section.A_mm2
    Z_mm3 = section.Z_mm3
    I_mm4 = section.I_mm4
    Av_mm2 = section.Av_mm2

    if any(v is None for v in (A_mm2, Z_mm3, I_mm4, Av_mm2)):
        A2, I2, Z2, Av2 = approx_h_section_props_mm(float(h), float(b), float(tw), float(tf))
        if A_mm2 is None:
            A_mm2 = A2
        if I_mm4 is None:
            I_mm4 = I2
        if Z_mm3 is None:
            Z_mm3 = Z2
        if Av_mm2 is None:
            Av_mm2 = Av2

    # Convert: mm2->m2 (1e-6), mm3->m3 (1e-9), mm4->m4 (1e-12)
    A = float(A_mm2) * 1e-6
    Av = float(Av_mm2) * 1e-6
    Z = float(Z_mm3) * 1e-9
    I = float(I_mm4) * 1e-12

    ensure_positive(f"A({section.name})", A)
    ensure_positive(f"Av({section.name})", Av)
    ensure_positive(f"Z({section.name})", Z)
    ensure_positive(f"I({section.name})", I)
    return A, Av, Z, I


# -----------------------------
# Geometry / beam layout
# -----------------------------
def make_pitch_candidates(cfg: Config) -> List[float]:
    if cfg.pitch_list:
        vals = [float(v) for v in cfg.pitch_list if v and float(v) > 0]
        return sorted(set(vals))

    ensure_positive("pitch_start", cfg.pitch_start)
    ensure_positive("pitch_end", cfg.pitch_end)
    ensure_positive("pitch_step", cfg.pitch_step)
    if cfg.pitch_end < cfg.pitch_start:
        raise ValueError("pitch_end must be >= pitch_start")

    out: List[float] = []
    x = cfg.pitch_start
    while x <= cfg.pitch_end + 1e-12:
        out.append(round(x, 10))
        x += cfg.pitch_step
    return sorted(set(out))


def positions_along(width: float, pitch: float, edge_beams: bool) -> List[float]:
    """
    Generate secondary beam centerlines along the pitch direction (0..width).

    Interpretation of edge_beams:
      - edge_beams=True:
          Secondary beams also exist at both edges (0 and width).
          Positions include 0 and width.
      - edge_beams=False:
          Perimeter is supported by primary beams (given), so secondary beams are placed
          only inside (0,width). Positions exclude 0 and width and are set to multiples
          of pitch: pitch, 2*pitch, 3*pitch, ...

    Note:
      - If pitch >= width and edge_beams=False, the returned list can be empty.
        Such a candidate will be treated as NG (no secondary beams).
    """
    ensure_positive("width", width)
    ensure_positive("pitch", pitch)

    pos: List[float] = []
    if edge_beams:
        pos = [0.0]
        k = 1
        while k * pitch < width - 1e-9:
            pos.append(k * pitch)
            k += 1
        if abs(pos[-1] - width) > 1e-9:
            pos.append(width)
    else:
        k = 1
        while k * pitch < width - 1e-9:
            pos.append(k * pitch)
            k += 1

    return sorted(set(round(p, 10) for p in pos))


def tributary_widths(pos: List[float], width: float) -> List[float]:
    """
    Tributary width based on midpoints between adjacent supports.

    Supports are:
      - primary beams at the perimeter (0 and width) are always treated as supports
      - secondary beams in `pos`

    Therefore:
      - if a secondary beam exists at the edge (pos includes 0 or width), its tributary
        becomes 0.5*pitch-like at the edge side (consistent with the fixed rule).
      - if no secondary beam exists at the edge, the first/last secondary beam tributary
        starts/ends at the midpoint between the perimeter primary beam and that secondary beam.

    This also handles the last bay not equal to pitch.
    """
    n = len(pos)
    if n == 0:
        return []
    trib: List[float] = []
    for i, p in enumerate(pos):
        left = 0.5 * (0.0 + p) if i == 0 else 0.5 * (pos[i - 1] + p)
        right = 0.5 * (p + width) if i == n - 1 else 0.5 * (p + pos[i + 1])
        trib.append(right - left)
    return trib


def pitch_direction_of(direction: str) -> str:
    # X方向配置: beams span X, are arrayed along Y => pitch_dir='Y'
    # Y方向配置: beams span Y, are arrayed along X => pitch_dir='X'
    return "Y" if direction == "X" else "X"


def short_side_axis(Lx: float, Ly: float) -> str:
    return "X" if Lx <= Ly else "Y"


# -----------------------------
# Load distribution (Rule-1 / Rule-2)
# -----------------------------
def allocate_point_loads_to_beams(
    beams: List[Beam],
    direction: str,
    loads: List[PointLoad],
    Lx: float,
    Ly: float,
    rule: int
) -> List[Tuple[str, str, float]]:
    """
    Fill beam.point_along and return allocation table rows (LoadID, BeamID, AllocP).
    """
    if rule not in (1, 2):
        raise ValueError("load_rule must be 1 or 2")

    if direction == "X":
        coord_pitch = "y"
        coord_span = "x"
        span_L = Lx
        pitch_positions = [b.pos for b in beams]
    else:
        coord_pitch = "x"
        coord_span = "y"
        span_L = Ly
        pitch_positions = [b.pos for b in beams]

    alloc_rows: List[Tuple[str, str, float]] = []

    def add_alloc(load_id: str, beam: Beam, P: float, a: float) -> None:
        beam.point_along.append((P, a))
        alloc_rows.append((load_id, beam.beam_id, P))

    for pl in loads:
        if not (0.0 <= pl.x <= Lx) or not (0.0 <= pl.y <= Ly):
            raise ValueError(f"Point load out of range: {pl}")

        c = pl.y if coord_pitch == "y" else pl.x
        a = pl.x if coord_span == "x" else pl.y
        if not (0.0 <= a <= span_L):
            raise ValueError(f"Point load along-span out of range: {pl}")

        if rule == 1:
            idx = min(range(len(beams)), key=lambda i: abs(pitch_positions[i] - c))
            add_alloc(pl.load_id, beams[idx], pl.P, a)
            continue

        # Rule-2
        if c <= pitch_positions[0] + 1e-9:
            add_alloc(pl.load_id, beams[0], pl.P, a)
            continue
        if c >= pitch_positions[-1] - 1e-9:
            add_alloc(pl.load_id, beams[-1], pl.P, a)
            continue

        for i in range(len(pitch_positions) - 1):
            p0 = pitch_positions[i]
            p1 = pitch_positions[i + 1]
            if p0 - 1e-9 <= c <= p1 + 1e-9:
                if abs(c - p0) < 1e-9:
                    add_alloc(pl.load_id, beams[i], pl.P, a)
                elif abs(c - p1) < 1e-9:
                    add_alloc(pl.load_id, beams[i + 1], pl.P, a)
                else:
                    Lseg = p1 - p0
                    if Lseg <= 0:
                        raise ValueError("Invalid beam positions (non-increasing).")
                    P0 = pl.P * (p1 - c) / Lseg
                    P1 = pl.P * (c - p0) / Lseg
                    add_alloc(pl.load_id, beams[i], P0, a)
                    add_alloc(pl.load_id, beams[i + 1], P1, a)
                break

    return alloc_rows


def build_beams(cfg: Config, direction: str, pitch: float) -> Tuple[List[Beam], List[Tuple[str, str, float]]]:
    if direction == "X":
        span = cfg.Lx
        width = cfg.Ly
    else:
        span = cfg.Ly
        width = cfg.Lx

    pos = positions_along(width, pitch, cfg.edge_beams)
    trib = tributary_widths(pos, width)

    if not pos:
        return [], []

    beams: List[Beam] = []
    for i, (p, tw) in enumerate(zip(pos, trib), start=1):
        beam_id = f"{direction}{i:02d}"
        beams.append(Beam(beam_id=beam_id, direction=direction, pos=p, span=span, trib_width=tw, point_along=[]))

    alloc_rows = allocate_point_loads_to_beams(
        beams=beams,
        direction=direction,
        loads=cfg.loads,
        Lx=cfg.Lx,
        Ly=cfg.Ly,
        rule=cfg.load_rule
    )
    return beams, alloc_rows


# -----------------------------
# Beam analysis (numerical)
# -----------------------------
def analyze_simply_supported(
    L: float,
    w_udl: float,
    point_loads: List[Tuple[float, float]],
    E: float,
    I: float,
    n_div: int
) -> Tuple[float, float, float]:
    """
    Returns (Mmax[kN*m], Vmax[kN], dmax[m]) for a simply-supported beam.
    Loads:
      - UDL w_udl [kN/m] along full span
      - point_loads: list of (P[kN], a[m] from left)
    """
    ensure_positive("L", L)
    if n_div < 50:
        raise ValueError("n_div too small (>=50 recommended)")

    pls = [(float(P), float(a)) for P, a in point_loads]
    for P, a in pls:
        if not (0.0 <= a <= L):
            raise ValueError(f"Point load position out of range: a={a}, L={L}")
        if P < 0:
            raise ValueError("Negative point load not supported in this template.")
    pls.sort(key=lambda x: x[1])

    # reactions
    Ra = w_udl * L / 2.0
    Rb = w_udl * L / 2.0
    for P, a in pls:
        Ra += P * (L - a) / L
        Rb += P * a / L

    # discretize
    n = n_div + 1
    dx = L / n_div
    xs = [i * dx for i in range(n)]

    V = [Ra - w_udl * x for x in xs]
    M = [Ra * x - w_udl * x * x / 2.0 for x in xs]

    # apply point loads
    for P, a in pls:
        for i, x in enumerate(xs):
            if x >= a - 1e-12:
                V[i] -= P
                M[i] -= P * (x - a)

    Mmax = max(abs(v) for v in M)
    Vmax = max(abs(v) for v in V)

    # curvature k = M/(E*I)
    k = [m / (E * I) for m in M]

    # integrate curvature -> slope (trapezoid)
    K1 = [0.0] * n
    for i in range(1, n):
        K1[i] = K1[i - 1] + 0.5 * (k[i - 1] + k[i]) * dx

    # integrate K1
    intK1 = [0.0] * n
    for i in range(1, n):
        intK1[i] = intK1[i - 1] + 0.5 * (K1[i - 1] + K1[i]) * dx

    # enforce y(L)=0
    theta0 = -intK1[-1] / L
    theta = [theta0 + v for v in K1]

    # integrate slope -> deflection (y(0)=0)
    y = [0.0] * n
    for i in range(1, n):
        y[i] = y[i - 1] + 0.5 * (theta[i - 1] + theta[i]) * dx

    dmax = max(abs(v) for v in y)
    return Mmax, Vmax, dmax


# -----------------------------
# Candidate evaluation / optimizer
# -----------------------------
def evaluate_candidate(
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    direction: str,
    pitch: float,
    section: Section
) -> CandidateResult:
    beams, alloc_rows = build_beams(cfg, direction, pitch)

    if not beams:
        # No secondary beams created (e.g., pitch >= width when edge beams are disabled).
        return CandidateResult(
            direction=direction,
            pitch=pitch,
            n_beams=0,
            section_rank=section.rank,
            section_name=section.name,
            w_total_max=0.0,
            Mmax=0.0,
            Vmax=0.0,
            dmax=0.0,
            util_M=float("inf"),
            util_V=float("inf"),
            util_d=float("inf"),
            util_max=float("inf"),
            total_weight=0.0,
            ok=False,
            ng_reason="no secondary beams generated (check pitch and edge setting)",
            allocation_rows=[],
        )

    # section props
    _, Av, Z, I = get_section_props_m(section)

    prev_key = None
    ng_reason = ""

    Mmax_all = Vmax_all = dmax_all = w_total_max = 0.0
    utilM_all = utilV_all = utild_all = 0.0

    for _it in range(1, setts.max_iter + 1):
        Mmax_all = Vmax_all = dmax_all = w_total_max = 0.0
        utilM_all = utilV_all = utild_all = 0.0

        for b in beams:
            w_q = cfg.q * b.trib_width
            w_total = w_q + section.w_g
            w_total_max = max(w_total_max, w_total)

            Mmax, Vmax, dmax = analyze_simply_supported(
                L=b.span,
                w_udl=w_total,
                point_loads=b.point_along,
                E=mat.E_kN_m2,
                I=I,
                n_div=setts.n_div
            )

            Mmax_all = max(Mmax_all, Mmax)
            Vmax_all = max(Vmax_all, Vmax)
            dmax_all = max(dmax_all, dmax)

            utilM = (Mmax / Z) / mat.fb_kN_m2
            utilV = (Vmax / Av) / mat.fv_kN_m2
            utilD = dmax / (b.span / mat.deflection_limit)

            utilM_all = max(utilM_all, utilM)
            utilV_all = max(utilV_all, utilV)
            utild_all = max(utild_all, utilD)

        key = (Mmax_all, dmax_all)
        if prev_key is not None:
            dM = abs(key[0] - prev_key[0]) / max(abs(prev_key[0]), 1e-9)
            dd = abs(key[1] - prev_key[1]) / max(abs(prev_key[1]), 1e-12)
            if max(dM, dd) < setts.tol:
                break
        prev_key = key

    if setts.max_iter > 0 and prev_key is None:
        ng_reason = "analysis not executed"

    util_max = max(utilM_all, utilV_all, utild_all)
    ok = util_max <= 1.0

    if not ok:
        reasons = []
        if utilM_all > 1.0:
            reasons.append(f"bending NG (util={utilM_all:.3f})")
        if utilV_all > 1.0:
            reasons.append(f"shear NG (util={utilV_all:.3f})")
        if utild_all > 1.0:
            reasons.append(f"deflection NG (util={utild_all:.3f})")
        ng_reason = "; ".join(reasons)

    total_weight = section.w_g * beams[0].span * len(beams) if beams else 0.0

    return CandidateResult(
        direction=direction,
        pitch=pitch,
        n_beams=len(beams),
        section_rank=section.rank,
        section_name=section.name,
        w_total_max=w_total_max,
        Mmax=Mmax_all,
        Vmax=Vmax_all,
        dmax=dmax_all,
        util_M=utilM_all,
        util_V=utilV_all,
        util_d=utild_all,
        util_max=util_max,
        total_weight=total_weight,
        ok=ok,
        ng_reason=ng_reason,
        allocation_rows=alloc_rows
    )


def optimize(cfg: Config, mat: Material, setts: SolverSettings, sections: List[Section]) -> Tuple[List[CandidateResult], Optional[CandidateResult], Optional[CandidateResult]]:
    pitches = make_pitch_candidates(cfg)
    s_axis = short_side_axis(cfg.Lx, cfg.Ly)

    directions: List[str] = []
    if cfg.enable_x:
        directions.append("X")
    if cfg.enable_y:
        directions.append("Y")
    if not directions:
        raise ValueError("Both X and Y directions are disabled.")

    results: List[CandidateResult] = []
    total_cases = len(directions) * len(pitches) * len(sections)
    done = 0
    ok_count = 0

    best_A: Optional[CandidateResult] = None
    best_B: Optional[CandidateResult] = None

    for d in directions:
        pdir = pitch_direction_of(d)
        for pitch in pitches:
            if pdir == s_axis and pitch > cfg.short_pitch_limit + 1e-9:
                continue

            for sec in sections:
                done += 1
                res = evaluate_candidate(cfg, mat, setts, d, pitch, sec)
                results.append(res)
                if res.ok:
                    ok_count += 1

                # Mode-A: min total_weight subject to OK
                if res.ok:
                    if best_A is None or (res.total_weight < best_A.total_weight - 1e-12) or (
                        abs(res.total_weight - best_A.total_weight) <= 1e-12 and res.util_max < best_A.util_max
                    ):
                        best_A = res
                        print(f"[Best-A updated] dir={res.direction} pitch={res.pitch} sec={res.section_name} W={res.total_weight:.3f} kN util={res.util_max:.3f}")

                # Mode-B: min section rank, tie-break min weight
                if res.ok:
                    if best_B is None or (res.section_rank < best_B.section_rank) or (
                        res.section_rank == best_B.section_rank and res.total_weight < best_B.total_weight - 1e-12
                    ):
                        best_B = res
                        print(f"[Best-B updated] dir={res.direction} pitch={res.pitch} rank={res.section_rank} sec={res.section_name} W={res.total_weight:.3f} kN")

                if done % max(1, total_cases // 20) == 0:
                    ng_rate = 1.0 - (ok_count / max(done, 1))
                    print(f"[Progress] {done}/{total_cases} evaluated, OK={ok_count}, NG rate={ng_rate:.1%}")

    return results, best_A, best_B


# -----------------------------
# Excel I/O (flex)
# -----------------------------
def read_pitch_list(ws) -> List[float]:
    """
    Optional pitch list reader.
    Strategy:
      1) Find a header cell like 'Pitch [m]' or 'Pitch' and read numbers below.
         If the column below is empty but adjacent column has numbers, use adjacent.
      2) Fallback: find a label like 'Optional pitch list' and read near-right column values below.
    """
    # (1) header 'Pitch [m]' / 'Pitch'
    hit = find_first_cell(ws, [r"^\s*pitch\s*\[\s*m\s*\]\s*$", r"^\s*pitch\s*$", r"ピッチ\s*\[\s*m\s*\]"])
    if hit:
        hr, hc, _ = hit
        cols_try = [hc, hc + 1]
        best_col = None
        best_vals: List[float] = []
        for col in cols_try:
            vals: List[float] = []
            blank = 0
            for r in range(hr + 1, min(ws.max_row, hr + 200) + 1):
                v = ws.cell(r, col).value
                if v is None or norm_text(v) == "":
                    blank += 1
                    if blank >= 3:
                        break
                    continue
                blank = 0
                try:
                    vals.append(float(v))
                except Exception:
                    break
            if len(vals) > len(best_vals):
                best_vals = vals
                best_col = col
        if best_vals:
            return [v for v in best_vals if v > 0]

    # (2) label 'Optional pitch list'
    hit2 = find_first_cell(ws, [r"optional\s+pitch\s+list", r"pitch\s+list", r"ピッチ.*リスト"])
    if hit2:
        lr, lc, _ = hit2
        col = lc + 1
        vals: List[float] = []
        blank = 0
        for r in range(lr + 1, min(ws.max_row, lr + 200) + 1):
            v = ws.cell(r, col).value
            if v is None or norm_text(v) == "":
                blank += 1
                if blank >= 3:
                    break
                continue
            blank = 0
            try:
                vals.append(float(v))
            except Exception:
                break
        return [v for v in vals if v > 0]

    return []


def read_input_xlsx(path: str) -> Tuple[Config, Material, SolverSettings, List[Section]]:
    wb = load_workbook(path, data_only=True)
    if "INPUT" not in wb.sheetnames:
        raise ValueError('Sheet "INPUT" not found.')
    ws = wb["INPUT"]

    # ---- scalar values by labels (row insert tolerant)
    Lx = to_float(read_value_right_of_label(ws, [r"^Lx\b"], name="Lx"), name="Lx", where="label(Lx)")
    Ly = to_float(read_value_right_of_label(ws, [r"^Ly\b"], name="Ly"), name="Ly", where="label(Ly)")
    q  = to_float(read_value_right_of_label(ws, [r"^q\b"],  name="q"),  name="q",  where="label(q)")

    enable_x = to_bool(read_value_right_of_label(ws, [r"enable\s*x", r"X-direction", r"X\s*方向"], name="enable_x", required=False, default=True))
    enable_y = to_bool(read_value_right_of_label(ws, [r"enable\s*y", r"Y-direction", r"Y\s*方向"], name="enable_y", required=False, default=True))
    edge_beams = to_bool(read_value_right_of_label(ws, [r"both\s+edges", r"edge", r"端部", r"外周"], name="edge_beams", required=False, default=True))

    # Perimeter is assumed to be supported by primary beams.
    # If perimeter_primary=True, we do NOT place secondary beams at x=0/Lx or y=0/Ly.
    perimeter_primary = to_bool(read_value_right_of_label(
        ws,
        [r"perimeter.*primary", r"primary.*perimeter", r"外周.*大梁", r"外周.*主梁", r"主梁.*外周"],
        name="perimeter_primary",
        required=False,
        default=True,
    ))
    if perimeter_primary:
        edge_beams = False


    load_rule = to_int(read_value_right_of_label(ws, [r"distribution\s+rule", r"load\s*rule", r"配分", r"最近傍", r"按分"], name="load_rule", required=False, default=2),
                       name="load_rule", where="label(load_rule)")

    pitch_start = to_float(read_value_right_of_label(ws, [r"pitch\s*start", r"開始.*ピッチ"], name="pitch_start", required=False, default=1.5),
                           name="pitch_start", where="label(pitch_start)")
    pitch_end   = to_float(read_value_right_of_label(ws, [r"pitch\s*end", r"終了.*ピッチ"], name="pitch_end", required=False, default=3.0),
                           name="pitch_end", where="label(pitch_end)")
    pitch_step  = to_float(read_value_right_of_label(ws, [r"pitch\s*step", r"刻み.*ピッチ"], name="pitch_step", required=False, default=0.5),
                           name="pitch_step", where="label(pitch_step)")

    short_pitch_limit = to_float(
        read_value_right_of_label(ws, [r"short[-\s]*side\s+pitch\s+limit", r"短辺.*ピッチ.*上限", r"pitch\s+limit"], name="short_pitch_limit", required=False, default=3.0),
        name="short_pitch_limit", where="label(short_pitch_limit)"
    )

    # ---- material (E with unit)
    E_val = to_float(read_value_right_of_label(ws, [r"^E\b"], name="E", required=False, default=205000.0),
                     name="E", where="label(E)")
    # unit is often in the next column (value in B, unit in C). If not found, default N/mm2.
    E_unit = read_value_right_of_label(ws, [r"^E\b"], value_offset_cols=2, name="E_unit", required=False, default="N/mm2")
    E_unit_s = norm_text(E_unit).lower() if E_unit is not None else "n/mm2"
    if E_unit_s in ("n/mm2", "n/mm^2", "mpa"):
        E_kN_m2 = nmm2_to_kN_m2(E_val)
    elif E_unit_s in ("kn/m2", "kn/m^2"):
        E_kN_m2 = float(E_val)
    else:
        raise ValueError('E_unit must be "N/mm2"(MPa) or "kN/m2".')

    fb = to_float(read_value_right_of_label(ws, [r"^fb\b", r"fb\s*\[?\s*n/mm2", r"曲げ.*許容"], name="fb", required=False, default=165.0),
                  name="fb", where="label(fb)")
    fv = to_float(read_value_right_of_label(ws, [r"^fv\b", r"fv\s*\[?\s*n/mm2", r"せん断.*許容"], name="fv", required=False, default=95.0),
                  name="fv", where="label(fv)")
    defl_lim = to_float(read_value_right_of_label(ws, [r"deflection[_\s]*limit", r"たわみ", r"L/"], name="deflection_limit", required=False, default=360.0),
                        name="deflection_limit", where="label(deflection_limit)")

    mat = Material(
        E_kN_m2=E_kN_m2,
        fb_kN_m2=nmm2_to_kN_m2(fb),
        fv_kN_m2=nmm2_to_kN_m2(fv),
        deflection_limit=defl_lim
    )

    tol = to_float(read_value_right_of_label(ws, [r"^tol\b", r"収束", r"tolerance"], name="tol", required=False, default=1e-6),
                   name="tol", where="label(tol)")
    max_iter = to_int(read_value_right_of_label(ws, [r"max[_\s]*iter", r"最大.*反復"], name="max_iter", required=False, default=5),
                      name="max_iter", where="label(max_iter)")
    n_div = to_int(read_value_right_of_label(ws, [r"^n[_\s]*div\b", r"分割", r"n_div"], name="n_div", required=False, default=2000),
                   name="n_div", where="label(n_div)")
    setts = SolverSettings(tol=tol, max_iter=max_iter, n_div=n_div)

    # ---- point load table (header search)
    # IMPORTANT: Table termination is "blank LoadID" OR "P/x/y not numeric".
    # This avoids accidentally treating the next block title (e.g., "Layout Candidates") as a load row.
    loads: List[PointLoad] = []
    pl_hdr = find_header_row(ws, ["LoadID", "P [kN]", "x [m]", "y [m]"])
    if pl_hdr:
        hr, cmap = pl_hdr
        cL = cmap["LoadID"]
        cP = cmap["P [kN]"]
        cX = cmap["x [m]"]
        cY = cmap["y [m]"]

        blank = 0
        for r in range(hr + 1, min(ws.max_row, hr + 500) + 1):
            lid = norm_text(ws.cell(r, cL).value)
            if not lid:
                blank += 1
                if blank >= 3:
                    break
                continue
            blank = 0

            P_raw = ws.cell(r, cP).value
            x_raw = ws.cell(r, cX).value
            y_raw = ws.cell(r, cY).value

            # If numeric columns are empty or non-numeric, treat as end-of-table.
            if P_raw is None or x_raw is None or y_raw is None:
                break
            try:
                P = float(P_raw); x = float(x_raw); y = float(y_raw)
            except Exception:
                break

            loads.append(PointLoad(load_id=lid, P=P, x=x, y=y))
    # If no header found: treat as zero point loads (not an error)

    # ---- pitch list (optional)
    pitch_list = read_pitch_list(ws)

    # ---- section table (header search)
    sec_hdr = find_header_row(ws, ["Rank", "SectionName", "w_g [kN/m]"])
    if not sec_hdr:
        raise ValueError("Section table header not found. Required headers: Rank, SectionName, w_g [kN/m].")
    sec_hr, sec_cmap = sec_hdr

    # Optional columns (if present, read; if not present, None)
    # We'll extend col_map with these when found in the header row.
    opt_headers = ["h [mm]", "b [mm]", "tw [mm]", "tf [mm]", "A [mm2]", "Z [mm3]", "I [mm4]", "Av [mm2]"]
    # build a header->col map for the sec header row
    header_texts = {}
    for c in range(1, min(80, ws.max_column) + 1):
        t = norm_text(ws.cell(sec_hr, c).value)
        if t:
            header_texts[t.strip().lower()] = c

    def find_opt_col(h: str) -> Optional[int]:
        key = h.strip().lower()
        # allow variations like "A [mm^2]" etc
        for k, col in header_texts.items():
            if k == key or k.startswith(key):
                return col
        return None

    sec_col_map = dict(sec_cmap)  # required
    opt_col_map = {h: find_opt_col(h) for h in opt_headers}

    # read rows below header
    # - key is Rank
    sections: List[Section] = []
    blank = 0
    r = sec_hr + 1
    while r <= ws.max_row and r <= sec_hr + 2000:
        rk = ws.cell(r, sec_col_map["Rank"]).value
        if rk is None or norm_text(rk) == "":
            blank += 1
            if blank >= 3:
                break
            r += 1
            continue
        blank = 0

        # skip repeated header row
        if isinstance(rk, str) and rk.strip().lower() == "rank":
            r += 1
            continue

        name = ws.cell(r, sec_col_map["SectionName"]).value
        w_g = ws.cell(r, sec_col_map["w_g [kN/m]"]).value
        if name is None or norm_text(name) == "":
            r += 1
            continue
        if w_g is None or norm_text(w_g) == "":
            raise ValueError(f"w_g missing at section row {r}")

        def get_opt_by_header(hname: str) -> Optional[float]:
            col = opt_col_map.get(hname)
            if not col:
                return None
            v = ws.cell(r, col).value
            if v is None or norm_text(v) == "":
                return None
            return float(v)

        sec = Section(
            rank=to_int(rk, name="Rank", where=f"Section row {r}"),
            name=norm_text(name),
            w_g=to_float(w_g, name="w_g", where=f"Section row {r}"),
            h=get_opt_by_header("h [mm]"),
            b=get_opt_by_header("b [mm]"),
            tw=get_opt_by_header("tw [mm]"),
            tf=get_opt_by_header("tf [mm]"),
            A_mm2=get_opt_by_header("A [mm2]"),
            Z_mm3=get_opt_by_header("Z [mm3]"),
            I_mm4=get_opt_by_header("I [mm4]"),
            Av_mm2=get_opt_by_header("Av [mm2]"),
        )
        sections.append(sec)
        r += 1

    sections = sorted(sections, key=lambda s: s.rank)
    if not sections:
        raise ValueError("No sections provided in section table.")

    cfg = Config(
        Lx=float(Lx), Ly=float(Ly), q=float(q), loads=loads,
        enable_x=bool(enable_x), enable_y=bool(enable_y),
        edge_beams=bool(edge_beams), perimeter_primary=bool(perimeter_primary), load_rule=int(load_rule),
        pitch_start=float(pitch_start), pitch_end=float(pitch_end), pitch_step=float(pitch_step),
        pitch_list=pitch_list,
        short_pitch_limit=float(short_pitch_limit)
    )

    return cfg, mat, setts, sections


# -----------------------------
# Reporting to Excel "RESULT"
# -----------------------------
def set_col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def write_header(ws, row: int, col: int, text: str) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor="DDDDDD")
    c.alignment = Alignment(horizontal="center", vertical="center")


def set_row_height(ws, row: int, height: float) -> None:
    ws.row_dimensions[row].height = height


def _plan_grid_resolution(L: float, max_cells: int, min_cells: int) -> int:
    """
    Decide number of grid cells (including boundaries) for one axis.
    Target is ~0.25 m per cell, but capped.
    """
    if L <= 0:
        return min_cells
    target = 0.25
    n = int(L / target) + 1
    return max(min_cells, min(max_cells, n))


def draw_plan(ws, top: int, left: int, cfg: Config, best: CandidateResult) -> int:
    """
    Improved plan view (cells only, no Excel shapes).

    What you get:
      - Boundary box with thick border
      - Beam lines with medium border + light fill
      - Beam IDs + coordinates (x=.. or y=..) placed on margins
      - x/y axis ticks (major ticks)
      - Point loads plotted as "●<LoadID>" at (x,y)
    """
    beams, _ = build_beams(cfg, best.direction, best.pitch)

    Lx, Ly = cfg.Lx, cfg.Ly

    # Grid size (including boundary cells)
    max_cols = 60
    max_rows = 40
    nx = _plan_grid_resolution(Lx, max_cols, 20)
    ny = _plan_grid_resolution(Ly, max_rows, 15)
    dx = Lx / (nx - 1)
    dy = Ly / (ny - 1)

    # Layout (with margins)
    title_row = top
    id_row = top + 1          # beam IDs for Y-direction (optional)
    x_axis_row = top + 2      # x ticks
    grid_row0 = top + 3       # grid starts here (includes boundary)
    y_label_col = left        # y labels and X-direction beam IDs
    grid_col0 = left + 1

    # Styling
    thick = Side(style="thick")
    med = Side(style="medium")
    thin = Side(style="thin")
    dotted = Side(style="dotted")

    border_box = Border(left=thick, right=thick, top=thick, bottom=thick)
    border_thin_box = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_beam_h = Border(bottom=med)
    border_beam_v = Border(right=med)

    fill_beam = PatternFill("solid", fgColor="E6EEF8")
    fill_boundary = PatternFill("solid", fgColor="F2F2F2")
    font_bold = Font(bold=True)
    font_small = Font(size=9)
    font_load = Font(bold=True, color="CC0000")
    fill_alloc = PatternFill("solid", fgColor="FFF2CC")  # highlight allocated beam points
    font_alloc = Font(bold=True, color="1F4E79")

    # Title
    ws.cell(row=title_row, column=left, value=f"PLAN: {best.direction}-dir, pitch={best.pitch} m, sec={best.section_name}").font = Font(bold=True, size=12)

    # Set column widths for plan area (LAYOUT sheet use)
    ws.column_dimensions[get_column_letter(y_label_col)].width = 18
    for c in range(grid_col0, grid_col0 + nx):
        ws.column_dimensions[get_column_letter(c)].width = 2.2
    for r in range(grid_row0, grid_row0 + ny):
        set_row_height(ws, r, 13)

    # Clear area (labels + grid)
    for r in range(title_row, grid_row0 + ny + 6):
        for c in range(y_label_col, grid_col0 + nx + 2):
            ws.cell(row=r, column=c, value=None)
            ws.cell(row=r, column=c).border = Border()
            ws.cell(row=r, column=c).fill = PatternFill()

    # Axis ticks (major)
    def major_tick_step(L: float) -> float:
        if L <= 4:
            return 0.5
        if L <= 10:
            return 1.0
        if L <= 20:
            return 2.0
        return 5.0

    x_step = major_tick_step(Lx)
    y_step = major_tick_step(Ly)

    # x-axis labels
    ws.cell(row=x_axis_row, column=y_label_col, value="y[m]\\x[m]").font = font_bold
    x = 0.0
    while x <= Lx + 1e-9:
        cc = grid_col0 + int(round(x / dx))
        cc = max(grid_col0, min(grid_col0 + nx - 1, cc))
        ws.cell(row=x_axis_row, column=cc, value=round(x, 2)).font = font_small
        ws.cell(row=x_axis_row, column=cc).alignment = Alignment(horizontal="center")
        x += x_step

    # y-axis labels
    y = 0.0
    while y <= Ly + 1e-9:
        rr = grid_row0 + int(round(y / dy))
        rr = max(grid_row0, min(grid_row0 + ny - 1, rr))
        ws.cell(row=rr, column=y_label_col, value=round(y, 2)).font = font_small
        ws.cell(row=rr, column=y_label_col).alignment = Alignment(horizontal="right")
        y += y_step

    # Build empty grid with thin borders
    for rr in range(grid_row0, grid_row0 + ny):
        for cc in range(grid_col0, grid_col0 + nx):
            cell = ws.cell(row=rr, column=cc, value="")
            cell.border = border_thin_box
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # light background at boundary for readability
            if rr in (grid_row0, grid_row0 + ny - 1) or cc in (grid_col0, grid_col0 + nx - 1):
                cell.fill = fill_boundary

    # Boundary thick box
    for cc in range(grid_col0, grid_col0 + nx):
        ws.cell(row=grid_row0, column=cc).border = Border(top=thick, left=thin, right=thin, bottom=thin)
        ws.cell(row=grid_row0 + ny - 1, column=cc).border = Border(bottom=thick, left=thin, right=thin, top=thin)
    for rr in range(grid_row0, grid_row0 + ny):
        ws.cell(row=rr, column=grid_col0).border = Border(left=thick, top=thin, bottom=thin, right=thin)
        ws.cell(row=rr, column=grid_col0 + nx - 1).border = Border(right=thick, top=thin, bottom=thin, left=thin)
    # corners
    ws.cell(row=grid_row0, column=grid_col0).border = Border(left=thick, top=thick)
    ws.cell(row=grid_row0, column=grid_col0 + nx - 1).border = Border(right=thick, top=thick)
    ws.cell(row=grid_row0 + ny - 1, column=grid_col0).border = Border(left=thick, bottom=thick)
    ws.cell(row=grid_row0 + ny - 1, column=grid_col0 + nx - 1).border = Border(right=thick, bottom=thick)

    # Beam drawing helpers
    def map_x(xm: float) -> int:
        return grid_col0 + int(round(xm / dx))

    def map_y(ym: float) -> int:
        return grid_row0 + int(round(ym / dy))

    def clamp_col(c: int) -> int:
        return max(grid_col0, min(grid_col0 + nx - 1, c))

    def clamp_row(r: int) -> int:
        return max(grid_row0, min(grid_row0 + ny - 1, r))

    # Tributary boundaries (midpoints) as dotted lines
    def draw_trib_boundaries():
        # Tributary boundaries as midpoints between adjacent supports (including perimeter primary beams).
        if not beams:
            return
        if best.direction == "X":
            width = cfg.Ly  # pitch axis = Y
            supports = sorted(set([0.0, width] + [b.pos for b in beams]))
            mids = [(supports[i] + supports[i + 1]) * 0.5 for i in range(len(supports) - 1)]
            for ym in mids:
                rr = clamp_row(map_y(ym))
                for cc in range(grid_col0 + 1, grid_col0 + nx - 1):
                    ws.cell(row=rr, column=cc).border = Border(top=dotted)
        else:
            width = cfg.Lx  # pitch axis = X
            supports = sorted(set([0.0, width] + [b.pos for b in beams]))
            mids = [(supports[i] + supports[i + 1]) * 0.5 for i in range(len(supports) - 1)]
            for xm in mids:
                cc = clamp_col(map_x(xm))
                for rr in range(grid_row0 + 1, grid_row0 + ny - 1):
                    ws.cell(row=rr, column=cc).border = Border(left=dotted)

    draw_trib_boundaries()

    # Draw beams with IDs on margins
    if best.direction == "X":
        # beams are horizontal at y=pos
        for b in beams:
            rr = clamp_row(map_y(b.pos))
            # label on margin (y_label_col)
            ws.cell(row=rr, column=y_label_col, value=f"{b.beam_id}  y={b.pos:.3f}").font = font_bold
            ws.cell(row=rr, column=y_label_col).alignment = Alignment(horizontal="right")
            for cc in range(grid_col0 + 1, grid_col0 + nx - 1):
                cell = ws.cell(row=rr, column=cc)
                cell.fill = fill_beam
                # emphasize beam line
                cell.border = Border(bottom=med)
    else:
        # beams are vertical at x=pos
        for b in beams:
            cc = clamp_col(map_x(b.pos))
            # label on id_row (top margin)
            ws.cell(row=id_row, column=cc, value=f"{b.beam_id}\nx={b.pos:.3f}").font = font_bold
            ws.cell(row=id_row, column=cc).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            set_row_height(ws, id_row, 28)
            for rr in range(grid_row0 + 1, grid_row0 + ny - 1):
                cell = ws.cell(row=rr, column=cc)
                cell.fill = fill_beam
                cell.border = Border(right=med)

    # Plot point loads
    # plot point loads (actual location): ●ID
    load_by_id = {pl.load_id: pl for pl in cfg.loads}
    for pl in cfg.loads:
        rr = clamp_row(map_y(pl.y))
        cc = clamp_col(map_x(pl.x))
        txt = f"●{pl.load_id}"
        cell = ws.cell(row=rr, column=cc, value=txt)
        cell.font = font_load
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # mark allocated beams (which beam(s) actually receive the point load per Rule-1/Rule-2):
    #   place ▲ID on the receiving beam line at the same span coordinate.
    beam_by_id = {b.beam_id: b for b in beams}
    for lid, bid, allocP in best.allocation_rows:
        pl = load_by_id.get(lid)
        b = beam_by_id.get(bid)
        if pl is None or b is None:
            continue
        if best.direction == "X":
            xm, ym = pl.x, b.pos
        else:
            xm, ym = b.pos, pl.y
        rr = clamp_row(map_y(ym))
        cc = clamp_col(map_x(xm))
        cell = ws.cell(row=rr, column=cc)
        add = f"▲{lid}"
        existing = "" if cell.value is None else str(cell.value)
        if add not in existing:
            cell.value = (existing + add) if existing else add
        cell.font = font_alloc
        cell.fill = fill_alloc
        cell.alignment = Alignment(horizontal="center", vertical="center")


    # Legend (below)
    leg_r = grid_row0 + ny + 1
    ws.cell(row=leg_r, column=left, value="Legend:").font = font_bold
    ws.cell(row=leg_r + 1, column=left, value="Beam: light fill + medium line, Tributary boundary: dotted line, Point load: red ●ID (actual location), Allocation marker: ▲ID on receiving beam").font = font_small

    return leg_r + 3


def write_layout_sheet(wb, cfg: Config, best_A: Optional[CandidateResult], best_B: Optional[CandidateResult]) -> None:
    """
    Create a dedicated sheet 'LAYOUT' for visual/layout outputs (A+E measures):
      - Improved PLAN view (with beam IDs/coords, axis ticks, point loads)
      - Beam layout table (ID, position, tributary width, UDL etc.)
      - Allocation table (LoadID -> BeamID mapping with coordinates)
    """
    if "LAYOUT" in wb.sheetnames:
        del wb["LAYOUT"]
    ws = wb.create_sheet("LAYOUT")

    # Basic column widths (plan will override for its own area)
    for c in range(1, 90):
        ws.column_dimensions[get_column_letter(c)].width = 11

    r = 1
    ws.cell(row=r, column=1, value="LAYOUT / PLAN VIEW (cells only)").font = Font(bold=True, size=14)
    r += 2

    # quick lookup for original point loads
    load_map = {pl.load_id: pl for pl in cfg.loads}

    def write_beam_table(row0: int, best: CandidateResult) -> int:
        beams, _ = build_beams(cfg, best.direction, best.pitch)
        if not beams:
            ws.cell(row=row0, column=1, value="No beams.").font = Font(bold=True)
            return row0 + 2

        span = beams[0].span
        w_g = best.total_weight / (span * len(beams)) if span > 0 and len(beams) > 0 else 0.0

        ws.cell(row=row0, column=1, value="Beam layout table").font = Font(bold=True)
        headers = ["BeamID", "Dir", "Span[m]", "PitchAxis", "Pos[m]", "TribWidth[m]", "w_q[kN/m]", "w_g[kN/m]", "w_total[kN/m]", "N_point", "SumP[kN]", "PointLoads(P@a[m])"]
        for j, h in enumerate(headers, start=1):
            write_header(ws, row0 + 1, j, h)

        rr = row0 + 2
        pitch_axis = "y" if best.direction == "X" else "x"
        for b in beams:
            w_q = cfg.q * b.trib_width
            w_total = w_q + w_g
            npt = len(b.point_along)
            sump = sum(P for P, _a in b.point_along)
            pts = ", ".join([f"{P:.2f}@{a:.2f}" for P, a in sorted(b.point_along, key=lambda t: t[1])]) if b.point_along else ""
            row = [b.beam_id, best.direction, round(span, 6), pitch_axis, round(b.pos, 6), round(b.trib_width, 6),
                   round(w_q, 6), round(w_g, 6), round(w_total, 6), npt, round(sump, 6), pts]
            for j, v in enumerate(row, start=1):
                ws.cell(row=rr, column=j, value=v)
            rr += 1

        return rr + 2

    def write_alloc_table(row0: int, best: CandidateResult) -> int:
        ws.cell(row=row0, column=1, value="Load allocation table").font = Font(bold=True)
        headers = ["LoadID", "P_orig[kN]", "x[m]", "y[m]", "BeamID", "P_alloc[kN]"]
        for j, h in enumerate(headers, start=1):
            write_header(ws, row0 + 1, j, h)

        rr = row0 + 2
        for lid, bid, p in best.allocation_rows:
            pl = load_map.get(lid)
            if pl is None:
                P0 = ""
                x = ""
                y = ""
            else:
                P0 = pl.P
                x = pl.x
                y = pl.y
            row = [lid, P0, x, y, bid, round(p, 6)]
            for j, v in enumerate(row, start=1):
                ws.cell(row=rr, column=j, value=v)
            rr += 1
        return rr + 2

    def write_mode_block(title: str, best: Optional[CandidateResult], row0: int) -> int:
        ws.cell(row=row0, column=1, value=title).font = Font(bold=True, size=12)
        row0 += 1
        if best is None:
            ws.cell(row=row0, column=1, value="NO FEASIBLE SOLUTION").font = Font(bold=True)
            return row0 + 2

        # PLAN
        row0 = draw_plan(ws, row0, 1, cfg, best)

        # Tables
        row0 = write_beam_table(row0, best)
        row0 = write_alloc_table(row0, best)
        return row0 + 2

    r = write_mode_block("Mode-A best (total weight minimum)", best_A, r)
    r = write_mode_block("Mode-B best (minimize maximum section rank)", best_B, r)


def write_result_xlsx(
    in_path: str,
    out_path: str,
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    results: List[CandidateResult],
    best_A: Optional[CandidateResult],
    best_B: Optional[CandidateResult]
) -> None:
    wb = load_workbook(in_path)
    if "RESULT" in wb.sheetnames:
        del wb["RESULT"]
    ws = wb.create_sheet("RESULT")

    for i, w in enumerate([10, 8, 8, 8, 22, 12, 12, 12, 12, 12, 10, 10, 10, 10, 35], start=1):
        set_col_width(ws, i, w)

    r = 1
    ws.cell(row=r, column=1, value="Secondary Beam Optimization RESULT").font = Font(bold=True, size=14)
    r += 2
    ws.cell(row=r, column=1, value="Notes: Units are kN, m. fb/fv/E are converted from N/mm2 to kN/m2.").font = Font(italic=True)
    r += 2

    def write_summary(title: str, best: Optional[CandidateResult], row0: int) -> int:
        ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
        headers = ["Direction", "Pitch[m]", "Section(Rank)", "TotalWeight[kN]", "Mmax[kN*m]", "Vmax[kN]", "dmax[mm]", "util(M)", "util(V)", "util(d)", "OK/NG"]
        for j, h in enumerate(headers, start=1):
            write_header(ws, row0 + 1, j, h)
        if best is None:
            ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
            return row0 + 4
        vals = [
            best.direction,
            best.pitch,
            f"{best.section_name} (R{best.section_rank})",
            round(best.total_weight, 6),
            round(best.Mmax, 6),
            round(best.Vmax, 6),
            round(best.dmax * 1000.0, 6),
            round(best.util_M, 6),
            round(best.util_V, 6),
            round(best.util_d, 6),
            "OK" if best.ok else f"NG: {best.ng_reason}",
        ]
        for j, v in enumerate(vals, start=1):
            ws.cell(row=row0 + 2, column=j, value=v)
        return row0 + 4

    r = write_summary("Mode-A: Total weight minimum", best_A, r)
    r = write_summary("Mode-B: Minimize maximum section rank (tie-break: weight)", best_B, r)

    ws.cell(row=r, column=1, value="All candidates").font = Font(bold=True)
    r += 1
    headers = ["Dir", "Pitch[m]", "Nbeams", "Rank", "Section", "w_total_max[kN/m]", "Mmax[kN*m]", "Vmax[kN]", "dmax[mm]",
               "utilM", "utilV", "utilD", "utilMax", "TotalWeight[kN]", "OK/NG (reason)"]
    for j, h in enumerate(headers, start=1):
        write_header(ws, r, j, h)
    r += 1

    results_sorted = sorted(results, key=lambda x: (not x.ok, x.total_weight, x.section_rank, x.direction, x.pitch))
    for res in results_sorted:
        row = [
            res.direction, res.pitch, res.n_beams, res.section_rank, res.section_name,
            round(res.w_total_max, 6),
            round(res.Mmax, 6),
            round(res.Vmax, 6),
            round(res.dmax * 1000.0, 6),
            round(res.util_M, 6),
            round(res.util_V, 6),
            round(res.util_d, 6),
            round(res.util_max, 6),
            round(res.total_weight, 6),
            "OK" if res.ok else f"NG: {res.ng_reason}",
        ]
        for j, v in enumerate(row, start=1):
            ws.cell(row=r, column=j, value=v)
        r += 1

    r += 2

    def write_alloc(title: str, best: Optional[CandidateResult], row0: int) -> int:
        ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
        write_header(ws, row0 + 1, 1, "LoadID")
        write_header(ws, row0 + 1, 2, "BeamID")
        write_header(ws, row0 + 1, 3, "AllocatedP[kN]")
        if best is None:
            ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
            return row0 + 4
        rr = row0 + 2
        for lid, bid, p in best.allocation_rows:
            ws.cell(row=rr, column=1, value=lid)
            ws.cell(row=rr, column=2, value=bid)
            ws.cell(row=rr, column=3, value=round(p, 6))
            rr += 1
        return rr + 1

    r = write_alloc("Allocation (Mode-A best)", best_A, r)
    r = write_alloc("Allocation (Mode-B best)", best_B, r)

    def write_member_list(title: str, best: Optional[CandidateResult], row0: int) -> int:
        ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
        write_header(ws, row0 + 1, 1, "Section")
        write_header(ws, row0 + 1, 2, "Count")
        write_header(ws, row0 + 1, 3, "Length each [m]")
        write_header(ws, row0 + 1, 4, "w_g [kN/m]")
        write_header(ws, row0 + 1, 5, "Weight [kN]")
        if best is None:
            ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
            return row0 + 4
        beams, _ = build_beams(cfg, best.direction, best.pitch)
        w_g = best.total_weight / (beams[0].span * len(beams)) if beams else 0.0
        ws.cell(row=row0 + 2, column=1, value=best.section_name)
        ws.cell(row=row0 + 2, column=2, value=len(beams))
        ws.cell(row=row0 + 2, column=3, value=beams[0].span if beams else 0.0)
        ws.cell(row=row0 + 2, column=4, value=round(w_g, 6))
        ws.cell(row=row0 + 2, column=5, value=round(best.total_weight, 6))
        return row0 + 5

    r = write_member_list("Member list (Mode-A best)", best_A, r)
    r = write_member_list("Member list (Mode-B best)", best_B, r)

    r += 1
    ws.cell(row=r, column=1, value="PLAN view has been moved to sheet 'LAYOUT' (improved readability).").font = Font(italic=True)
    r += 2

    # Create dedicated layout sheet (A+E measures)
    write_layout_sheet(wb, cfg, best_A, best_B)

    wb.save(out_path)



# -----------------------------
# Main
# -----------------------------
def main(argv: List[str]) -> int:
    if len(argv) < 2:
        print("Usage: python beam_optimizer_b1_flex.py input.xlsx output.xlsx")
        return 2

    in_path = argv[1]
    out_path = argv[2] if len(argv) >= 3 else (in_path.replace(".xlsx", "") + "_out.xlsx")

    cfg, mat, setts, sections = read_input_xlsx(in_path)

    ensure_positive("Lx", cfg.Lx)
    ensure_positive("Ly", cfg.Ly)
    ensure_positive("q", cfg.q)
    ensure_positive("E", mat.E_kN_m2)
    ensure_positive("fb", mat.fb_kN_m2)
    ensure_positive("fv", mat.fv_kN_m2)
    ensure_positive("deflection_limit", mat.deflection_limit)

    results, best_A, best_B = optimize(cfg, mat, setts, sections)
    write_result_xlsx(in_path, out_path, cfg, mat, setts, results, best_A, best_B)

    print(f"Done. Output written to: {out_path}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main(sys.argv))
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)
