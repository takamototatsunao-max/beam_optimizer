#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
beam_optimizer_b9.py

README / 実行方法
-----------------
- 前提: Python 3.x, openpyxl
- 実行例:
    python beam_optimizer_b9.py input.xlsx output.xlsx

入出力
------
- 入力:  Excelファイル（.xlsx）のシート "INPUT"
- 出力:  output.xlsx にシート "RESULT" と "LAYOUT" を生成して書き込み
         ※Excel図形は不使用。LAYOUTはセルのみの簡易伏図（表はRESULTへ）。

最適化（本版は Mode-A のみ）
---------------------------
(Mode-A) 総重量最小:
    - MAIN小梁/TRANS小梁とも「部材ごとに」断面を選定（1本ごとに断面が違ってOK）。
    - 方向×ピッチの組合せを全探索し、総重量（Σ w_g[kN/m]×長さ[m]）最小の案を採用。
    - 各部材の断面選定は「成立する候補の中で最軽量（w_g最小）」を採用。

荷重モデル（固定仮定）
----------------------
- 面荷重: w_q = q[kN/m2] × 支配幅[m]（支配幅は隣接梁の中間で区切った領域幅）
- 自重: 断面候補の w_g[kN/m] を等分布荷重として加算
- 集中荷重:
    - 配置ライン（MAIN小梁の中心線）に乗っている場合: そのMAIN小梁へ付与
    - 配置ライン外の場合: その荷重点を通る TRANS小梁（黄色）を追加し、
      両側の支持（MAIN小梁または外周大梁=PERIM）へ反力として配分

LAYOUT 表現
-----------
- 同一セル座標に「実荷重点●」と「割当▲」が重なる場合、1セルにまとめて表示（例: ●▲P4）。
- 梁ラベルに util（util_max）を表示。TRANS梁も "TB_P4 u=0.03" のように表示。
- util最大の部材は赤塗り。

断面諸元入力が空欄の場合の内部計算
----------------------------------
- SectionName が "H-350x175x7x11" の形式であれば、(h,b,tw,tf) を文字列から抽出。
- A/Z/I/Av が空欄の場合は、フィレット無視の近似断面として以下で自動算定:
    A  = 2*b*tf + (h-2*tf)*tw
    I  = 2*(b*tf^3/12 + b*tf*(h/2-tf/2)^2) + tw*(h-2*tf)^3/12
    Z  = I/(h/2)
    Av = tw*(h-2*tf)  （せん断有効断面: ウェブ面積近似）
  ※単位はmm系で計算→m系に換算して照査。
"""


from __future__ import annotations

import math
import os
import re
import sys
import json
from datetime import datetime
from dataclasses import dataclass, replace
from typing import Dict, List, Optional, Tuple
from concurrent.futures import ProcessPoolExecutor, as_completed

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter



# -----------------------------
# Trace / Debug (verbose output)
# -----------------------------
_TRACE: List[Dict[str, object]] = []
_DBG_MEMBER_TRIALS: List[Dict[str, object]] = []
_DBG_MAIN_GEOMS: List[Dict[str, object]] = []
_DBG_TRANS_DEFS: List[Dict[str, object]] = []
_DBG_MEMBER_FINAL: List[Dict[str, object]] = []
_DBG_ALLOC_FINAL: List[Dict[str, object]] = []
_DBG_CTX: Dict[str, object] = {
    "cand_id": "",
    "direction": "",
    "pitch": None,
    "member_id": "",
    "member_type": "",
}


def _clear_debug() -> None:
    _TRACE.clear()
    _DBG_MEMBER_TRIALS.clear()
    _DBG_MAIN_GEOMS.clear()
    _DBG_TRANS_DEFS.clear()
    _DBG_MEMBER_FINAL.clear()
    _DBG_ALLOC_FINAL.clear()
    _DBG_CTX.update({"cand_id": "", "direction": "", "pitch": None, "member_id": "", "member_type": ""})


def _set_dbg_context(
    cand_id: str,
    direction: str,
    pitch: float,
    member_id: str = "",
    member_type: str = "",
) -> None:
    _DBG_CTX["cand_id"] = cand_id
    _DBG_CTX["direction"] = direction
    _DBG_CTX["pitch"] = pitch
    _DBG_CTX["member_id"] = member_id
    _DBG_CTX["member_type"] = member_type


def trace(phase: str, message: str, **data: object) -> None:
    """Collect verbose trace rows to be written to Excel."""
    try:
        t = datetime.now().isoformat(timespec="seconds")
    except Exception:
        t = ""
    row = {
        "seq": len(_TRACE) + 1,
        "time": t,
        "cand_id": _DBG_CTX.get("cand_id", ""),
        "direction": _DBG_CTX.get("direction", ""),
        "pitch": _DBG_CTX.get("pitch", None),
        "member_id": _DBG_CTX.get("member_id", ""),
        "member_type": _DBG_CTX.get("member_type", ""),
        "phase": phase,
        "message": message,
        "data": data,
    }
    _TRACE.append(row)


def _dbg_add_member_trial(sec: Section, bc: Optional["MemberCheck"], ok: bool, note: str = "") -> None:
    row = {
        "cand_id": _DBG_CTX.get("cand_id", ""),
        "direction": _DBG_CTX.get("direction", ""),
        "pitch": _DBG_CTX.get("pitch", None),
        "member_id": _DBG_CTX.get("member_id", ""),
        "member_type": _DBG_CTX.get("member_type", ""),
        "sec_rank": sec.rank,
        "sec_name": sec.name,
        "w_g": sec.w_g,
        "span": (bc.span if bc else None),
        "w_udl": (bc.w_udl if bc else None),
        "w_udl_max": (bc.w_udl_max if bc else None),
        "Ra": (bc.Ra if bc else None),
        "Rb": (bc.Rb if bc else None),
        "Mmax": (bc.Mmax if bc else None),
        "x_Mmax": (bc.x_Mmax if bc else None),
        "Vmax": (bc.Vmax if bc else None),
        "x_Vmax": (bc.x_Vmax if bc else None),
        "dmax": (bc.dmax if bc else None),
        "x_dmax": (bc.x_dmax if bc else None),
        "util_M": (bc.util_M if bc else None),
        "util_V": (bc.util_V if bc else None),
        "util_d": (bc.util_d if bc else None),
        "util_max": (bc.util_max if bc else None),
        "ok": ok,
        "note": note,
    }
    _DBG_MEMBER_TRIALS.append(row)


def _dbg_add_main_geoms(cand_id: str, direction: str, pitch: float, beams: List["BeamGeom"]) -> None:
    for b in beams:
        _DBG_MAIN_GEOMS.append({
            "cand_id": cand_id,
            "direction": direction,
            "pitch": pitch,
            "beam_id": b.beam_id,
            "pos": b.pos,
            "span": b.span,
            "trib_width": b.trib_width,
            "trib_left": b.trib_left,
            "trib_right": b.trib_right,
            "n_point": len(b.point_along),
        })


def _dbg_add_trans_defs(cand_id: str, direction: str, pitch: float, tdefs: List["TransferDef"]) -> None:
    for td in tdefs:
        _DBG_TRANS_DEFS.append({
            "cand_id": cand_id,
            "direction": direction,
            "pitch": pitch,
            "tb_id": td.tb_id,
            "span_dir": td.span_dir,
            "fixed_coord": td.fixed_coord,
            "left_support": td.left_support,
            "right_support": td.right_support,
            "left_pos": td.left_pos,
            "right_pos": td.right_pos,
            "a_tb": td.a_tb,
            "a_on_support": td.a_on_support,
            "load_id": td.load_id,
            "P": td.P,
        })


def _dbg_add_member_final(cand_id: str, direction: str, pitch: float, mc: "MemberCheck") -> None:
    _DBG_MEMBER_FINAL.append({
        "cand_id": cand_id,
        "direction": direction,
        "pitch": pitch,
        "beam_id": mc.beam_id,
        "beam_no": mc.beam_no,
        "member_type": mc.member_type,
        "span_dir": mc.direction,
        "pos_or_fixed": mc.pos_or_fixed,
        "left": mc.left,
        "right": mc.right,
        "span": mc.span,
        "sec_rank": mc.section_rank,
        "sec_name": mc.section_name,
        "w_g": mc.w_g,
        "w_udl": mc.w_udl,
        "w_udl_max": mc.w_udl_max,
        "Ra": mc.Ra,
        "Rb": mc.Rb,
        "n_point": mc.n_point,
        "Mmax": mc.Mmax,
        "x_Mmax": mc.x_Mmax,
        "Vmax": mc.Vmax,
        "x_Vmax": mc.x_Vmax,
        "dmax": mc.dmax,
        "x_dmax": mc.x_dmax,
        "util_max": mc.util_max,
        "ok": mc.ok,
    })


def _dbg_add_alloc_final(cand_id: str, direction: str, pitch: float, alloc_rows: List[Tuple[str, str, float]]) -> None:
    for lid, sid, p in alloc_rows:
        _DBG_ALLOC_FINAL.append({
            "cand_id": cand_id,
            "direction": direction,
            "pitch": pitch,
            "load_id": lid,
            "support": sid,
            "P_alloc": p,
        })


# -----------------------------
# Data models
# -----------------------------
@dataclass(frozen=True)
class PointLoad:
    load_id: str
    P: float  # kN
    x: float  # m (0..Lx)
    y: float  # m (0..Ly)


@dataclass(frozen=True)
class Section:
    rank: int
    name: str
    w_g: float      # kN/m
    h: Optional[float] = None   # mm
    b: Optional[float] = None   # mm
    tw: Optional[float] = None  # mm
    tf: Optional[float] = None  # mm
    A_mm2: Optional[float] = None
    Z_mm3: Optional[float] = None
    I_mm4: Optional[float] = None
    Av_mm2: Optional[float] = None


@dataclass(frozen=True)
class Material:
    E_kN_m2: float     # kN/m^2
    fb_kN_m2: float    # kN/m^2
    fv_kN_m2: float    # kN/m^2
    deflection_limit: float


@dataclass(frozen=True)
class SolverSettings:
    tol: float
    max_iter: int
    n_div: int


@dataclass(frozen=True)
class Config:
    Lx: float
    Ly: float
    q: float
    loads: List[PointLoad]

    # load sharing model for slab -> MAIN beams
    #   - 'KAMEKKO': tributary width varies along span (turtle-shell; 45deg) 
    #   - 'ONEWAY' : constant tributary width (b8-compatible)
    load_share_model: str

    enable_x: bool
    enable_y: bool

    # edge_beams=True: ピッチ起点を外周(0)に取り、0, pitch, 2pitch...を生成（ただし0とLはPERIM扱いでMAINには置かない）
    # edge_beams=False: ピッチ/2起点で内側に配置
    edge_beams: bool

    # 旧仕様互換（現状は TRANS追加ロジック優先）。残しておくが、load_ruleは直接使わない。
    load_rule: int

    pitch_start: float
    pitch_end: float
    pitch_step: float
    pitch_list: List[float]

    short_pitch_limit: float  # 3.0
    snap_tol: float           # "on-beam" 判定の許容[m]


@dataclass
class BeamGeom:
    beam_id: str
    direction: str  # 'X' or 'Y'  (MAIN span direction)
    pos: float      # position along pitch axis (m)
    span: float     # m
    trib_width: float  # m
    trib_left: float  # m (half spacing to left boundary/adjacent beam)
    trib_right: float  # m (half spacing to right boundary/adjacent beam)
    # point loads list for analysis: (P[kN], a[m] from left)
    point_along: List[Tuple[float, float]]


@dataclass(frozen=True)
class TransferDef:
    tb_id: str
    span_dir: str         # 'X' or 'Y' (TRANS span direction)
    fixed_coord: float    # m (coordinate perpendicular to span_dir)
    left_support: str     # BeamID or 'PERIM'
    right_support: str    # BeamID or 'PERIM'
    left_pos: float       # m coordinate on span_dir
    right_pos: float      # m coordinate on span_dir
    a_tb: float           # m from left support along TRANS
    a_on_support: float   # m along support span direction (to place reaction)
    load_id: str
    P: float              # kN (original point load)


@dataclass(frozen=True)
class MemberCheck:
    beam_id: str
    beam_no: int          # MAIN: 1.., TRANS: 101..
    member_type: str      # 'MAIN' or 'TRANS'
    direction: str        # span direction ('X'/'Y')
    pos_or_fixed: Optional[float]  # MAIN: pos along pitch, TRANS: fixed_coord
    left: Optional[float]          # TRANS: left_pos, else None
    right: Optional[float]         # TRANS: right_pos, else None
    span: float
    section_rank: int
    section_name: str
    w_g: float

    # distributed load summary [kN/m]
    w_udl: float          # average (b8-compatible)
    w_udl_max: float      # peak (KAMEKKO)

    # support reactions [kN]
    Ra: float
    Rb: float

    n_point: int

    # maxima
    Mmax: float
    x_Mmax: float
    Vmax: float
    x_Vmax: float
    dmax: float
    x_dmax: float

    util_M: float
    util_V: float
    util_d: float
    util_max: float
    ok: bool

@dataclass(frozen=True)
class Solution:
    # Mode-B is removed. This solution is always "Mode-A (total weight minimum)".
    direction: str
    pitch: float

    total_weight: float
    max_rank_used: int

    Mmax: float
    Vmax: float
    dmax: float
    util_max: float

    ok: bool
    ng_reason: str

    member_checks: List[MemberCheck]
    # For plan view and reporting
    main_geoms: List[BeamGeom]
    transfer_defs: List[TransferDef]
    allocation_rows: List[Tuple[str, str, float]]  # (LoadID, SupportID, AllocatedP[kN])
    worst_member_id: str


# -----------------------------
# Utility: unit conversions
# -----------------------------
def nmm2_to_kN_m2(val_nmm2: float) -> float:
    # 1 N/mm2 = 1 MPa = 1e6 N/m2 = 1e3 kN/m2
    return float(val_nmm2) * 1000.0


def ensure_positive(name: str, v: float) -> None:
    if v is None or not isinstance(v, (int, float)) or float(v) <= 0.0:
        raise ValueError(f"{name} must be positive. got={v}")


# -----------------------------
# Section properties (approx)
# -----------------------------
def parse_h_section_dims(name: str) -> Optional[Tuple[float, float, float, float]]:
    """
    Parse "H-350x175x7x11" or similar (x, X, ×).
    Returns (h,b,tw,tf) in mm if parse succeeds.
    """
    s = str(name).replace("×", "x").replace("X", "x")
    m = re.search(r"(\d+(\.\d+)?)x(\d+(\.\d+)?)x(\d+(\.\d+)?)x(\d+(\.\d+)?)", s)
    if not m:
        return None
    h = float(m.group(1))
    b = float(m.group(3))
    tw = float(m.group(5))
    tf = float(m.group(7))
    return h, b, tw, tf


def approx_h_section_props_mm(h: float, b: float, tw: float, tf: float) -> Tuple[float, float, float, float]:
    """
    Fillet ignored. Strong-axis bending (Ixx, Zx).
    Returns (A_mm2, I_mm4, Z_mm3, Av_mm2).
    """
    A = 2.0 * b * tf + (h - 2.0 * tf) * tw

    # Ixx = 2*(b*tf^3/12 + b*tf*y^2) + tw*(h-2tf)^3/12
    I_flange = b * tf**3 / 12.0
    y = h / 2.0 - tf / 2.0
    I = 2.0 * (I_flange + b * tf * y**2) + tw * (h - 2.0 * tf)**3 / 12.0
    Z = I / (h / 2.0)

    Av = tw * (h - 2.0 * tf)  # web area approximation
    return A, I, Z, Av


def get_section_props_m(section: Section) -> Tuple[float, float, float, float]:
    """
    Returns (A[m2], Av[m2], Z[m3], I[m4]).
    If inputs missing, tries to parse from name and approximate.
    """
    h, b, tw, tf = section.h, section.b, section.tw, section.tf

    if any(v is None for v in (h, b, tw, tf)):
        dims = parse_h_section_dims(section.name)
        if dims:
            h, b, tw, tf = dims

    if any(v is None for v in (h, b, tw, tf)):
        raise ValueError(f"Section dims missing and cannot parse from name: {section.name}")

    A_mm2 = section.A_mm2
    Z_mm3 = section.Z_mm3
    I_mm4 = section.I_mm4
    Av_mm2 = section.Av_mm2

    if any(v is None for v in (A_mm2, Z_mm3, I_mm4, Av_mm2)):
        A2, I2, Z2, Av2 = approx_h_section_props_mm(float(h), float(b), float(tw), float(tf))
        if A_mm2 is None:
            A_mm2 = A2
        if I_mm4 is None:
            I_mm4 = I2
        if Z_mm3 is None:
            Z_mm3 = Z2
        if Av_mm2 is None:
            Av_mm2 = Av2

    # Convert: mm2->m2 (1e-6), mm3->m3 (1e-9), mm4->m4 (1e-12)
    A = float(A_mm2) * 1e-6
    Av = float(Av_mm2) * 1e-6
    Z = float(Z_mm3) * 1e-9
    I = float(I_mm4) * 1e-12

    ensure_positive(f"A({section.name})", A)
    ensure_positive(f"Av({section.name})", Av)
    ensure_positive(f"Z({section.name})", Z)
    ensure_positive(f"I({section.name})", I)
    return A, Av, Z, I


# -----------------------------
# Geometry / beam layout
# -----------------------------
def pitch_direction_of(direction: str) -> str:
    # X方向配置: beams span X, are arrayed along Y => pitch_dir='Y'
    # Y方向配置: beams span Y, are arrayed along X => pitch_dir='X'
    return "Y" if direction == "X" else "X"


def short_side_axis(Lx: float, Ly: float) -> str:
    return "X" if Lx <= Ly else "Y"


def make_pitch_candidates(cfg: Config) -> List[float]:
    if cfg.pitch_list:
        vals = [float(v) for v in cfg.pitch_list if v is not None and str(v).strip() != "" and float(v) > 0]
        return sorted(set(vals))
    ensure_positive("pitch_start", cfg.pitch_start)
    ensure_positive("pitch_end", cfg.pitch_end)
    ensure_positive("pitch_step", cfg.pitch_step)
    if cfg.pitch_end < cfg.pitch_start:
        raise ValueError("pitch_end must be >= pitch_start")

    out = []
    x = cfg.pitch_start
    while x <= cfg.pitch_end + 1e-12:
        out.append(round(x, 10))
        x += cfg.pitch_step
    return sorted(set(out))


def positions_along(width: float, pitch: float, edge_beams: bool) -> List[float]:
    """
    Generate beam centerlines along pitch direction from 0..width.
    edge_beams=True: includes 0 and width (later excluded as PERIM).
    edge_beams=False: starts at pitch/2.
    """
    ensure_positive("width", width)
    ensure_positive("pitch", pitch)

    pos: List[float] = []
    if edge_beams:
        pos = [0.0]
        k = 1
        while k * pitch < width - 1e-9:
            pos.append(k * pitch)
            k += 1
        if abs(pos[-1] - width) > 1e-9:
            pos.append(width)
    else:
        k = 0
        while True:
            p = pitch / 2.0 + k * pitch
            if p > width - 1e-9:
                break
            pos.append(p)
            k += 1
        if not pos:
            pos = [width / 2.0]

    pos = sorted(set(round(p, 10) for p in pos))
    return pos


def tributary_widths(pos: List[float], width: float) -> List[float]:
    """
    Tributary width from midpoints between adjacent beams and boundaries (0,width).
    Note: PERIM strips are excluded automatically (handled by boundaries).
    """
    n = len(pos)
    trib: List[float] = []
    for i, p in enumerate(pos):
        left = 0.0 if i == 0 else 0.5 * (pos[i - 1] + p)
        right = width if i == n - 1 else 0.5 * (p + pos[i + 1])
        trib.append(right - left)
    return trib


def build_main_beams(cfg: Config, direction: str, pitch: float) -> List[BeamGeom]:
    """
    Build MAIN secondary beams (excluding PERIM at 0 and width).
    Notes:
      - PERIM strips at 0 and width are excluded from MAIN beams.
      - Tributary halves (trib_left/right) are defined by midpoints between adjacent beams and boundaries.
    """
    if direction == "X":
        span = cfg.Lx
        width = cfg.Ly
    else:
        span = cfg.Ly
        width = cfg.Lx

    pos_all = positions_along(width, pitch, cfg.edge_beams)

    # Exclude PERIM lines (0 and width) from MAIN beams
    pos = [p for p in pos_all if p > cfg.snap_tol and p < width - cfg.snap_tol]
    if not pos:
        pos = []

    beams: List[BeamGeom] = []
    n = len(pos)
    for i, p in enumerate(pos, start=1):
        left_bd = 0.0 if i == 1 else 0.5 * (pos[i - 2] + p)
        right_bd = width if i == n else 0.5 * (p + pos[i])
        trib_left = p - left_bd
        trib_right = right_bd - p
        trib_width = trib_left + trib_right
        beam_id = f"{direction}{i:02d}"
        beams.append(BeamGeom(
            beam_id=beam_id,
            direction=direction,
            pos=p,
            span=span,
            trib_width=trib_width,
            trib_left=trib_left,
            trib_right=trib_right,
            point_along=[]
        ))
    return beams
# -----------------------------
# Beam analysis (numerical)
# -----------------------------
def analyze_simply_supported(
    L: float,
    w_udl: float,
    point_loads: List[Tuple[float, float]],
    E: float,
    I: float,
    n_div: int
) -> Tuple[float, float, float]:
    """
    Returns (Mmax[kN*m], Vmax[kN], dmax[m]) for a simply-supported beam.
    Loads:
      - UDL w_udl [kN/m] along full span
      - point_loads: list of (P[kN], a[m] from left)
    """
    ensure_positive("L", L)
    if n_div < 50:
        raise ValueError("n_div too small (>=50 recommended)")

    pls = [(float(P), float(a)) for P, a in point_loads]
    for P, a in pls:
        if not (0.0 <= a <= L):
            raise ValueError(f"Point load position out of range: a={a}, L={L}")
        if P < 0:
            raise ValueError("Negative point load not supported.")
    pls.sort(key=lambda x: x[1])

    # reactions
    Ra = w_udl * L / 2.0
    Rb = w_udl * L / 2.0
    for P, a in pls:
        Ra += P * (L - a) / L
        Rb += P * a / L

    # discretize
    n = n_div + 1
    dx = L / n_div
    xs = [i * dx for i in range(n)]

    V = [Ra - w_udl * x for x in xs]
    M = [Ra * x - w_udl * x * x / 2.0 for x in xs]

    for P, a in pls:
        for i, x in enumerate(xs):
            if x >= a - 1e-12:
                V[i] -= P
                M[i] -= P * (x - a)

    Mmax = max(abs(v) for v in M)
    Vmax = max(abs(v) for v in V)

    # curvature k = M/(E*I)
    k = [m / (E * I) for m in M]

    # integrate curvature -> slope with trapezoid
    K1 = [0.0] * n
    for i in range(1, n):
        K1[i] = K1[i - 1] + 0.5 * (k[i - 1] + k[i]) * dx

    # integrate K1 -> helper integral
    intK1 = [0.0] * n
    for i in range(1, n):
        intK1[i] = intK1[i - 1] + 0.5 * (K1[i - 1] + K1[i]) * dx

    # enforce y(L)=0 => theta0 = -int_0^L K1(s) ds / L
    theta0 = -intK1[-1] / L
    theta = [theta0 + v for v in K1]

    # integrate slope -> deflection (y(0)=0)
    y = [0.0] * n
    for i in range(1, n):
        y[i] = y[i - 1] + 0.5 * (theta[i - 1] + theta[i]) * dx

    dmax = max(abs(v) for v in y)
    return Mmax, Vmax, dmax



def analyze_simply_supported_general(
    L: float,
    w_vals: List[float],
    point_loads: List[Tuple[float, float]],
    E: float,
    I: float,
    n_div: int,
    return_arrays: bool = False,
) -> Dict[str, object]:
    """
    General simply-supported beam analysis by numerical integration.

    Parameters
    ----------
    L : span [m]
    w_vals : distributed load values [kN/m] at nodes x=i*L/n_div (len = n_div+1)
             piecewise-linear interpolation is assumed.
    point_loads : list of (P[kN], a[m] from left)
    E, I : stiffness (kN/m^2, m^4)
    n_div : number of divisions (>=50 recommended)
    return_arrays : if True, returns x, V, M, y arrays for reporting

    Returns keys
    ------------
      Ra, Rb [kN]
      w_avg, w_max [kN/m]
      Mmax, Vmax, dmax [kN*m, kN, m]
      x_Mmax, x_Vmax, x_dmax [m]
      (optional) xs, V, M, y arrays
    """
    ensure_positive("L", L)
    if n_div < 50:
        raise ValueError("n_div too small (>=50 recommended)")
    n = n_div + 1
    if len(w_vals) != n:
        raise ValueError(f"w_vals length must be n_div+1 (= {n}). got={len(w_vals)}")

    dx = L / n_div
    xs = [i * dx for i in range(n)]
    w = [float(v) for v in w_vals]

    pls = [(float(P), float(a)) for P, a in point_loads]
    for P, a in pls:
        if not (0.0 <= a <= L):
            raise ValueError(f"Point load position out of range: a={a}, L={L}")
        if P < 0:
            raise ValueError("Negative point load not supported.")
    pls.sort(key=lambda x: x[1])

    # distributed load integrals:
    # I0(x)=∫ w ds, I1(x)=∫ w*s ds  (trapezoid on nodes)
    I0 = [0.0] * n
    I1 = [0.0] * n
    for i in range(1, n):
        w0, w1 = w[i - 1], w[i]
        x0, x1 = xs[i - 1], xs[i]
        I0[i] = I0[i - 1] + 0.5 * (w0 + w1) * dx
        # integral of w*s over segment using trapezoid on w*s
        I1[i] = I1[i - 1] + 0.5 * (w0 * x0 + w1 * x1) * dx

    Wtot = I0[-1]
    Mtot = I1[-1]  # about left support

    # reactions from distributed load
    Rb = Mtot / L
    Ra = Wtot - Rb

    # add point loads reactions
    for P, a in pls:
        Ra += P * (L - a) / L
        Rb += P * a / L

    # shear and moment along span
    V = [0.0] * n
    M = [0.0] * n
    for i, x in enumerate(xs):
        v = Ra - I0[i]
        m = Ra * x - (x * I0[i] - I1[i])
        for P, a in pls:
            if x >= a - 1e-12:
                v -= P
                m -= P * (x - a)
        V[i] = v
        M[i] = m

    # maxima and positions
    absM = [abs(v) for v in M]
    absV = [abs(v) for v in V]
    iM = int(max(range(n), key=lambda i: absM[i]))
    iV = int(max(range(n), key=lambda i: absV[i]))
    Mmax = absM[iM]
    Vmax = absV[iV]
    x_Mmax = xs[iM]
    x_Vmax = xs[iV]

    # curvature k = M/(E*I)
    k = [m / (E * I) for m in M]

    # integrate curvature -> slope with trapezoid
    K1 = [0.0] * n
    for i in range(1, n):
        K1[i] = K1[i - 1] + 0.5 * (k[i - 1] + k[i]) * dx

    # integrate K1 -> helper integral
    intK1 = [0.0] * n
    for i in range(1, n):
        intK1[i] = intK1[i - 1] + 0.5 * (K1[i - 1] + K1[i]) * dx

    # enforce y(L)=0 => theta0 = -∫K1 ds / L
    theta0 = -intK1[-1] / L
    theta = [theta0 + v for v in K1]

    # integrate slope -> deflection (y(0)=0)
    y = [0.0] * n
    for i in range(1, n):
        y[i] = y[i - 1] + 0.5 * (theta[i - 1] + theta[i]) * dx

    absy = [abs(v) for v in y]
    iy = int(max(range(n), key=lambda i: absy[i]))
    dmax = absy[iy]
    x_dmax = xs[iy]

    w_avg = Wtot / L
    w_max = max(w) if w else 0.0

    out: Dict[str, object] = {
        "Ra": Ra,
        "Rb": Rb,
        "w_avg": w_avg,
        "w_max": w_max,
        "Mmax": Mmax,
        "Vmax": Vmax,
        "dmax": dmax,
        "x_Mmax": x_Mmax,
        "x_Vmax": x_Vmax,
        "x_dmax": x_dmax,
    }
    if return_arrays:
        out.update({"xs": xs, "V": V, "M": M, "y": y})
    return out

def check_member(
    member_type: str,
    span_dir: str,
    span: float,
    pos_or_fixed: Optional[float],
    left: Optional[float],
    right: Optional[float],
    trib_width: float,
    trib_left: float,
    trib_right: float,
    load_share_model: str,
    q: float,
    point_loads: List[Tuple[float, float]],
    section: Section,
    mat: Material,
    setts: SolverSettings,
    return_arrays: bool = False,
) -> Tuple[MemberCheck, Optional[Dict[str, object]]]:
    """
    Member check (MAIN / TRANS).

    Load model:
      - TRANS: self-weight only (w_g).
      - MAIN:
          * ONEWAY : constant tributary width (b8 compatible)
          * KAMEKKO: tributary width varies along span (45deg turtle-shell model)
    """
    _, Av, Z, I = get_section_props_m(section)

    # distributed load values on nodes
    n = setts.n_div + 1
    dx = span / setts.n_div
    xs = [i * dx for i in range(n)]

    lsm = (load_share_model or "KAMEKKO").strip().upper()

    if member_type == "MAIN":
        if lsm == "KAMEKKO":
            tl = max(0.0, float(trib_left))
            tr = max(0.0, float(trib_right))

            def trib_at(x: float) -> float:
                # min(trib_side, x, L-x)
                return min(tl, x, span - x) + min(tr, x, span - x)

            w_vals = [q * trib_at(x) + section.w_g for x in xs]
        else:
            w_udl = q * trib_width + section.w_g
            w_vals = [w_udl for _ in xs]
    else:
        w_udl = section.w_g
        w_vals = [w_udl for _ in xs]

    ana = analyze_simply_supported_general(
        L=span,
        w_vals=w_vals,
        point_loads=point_loads,
        E=mat.E_kN_m2,
        I=I,
        n_div=setts.n_div,
        return_arrays=return_arrays,
    )

    Mmax = float(ana["Mmax"])
    Vmax = float(ana["Vmax"])
    dmax = float(ana["dmax"])
    w_avg = float(ana["w_avg"])
    w_max = float(ana["w_max"])
    Ra = float(ana["Ra"])
    Rb = float(ana["Rb"])
    x_Mmax = float(ana["x_Mmax"])
    x_Vmax = float(ana["x_Vmax"])
    x_dmax = float(ana["x_dmax"])

    utilM = (Mmax / Z) / mat.fb_kN_m2
    utilV = (Vmax / Av) / mat.fv_kN_m2
    utilD = dmax / (span / mat.deflection_limit)
    utilMax = max(utilM, utilV, utilD)
    ok = utilMax <= 1.0

    mc = MemberCheck(
        beam_id="",
        beam_no=0,
        member_type=member_type,
        direction=span_dir,
        pos_or_fixed=pos_or_fixed,
        left=left,
        right=right,
        span=span,
        section_rank=section.rank,
        section_name=section.name,
        w_g=section.w_g,
        w_udl=w_avg,
        w_udl_max=w_max,
        Ra=Ra,
        Rb=Rb,
        n_point=len(point_loads),
        Mmax=Mmax,
        x_Mmax=x_Mmax,
        Vmax=Vmax,
        x_Vmax=x_Vmax,
        dmax=dmax,
        x_dmax=x_dmax,
        util_M=utilM,
        util_V=utilV,
        util_d=utilD,
        util_max=utilMax,
        ok=ok,
    )
    return mc, (ana if return_arrays else None)

# -----------------------------
# Load distribution with TRANS beams
# -----------------------------
def build_system_defs(
    cfg: Config,
    direction: str,
    pitch: float,
) -> Tuple[List[BeamGeom], List[TransferDef], List[Tuple[str, str, float]]]:
    """
    Build MAIN geometries and TRANS definitions.
    Returns:
      main_beams: BeamGeom[] (point_along filled for on-beam loads only)
      transfer_defs: TransferDef[] for off-beam loads
      base_alloc_rows: allocations for on-beam and PERIM-direct loads only
                       (TRANS reactions added later after selecting TRANS section)
    """
    main_beams = build_main_beams(cfg, direction, pitch)

    # axis setup
    if direction == "X":
        pitch_axis = "Y"
        span_axis = "X"
        width = cfg.Ly
        span_L = cfg.Lx
        main_pos = [b.pos for b in main_beams]  # y
    else:
        pitch_axis = "X"
        span_axis = "Y"
        width = cfg.Lx
        span_L = cfg.Ly
        main_pos = [b.pos for b in main_beams]  # x

    # snap tol
    snap = max(cfg.snap_tol, 1e-6)

    # helper
    def add_point_to_main(beam_id: str, P: float, a: float) -> None:
        for b in main_beams:
            if b.beam_id == beam_id:
                b.point_along.append((P, a))
                return
        raise ValueError(f"Internal: main beam not found: {beam_id}")

    base_alloc: List[Tuple[str, str, float]] = []
    transfer_defs: List[TransferDef] = []
    tb_index = 0

    # support list (including PERIM as virtual supports at 0 and width)
    support_positions: List[Tuple[str, float]] = [("PERIM_L", 0.0)] + [(b.beam_id, b.pos) for b in main_beams] + [("PERIM_R", width)]
    trace("LOAD_DIST_START", "Distribute point loads", n_loads=len(cfg.loads), n_supports=len(support_positions), supports=[s[0] for s in support_positions])

    for pl in cfg.loads:
        if not (0.0 <= pl.x <= cfg.Lx) or not (0.0 <= pl.y <= cfg.Ly):
            raise ValueError(f"Point load out of range: {pl}")

        c_pitch = pl.y if pitch_axis == "Y" else pl.x
        a_span = pl.x if span_axis == "X" else pl.y

        if not (0.0 <= a_span <= span_L):
            raise ValueError(f"Point load along-span out of range: {pl}")

        # PERIM direct if on perimeter line
        if c_pitch <= snap or c_pitch >= width - snap:
            base_alloc.append((pl.load_id, "PERIM", pl.P))
            trace("LOAD_PERIM", f"{pl.load_id} on PERIM", P=pl.P, x=pl.x, y=pl.y)
            continue

        # on-beam if matches a MAIN beam position
        on_beam_id: Optional[str] = None
        for b in main_beams:
            if abs(b.pos - c_pitch) <= snap:
                on_beam_id = b.beam_id
                break

        if on_beam_id is not None:
            add_point_to_main(on_beam_id, pl.P, a_span)
            base_alloc.append((pl.load_id, on_beam_id, pl.P))
            trace("LOAD_ON_BEAM", f"{pl.load_id} on MAIN", P=pl.P, beam=on_beam_id, a_span=a_span, c_pitch=c_pitch)
            continue

        # off-beam => create TRANS beam at fixed coord = a_span? no: fixed coord is along span axis (y for Y-dir system, x for X-dir system)
        # TRANS spans along pitch axis between adjacent supports around c_pitch.
        # Find adjacent supports
        left_sup, left_pos = "PERIM", 0.0
        right_sup, right_pos = "PERIM", width

        # scan support_positions
        for (sid0, p0), (sid1, p1) in zip(support_positions[:-1], support_positions[1:]):
            if p0 - 1e-12 <= c_pitch <= p1 + 1e-12:
                # map PERIM_L/ PERIM_R to PERIM for reporting
                left_sup = "PERIM" if sid0.startswith("PERIM") else sid0
                right_sup = "PERIM" if sid1.startswith("PERIM") else sid1
                left_pos, right_pos = p0, p1
                break

        Ltb = right_pos - left_pos
        if Ltb <= 1e-9:
            # should not happen
            base_alloc.append((pl.load_id, "PERIM", pl.P))
            continue

        a_tb = c_pitch - left_pos

        tb_index += 1
        tb_id = f"TB_{pl.load_id}"
        # ensure uniqueness
        if any(t.tb_id == tb_id for t in transfer_defs):
            tb_id = f"TB_{pl.load_id}_{tb_index}"

        # TRANS span direction is pitch axis
        span_dir = pitch_axis

        transfer_defs.append(TransferDef(
            tb_id=tb_id,
            span_dir=span_dir,
            fixed_coord=a_span,  # fixed at span-axis coordinate
            left_support=left_sup,
            right_support=right_sup,
            left_pos=left_pos,
            right_pos=right_pos,
            a_tb=a_tb,
            a_on_support=a_span,
            load_id=pl.load_id,
            P=pl.P,
        ))
        trace("LOAD_OFF_BEAM", f"{pl.load_id} off MAIN -> TRANS", P=pl.P, tb_id=tb_id, span_dir=span_dir, fixed_coord=a_span, left_sup=left_sup, right_sup=right_sup, left_pos=left_pos, right_pos=right_pos, a_tb=a_tb)

    return main_beams, transfer_defs, base_alloc


# -----------------------------
# Design / selection per mode
# -----------------------------
def choose_section_for_member(
    member_type: str,
    span_dir: str,
    span: float,
    pos_or_fixed: Optional[float],
    left: Optional[float],
    right: Optional[float],
    trib_width: float,
    trib_left: float,
    trib_right: float,
    load_share_model: str,
    q: float,
    point_loads: List[Tuple[float, float]],
    sections: List[Section],
    mat: Material,
    setts: SolverSettings,
) -> Optional[MemberCheck]:
    """
    Return chosen MemberCheck (beam_id is filled by caller), or None if no feasible section.

    Selection rule (Mode-A only):
      - Among feasible sections, choose minimum member weight = w_g * span.
      - Tie-break: smaller rank (stabilizer only).
    """
    best: Optional[MemberCheck] = None
    best_weight = math.inf

    for sec in sections:
        try:
            bc, _ana = check_member(
                member_type=member_type,
                span_dir=span_dir,
                span=span,
                pos_or_fixed=pos_or_fixed,
                left=left,
                right=right,
                trib_width=trib_width,
                trib_left=trib_left,
                trib_right=trib_right,
                load_share_model=load_share_model,
                q=q,
                point_loads=point_loads,
                section=sec,
                mat=mat,
                setts=setts,
                return_arrays=False,
            )
            _dbg_add_member_trial(sec, bc, bc.ok, note="")
        except Exception as e:
            _dbg_add_member_trial(sec, None, False, note=str(e))
            trace("MEMBER_TRY_ERROR", "check_member failed", sec_rank=sec.rank, sec_name=sec.name, error=str(e))
            continue

        # keep full record even if NG
        if not bc.ok:
            continue

        w = sec.w_g * span
        if w < best_weight - 1e-12:
            best = bc
            best_weight = w
        elif abs(w - best_weight) <= 1e-12 and best is not None and sec.rank < best.section_rank:
            best = bc
            best_weight = w

    if best is None:
        trace("MEMBER_NO_FEASIBLE", "No feasible section", member_type=member_type, span_dir=span_dir, span=span, n_point=len(point_loads))
    else:
        trace("MEMBER_BEST", "Selected section", sec_rank=best.section_rank, sec_name=best.section_name, util=best.util_max)


    return best


def reactions_simply_supported(L: float, w_udl: float, P: float, a: float) -> Tuple[float, float]:
    """Return (Ra,Rb) for UDL + one point load at a from left."""
    Ra = w_udl * L / 2.0 + P * (L - a) / L
    Rb = w_udl * L / 2.0 + P * a / L
    return Ra, Rb


def solve_layout(
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    sections: List[Section],
    direction: str,
    pitch: float,
    cand_id: str,
) -> Solution:
    """
    Solve a given (direction, pitch) for Mode-A only (total weight minimum).
    - MAIN beams: per-beam section selection (minimum member weight among feasible).
    - TRANS beams: generated only when a point load is off the MAIN beam line.
    """


    # verbose trace context
    _set_dbg_context(cand_id=cand_id, direction=direction, pitch=pitch)
    trace("CAND_START", f"Start candidate {cand_id}", direction=direction, pitch=pitch, load_share_model=cfg.load_share_model)

    # fresh system definition
    main_beams, transfer_defs, base_alloc = build_system_defs(cfg, direction, pitch)
    trace("SYSTEM_BUILT", "System definitions built", n_main=len(main_beams), n_trans=len(transfer_defs), n_base_alloc=len(base_alloc))
    _dbg_add_main_geoms(cand_id, direction, pitch, main_beams)
    _dbg_add_trans_defs(cand_id, direction, pitch, transfer_defs)

    member_checks: List[MemberCheck] = []
    alloc_rows: List[Tuple[str, str, float]] = list(base_alloc)

    # support reactions to MAIN beams: map beam_id -> list[(P,a)]
    add_to_main: Dict[str, List[Tuple[float, float]]] = {}

    # 1) TRANS beams first (choose section -> reactions -> add to MAIN)
    for td in transfer_defs:
        Ltb = td.right_pos - td.left_pos
        if Ltb <= 1e-9:
            continue

        pts = [(td.P, td.a_tb)]  # original point load; self-weight included in chosen (w_udl)
        _set_dbg_context(cand_id=cand_id, direction=direction, pitch=pitch, member_id=td.tb_id, member_type="TRANS")
        trace("TRANS_START", f"Design TRANS {td.tb_id}", span=td.right_pos-td.left_pos, load_id=td.load_id, P=td.P, a_tb=td.a_tb, left=td.left_support, right=td.right_support)
        chosen = choose_section_for_member(
            member_type="TRANS",
            span_dir=td.span_dir,
            span=Ltb,
            pos_or_fixed=td.fixed_coord,
            left=td.left_pos,
            right=td.right_pos,
            trib_width=0.0,
            trib_left=0.0,
            trib_right=0.0,
            load_share_model="ONEWAY",
            q=cfg.q,
            point_loads=pts,
            sections=sections,
            mat=mat,
            setts=setts,
        )
        if chosen is None:
            return Solution(
                direction=direction,
                pitch=pitch,
                total_weight=math.inf,
                max_rank_used=10**9,
                Mmax=0.0,
                Vmax=0.0,
                dmax=0.0,
                util_max=math.inf,
                ok=False,
                ng_reason=f"No feasible TRANS section for {td.tb_id}",
                member_checks=[],
                main_geoms=main_beams,
                transfer_defs=transfer_defs,
                allocation_rows=alloc_rows,
                worst_member_id="",
            )

        Ra, Rb = chosen.Ra, chosen.Rb
        trace("TRANS_CHOSEN", f"Chosen TRANS section rank={chosen.section_rank}", sec=chosen.section_name, Ra=Ra, Rb=Rb, util=chosen.util_max)

        alloc_rows.append((td.load_id, td.left_support, Ra))
        alloc_rows.append((td.load_id, td.right_support, Rb))

        if td.left_support != "PERIM":
            add_to_main.setdefault(td.left_support, []).append((Ra, td.a_on_support))
        if td.right_support != "PERIM":
            add_to_main.setdefault(td.right_support, []).append((Rb, td.a_on_support))

        member_checks.append(replace(chosen, beam_id=td.tb_id))

    # 2) Add TRANS reactions to MAIN as point loads
    for b in main_beams:
        for P, a in add_to_main.get(b.beam_id, []):
            b.point_along.append((P, a))

    # 3) Design MAIN beams (per beam)
    for b in main_beams:
        _set_dbg_context(cand_id=cand_id, direction=direction, pitch=pitch, member_id=b.beam_id, member_type="MAIN")
        trace("MAIN_START", f"Design MAIN {b.beam_id}", span=b.span, pos=b.pos, trib_width=b.trib_width, trib_left=b.trib_left, trib_right=b.trib_right, n_point=len(b.point_along))
        chosen = choose_section_for_member(
            member_type="MAIN",
            span_dir=b.direction,
            span=b.span,
            pos_or_fixed=b.pos,
            left=None,
            right=None,
            trib_width=b.trib_width,
            trib_left=b.trib_left,
            trib_right=b.trib_right,
            load_share_model=cfg.load_share_model,
            q=cfg.q,
            point_loads=b.point_along,
            sections=sections,
            mat=mat,
            setts=setts,
        )
        if chosen is None:
            return Solution(
                direction=direction,
                pitch=pitch,
                total_weight=math.inf,
                max_rank_used=10**9,
                Mmax=0.0,
                Vmax=0.0,
                dmax=0.0,
                util_max=math.inf,
                ok=False,
                ng_reason=f"No feasible MAIN section for {b.beam_id}",
                member_checks=[],
                main_geoms=main_beams,
                transfer_defs=transfer_defs,
                allocation_rows=alloc_rows,
                worst_member_id="",
            )

        trace("MAIN_CHOSEN", f"Chosen MAIN section rank={chosen.section_rank}", sec=chosen.section_name, util=chosen.util_max, w_avg=chosen.w_udl, w_max=chosen.w_udl_max)
        member_checks.append(replace(chosen, beam_id=b.beam_id))

    # beam numbering (for plan view / mapping)
    main_ids = sorted([mc.beam_id for mc in member_checks if mc.member_type == "MAIN"])
    trans_ids = sorted([mc.beam_id for mc in member_checks if mc.member_type == "TRANS"])
    no_map: Dict[str, int] = {}
    for i, bid in enumerate(main_ids, start=1):
        no_map[bid] = i
    for i, bid in enumerate(trans_ids, start=101):
        no_map[bid] = i
    member_checks = [replace(mc, beam_no=no_map.get(mc.beam_id, 0)) for mc in member_checks]

    # aggregate
    total_weight = 0.0
    max_rank_used = 0
    Mmax = Vmax = dmax = 0.0
    util_max = 0.0
    worst_id = ""
    worst_u = -1.0

    for mc in member_checks:
        total_weight += mc.w_g * mc.span
        max_rank_used = max(max_rank_used, mc.section_rank)
        Mmax = max(Mmax, mc.Mmax)
        Vmax = max(Vmax, mc.Vmax)
        dmax = max(dmax, mc.dmax)
        util_max = max(util_max, mc.util_max)
        if mc.util_max > worst_u:
            worst_u = mc.util_max
            worst_id = mc.beam_id

    ok = util_max <= 1.0
    ng_reason = "" if ok else "Some member failed."

    # verbose outputs (tables)
    for mc in member_checks:
        _dbg_add_member_final(cand_id, direction, pitch, mc)
    _dbg_add_alloc_final(cand_id, direction, pitch, alloc_rows)
    trace("CAND_END", f"End candidate {cand_id}", ok=ok, total_weight=total_weight, util_max=util_max, max_rank=max_rank_used, worst_member=worst_id)

    return Solution(
        direction=direction,
        pitch=pitch,
        total_weight=total_weight,
        max_rank_used=max_rank_used,
        Mmax=Mmax,
        Vmax=Vmax,
        dmax=dmax,
        util_max=util_max,
        ok=ok,
        ng_reason=ng_reason,
        member_checks=sorted(member_checks, key=lambda x: (0 if x.member_type == "MAIN" else 1, x.beam_no, x.beam_id)),
        main_geoms=main_beams,
        transfer_defs=transfer_defs,
        allocation_rows=alloc_rows,
        worst_member_id=worst_id,
    )

# -----------------------------
# Optimizer (direction × pitch)
# -----------------------------
@dataclass(frozen=True)
class CandidateRow:
    direction: str
    pitch: float
    n_main: int
    n_trans: int
    max_rank_used: int
    total_weight: float
    Mmax: float
    Vmax: float
    dmax_mm: float
    util_max: float
    ok: bool
    ng_reason: str


@dataclass(frozen=True)
class CandidateSpec:
    cand_no: int
    cand_id: str
    direction: str
    pitch: float


@dataclass
class WorkerResult:
    cand_no: int
    row: CandidateRow
    sol: Solution
    trace: List[Dict[str, object]]
    dbg_trials: List[Dict[str, object]]
    dbg_main_geoms: List[Dict[str, object]]
    dbg_trans_defs: List[Dict[str, object]]
    dbg_member_final: List[Dict[str, object]]
    dbg_alloc_final: List[Dict[str, object]]


def input_stage(in_path: str) -> Tuple[Config, Material, SolverSettings, List[Section]]:
    cfg, mat, setts, sections = read_input_xlsx(in_path)

    ensure_positive("Lx", cfg.Lx)
    ensure_positive("Ly", cfg.Ly)
    ensure_positive("q", cfg.q)
    ensure_positive("E", mat.E_kN_m2)
    ensure_positive("fb", mat.fb_kN_m2)
    ensure_positive("fv", mat.fv_kN_m2)
    ensure_positive("deflection_limit", mat.deflection_limit)

    return cfg, mat, setts, sections


def _fill_section_props(sections: List[Section]) -> List[Section]:
    out: List[Section] = []
    for s in sections:
        h, b, tw, tf = s.h, s.b, s.tw, s.tf
        A_mm2, I_mm4, Z_mm3, Av_mm2 = s.A_mm2, s.I_mm4, s.Z_mm3, s.Av_mm2

        if any(v is None for v in (h, b, tw, tf)):
            dims = parse_h_section_dims(s.name)
            if dims:
                h, b, tw, tf = dims

        if any(v is None for v in (A_mm2, I_mm4, Z_mm3, Av_mm2)) and all(v is not None for v in (h, b, tw, tf)):
            A2, I2, Z2, Av2 = approx_h_section_props_mm(float(h), float(b), float(tw), float(tf))
            if A_mm2 is None:
                A_mm2 = A2
            if I_mm4 is None:
                I_mm4 = I2
            if Z_mm3 is None:
                Z_mm3 = Z2
            if Av_mm2 is None:
                Av_mm2 = Av2

        out.append(replace(
            s,
            h=h,
            b=b,
            tw=tw,
            tf=tf,
            A_mm2=A_mm2,
            I_mm4=I_mm4,
            Z_mm3=Z_mm3,
            Av_mm2=Av_mm2,
        ))
    return out


def prepare_stage(cfg: Config, sections: List[Section]) -> Tuple[List[CandidateSpec], List[Section]]:
    sections2 = _fill_section_props(sections)
    pitches = make_pitch_candidates(cfg)

    directions: List[str] = []
    if cfg.enable_x:
        directions.append("X")
    if cfg.enable_y:
        directions.append("Y")
    if not directions:
        raise ValueError("Both X and Y directions are disabled.")

    s_axis = short_side_axis(cfg.Lx, cfg.Ly)
    specs: List[CandidateSpec] = []
    cand_no = 0
    for d in directions:
        pdir = pitch_direction_of(d)
        for pitch in pitches:
            if pdir == s_axis and pitch > cfg.short_pitch_limit + 1e-9:
                continue
            cand_no += 1
            specs.append(CandidateSpec(cand_no=cand_no, cand_id=f"C{cand_no:03d}", direction=d, pitch=float(pitch)))

    return specs, sections2


def _worker_eval(args) -> WorkerResult:
    cfg, mat, setts, sections, spec = args
    _clear_debug()

    sol = solve_layout(cfg, mat, setts, sections, spec.direction, spec.pitch, cand_id=spec.cand_id)

    n_main = len([m for m in sol.member_checks if m.member_type == "MAIN"])
    n_trans = len([m for m in sol.member_checks if m.member_type == "TRANS"])
    row = CandidateRow(
        direction=spec.direction,
        pitch=spec.pitch,
        n_main=n_main,
        n_trans=n_trans,
        max_rank_used=sol.max_rank_used if sol.ok else 0,
        total_weight=sol.total_weight,
        Mmax=sol.Mmax,
        Vmax=sol.Vmax,
        dmax_mm=sol.dmax * 1000.0,
        util_max=sol.util_max,
        ok=sol.ok,
        ng_reason=sol.ng_reason,
    )

    return WorkerResult(
        cand_no=spec.cand_no,
        row=row,
        sol=sol,
        trace=list(_TRACE),
        dbg_trials=list(_DBG_MEMBER_TRIALS),
        dbg_main_geoms=list(_DBG_MAIN_GEOMS),
        dbg_trans_defs=list(_DBG_TRANS_DEFS),
        dbg_member_final=list(_DBG_MEMBER_FINAL),
        dbg_alloc_final=list(_DBG_ALLOC_FINAL),
    )


def parallel_stage(
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    sections: List[Section],
    specs: List[CandidateSpec],
    max_workers: Optional[int] = None,
) -> List[WorkerResult]:
    if not specs:
        return []

    if max_workers is None:
        cpu = os.cpu_count() or 1
        max_workers = max(1, min(cpu, 8))
    max_workers = max(1, min(int(max_workers), len(specs)))

    tasks = [(cfg, mat, setts, sections, s) for s in specs]
    results: List[WorkerResult] = []
    done = 0
    total = len(tasks)

    with ProcessPoolExecutor(max_workers=max_workers) as ex:
        futs = [ex.submit(_worker_eval, t) for t in tasks]
        for fut in as_completed(futs):
            results.append(fut.result())
            done += 1
            if done % max(1, total // 20) == 0 or done == total:
                print(f"[Progress] {done}/{total} evaluated")

    results.sort(key=lambda x: x.cand_no)
    return results


def postprocess_stage(worker_results: List[WorkerResult]) -> Tuple[List[CandidateRow], Optional[Solution]]:
    _clear_debug()

    merged_trace: List[Dict[str, object]] = []
    for wr in worker_results:
        merged_trace.extend(wr.trace)
        _DBG_MEMBER_TRIALS.extend(wr.dbg_trials)
        _DBG_MAIN_GEOMS.extend(wr.dbg_main_geoms)
        _DBG_TRANS_DEFS.extend(wr.dbg_trans_defs)
        _DBG_MEMBER_FINAL.extend(wr.dbg_member_final)
        _DBG_ALLOC_FINAL.extend(wr.dbg_alloc_final)

    for i, e in enumerate(merged_trace, start=1):
        e["seq"] = i
    _TRACE.extend(merged_trace)

    cand_rows = [wr.row for wr in worker_results]
    best: Optional[Solution] = None
    for wr in worker_results:
        sol = wr.sol
        if not sol.ok:
            continue
        if best is None:
            best = sol
            continue
        if sol.total_weight < best.total_weight - 1e-12:
            best = sol
            continue
        if abs(sol.total_weight - best.total_weight) <= 1e-12 and sol.util_max < best.util_max - 1e-12:
            best = sol

    return cand_rows, best


def output_stage(
    in_path: str,
    out_path: str,
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    sections: List[Section],
    cand_rows: List[CandidateRow],
    best: Optional[Solution],
) -> None:
    write_result_xlsx(in_path, out_path, cfg, mat, setts, sections, cand_rows, best)


def optimize(
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    sections: List[Section],
    max_workers: Optional[int] = None,
) -> Tuple[List[CandidateRow], Optional[Solution]]:
    specs, sections2 = prepare_stage(cfg, sections)
    worker_results = parallel_stage(cfg, mat, setts, sections2, specs, max_workers=max_workers)
    return postprocess_stage(worker_results)


def run(in_path: str, out_path: str, max_workers: Optional[int] = None) -> None:
    _clear_debug()
    cfg, mat, setts, sections = input_stage(in_path)
    specs, sections2 = prepare_stage(cfg, sections)
    worker_results = parallel_stage(cfg, mat, setts, sections2, specs, max_workers=max_workers)
    cand_rows, best = postprocess_stage(worker_results)
    trace(
        "RUN_OPT_DONE",
        "Optimization finished",
        n_candidates=len(cand_rows),
        best_ok=(best.ok if best else False),
        best_direction=(best.direction if best else None),
        best_pitch=(best.pitch if best else None),
    )
    output_stage(in_path, out_path, cfg, mat, setts, sections2, cand_rows, best)


# -----------------------------
# Excel I/O
# -----------------------------
def read_cell(ws, addr: str):
    return ws[addr].value


def read_bool(ws, addr: str, default: bool = False) -> bool:
    v = read_cell(ws, addr)
    if v is None:
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "y", "on"):
        return True
    if s in ("false", "0", "no", "n", "off"):
        return False
    return default


def read_float(ws, addr: str, name: str, default: Optional[float] = None) -> float:
    v = read_cell(ws, addr)
    if v is None or str(v).strip() == "":
        if default is None:
            raise ValueError(f"Missing required value: {name} at {addr}")
        return float(default)
    try:
        return float(v)
    except Exception as e:
        raise ValueError(f"Invalid number for {name} at {addr}: {v}") from e


def _norm(s: object) -> str:
    if s is None:
        return ""
    return str(s).strip().replace("\u3000", " ").strip().lower()


def _norm_header(s: object) -> str:
    t = _norm(s)
    t = re.sub(r"\[.*?\]", "", t)  # drop [unit]
    t = t.replace(" ", "").replace("-", "").replace("/", "")
    t = re.sub(r"[^0-9a-z_]", "", t)
    return t


def _to_bool(v: object, default: bool = False) -> bool:
    if v is None or str(v).strip() == "":
        return default
    if isinstance(v, bool):
        return v
    s = _norm(v)
    if s in ("true", "1", "yes", "y", "on", "ok"):
        return True
    if s in ("false", "0", "no", "n", "off", "ng"):
        return False
    return default


def find_value_right_of_label(
    ws,
    label_regex: str,
    *,
    default: Optional[object] = None,
    required: bool = False,
    search_rows: int = 500,
    search_cols: Tuple[int, ...] = (1, 2, 3, 4),
    value_offset: int = 1,
) -> object:
    """
    行位置が変わっても読めるよう、ラベル文字列を正規表現で探索して右隣セルの値を返す。
    """
    rgx = re.compile(label_regex, flags=re.IGNORECASE)
    for r in range(1, search_rows + 1):
        for c in search_cols:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and rgx.search(v):
                return ws.cell(row=r, column=c + value_offset).value
    if required:
        raise ValueError(f"Required label not found: /{label_regex}/")
    return default


def find_cell(ws, label_regex: str, *, search_rows: int = 600, search_cols: int = 60) -> Optional[Tuple[int, int]]:
    rgx = re.compile(label_regex, flags=re.IGNORECASE)
    for r in range(1, search_rows + 1):
        for c in range(1, search_cols + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and rgx.search(v):
                return r, c
    return None


def find_table_header(
    ws,
    required_aliases: Dict[str, List[str]],
    *,
    search_rows: int = 600,
    search_cols: int = 60
) -> Optional[Tuple[int, Dict[str, int], Dict[str, int]]]:
    """
    見出し行を探索して、必須列が揃う行を返す。
    Returns: (header_row, col_map_canonical, row_map_normheader->col)
    """
    for r in range(1, search_rows + 1):
        row_map: Dict[str, int] = {}
        for c in range(1, search_cols + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or str(v).strip() == "":
                continue
            h = _norm_header(v)
            if h:
                row_map[h] = c

        col_map: Dict[str, int] = {}
        ok = True
        for key, aliases in required_aliases.items():
            found = None
            for a in aliases:
                if a in row_map:
                    found = row_map[a]
                    break
            if found is None:
                ok = False
                break
            col_map[key] = found

        if ok:
            return r, col_map, row_map
    return None


def read_point_loads_table(ws) -> List[PointLoad]:
    hdr = find_table_header(ws, {
        "load_id": ["loadid", "id"],
        "p": ["p", "pkn", "pkn", "pkn"],  # normalized: "P [kN]" -> "p"
        "x": ["x", "xm"],
        "y": ["y", "ym"],
    }, search_rows=300, search_cols=20)
    if hdr is None:
        return []
    header_row, cols, _ = hdr

    loads: List[PointLoad] = []
    r = header_row + 1
    while True:
        lid = ws.cell(row=r, column=cols["load_id"]).value
        if lid is None or str(lid).strip() == "":
            break
        P = ws.cell(row=r, column=cols["p"]).value
        x = ws.cell(row=r, column=cols["x"]).value
        y = ws.cell(row=r, column=cols["y"]).value
        try:
            loads.append(PointLoad(load_id=str(lid).strip(), P=float(P), x=float(x), y=float(y)))
        except Exception as e:
            raise ValueError(f"Invalid point load row {r}: {lid},{P},{x},{y}") from e
        r += 1
    return loads


def read_sections_table(ws) -> List[Section]:
    hdr = find_table_header(ws, {
        "rank": ["rank"],
        "name": ["sectionname", "section", "name"],
        "wg": ["w_g", "wg"],
    }, search_rows=700, search_cols=40)
    if hdr is None:
        raise ValueError("Section Candidates table not found (need Rank, SectionName, w_g).")
    header_row, cols_req, row_map = hdr

    def col_of(*aliases: str) -> Optional[int]:
        for a in aliases:
            if a in row_map:
                return row_map[a]
        return None

    c_rank = cols_req["rank"]
    c_name = cols_req["name"]
    c_wg = cols_req["wg"]

    c_h = col_of("h", "hmm")
    c_b = col_of("b", "bmm")
    c_tw = col_of("tw", "twmm")
    c_tf = col_of("tf", "tfmm")
    c_A = col_of("a", "amm2")
    c_Z = col_of("z", "zmm3")
    c_I = col_of("i", "imm4")
    c_Av = col_of("av", "avmm2")

    sections: List[Section] = []
    r = header_row + 1
    while True:
        rk = ws.cell(row=r, column=c_rank).value
        nm = ws.cell(row=r, column=c_name).value
        wg = ws.cell(row=r, column=c_wg).value
        if (rk is None or str(rk).strip() == "") and (nm is None or str(nm).strip() == ""):
            break
        if nm is None or str(nm).strip() == "":
            raise ValueError(f"SectionName missing at row {r}")
        if wg is None or str(wg).strip() == "":
            raise ValueError(f"w_g missing at row {r}")
        try:
            rank = int(float(rk))
        except Exception as e:
            raise ValueError(f"Rank invalid at row {r}: {rk}") from e

        def opt(col: Optional[int]) -> Optional[float]:
            if col is None:
                return None
            v = ws.cell(row=r, column=col).value
            if v is None or str(v).strip() == "":
                return None
            return float(v)

        sections.append(Section(
            rank=rank,
            name=str(nm).strip(),
            w_g=float(wg),
            h=opt(c_h),
            b=opt(c_b),
            tw=opt(c_tw),
            tf=opt(c_tf),
            A_mm2=opt(c_A),
            Z_mm3=opt(c_Z),
            I_mm4=opt(c_I),
            Av_mm2=opt(c_Av),
        ))
        r += 1

    sections = sorted(sections, key=lambda s: s.rank)
    if not sections:
        raise ValueError("No sections provided.")
    return sections


def read_input_xlsx(path: str) -> Tuple[Config, Material, SolverSettings, List[Section]]:
    wb = load_workbook(path, data_only=True)
    if "INPUT" not in wb.sheetnames:
        raise ValueError('Sheet "INPUT" not found.')
    ws = wb["INPUT"]

    # --- 基本寸法・荷重（ラベル検索） ---
    Lx = float(find_value_right_of_label(ws, r"^\s*lx\b", required=True))
    Ly = float(find_value_right_of_label(ws, r"^\s*ly\b", required=True))
    q = float(find_value_right_of_label(ws, r"^\s*q\b", required=True))

    # --- load sharing model (optional) ---
    lsm_raw = find_value_right_of_label(ws, r"load\s*sharing\s*model", default="KAMEKKO")
    load_share_model = str(lsm_raw).strip().upper() if lsm_raw is not None else "KAMEKKO"
    if load_share_model not in ("KAMEKKO", "ONEWAY"):
        load_share_model = "KAMEKKO"

    # --- 点荷重（テーブル探索） ---
    loads = read_point_loads_table(ws)

    # --- レイアウト候補 ---
    enable_x = _to_bool(find_value_right_of_label(ws, r"enable\s*x", default=True), True)
    enable_y = _to_bool(find_value_right_of_label(ws, r"enable\s*y", default=True), True)
    edge_beams = _to_bool(find_value_right_of_label(ws, r"place\s*beams.*edges", default=True), True)

    rule_raw = find_value_right_of_label(ws, r"point\s*load.*rule", default=2)
    try:
        load_rule = int(float(rule_raw))
    except Exception:
        load_rule = 2

    pitch_start = float(find_value_right_of_label(ws, r"pitch\s*start", default=1.5))
    pitch_end = float(find_value_right_of_label(ws, r"pitch\s*end", default=3.0))
    pitch_step = float(find_value_right_of_label(ws, r"pitch\s*step", default=0.5))
    short_pitch_limit = float(find_value_right_of_label(ws, r"short-?side.*pitch.*limit", default=3.0))

    # 任意：ピッチ候補リスト（見出し "Pitch [m]" を探索）
    pitch_list: List[float] = []
    pitch_hdr = find_cell(ws, r"^\s*pitch\s*\[m\]\s*$", search_rows=500, search_cols=30)
    if pitch_hdr is not None:
        hr, hc = pitch_hdr
        if hc != 1:
            data_col = None
            for cand in (hc, hc + 1, hc + 2):
                v0 = ws.cell(row=hr + 1, column=cand).value
                try:
                    float(v0)
                    data_col = cand
                    break
                except Exception:
                    continue
            if data_col is not None:
                rr = hr + 1
                while True:
                    v = ws.cell(row=rr, column=data_col).value
                    if v is None or str(v).strip() == "":
                        break
                    pitch_list.append(float(v))
                    rr += 1

    # 任意：snap tolerance（無ければ 1e-3 m）
    snap_raw = find_value_right_of_label(ws, r"snap.*tol", default=None)
    snap_tol = 1e-3 if (snap_raw is None or str(snap_raw).strip() == "") else float(snap_raw)
    snap_tol = max(float(snap_tol), 1e-6)

    # --- 材料・許容 ---
    E_val = float(find_value_right_of_label(ws, r"^\s*e\s*$", default=205000.0))
    e_cell = find_cell(ws, r"^\s*e\s*$", search_rows=400, search_cols=8)
    E_unit = ws.cell(row=e_cell[0], column=e_cell[1] + 2).value if e_cell else "N/mm2"
    E_unit_s = _norm(E_unit)

    if E_unit_s in ("n/mm2", "n/mm^2", "mpa", ""):
        E_kN_m2 = nmm2_to_kN_m2(E_val)
    elif E_unit_s in ("kn/m2", "kn/m^2"):
        E_kN_m2 = float(E_val)
    else:
        raise ValueError('E_unit must be "N/mm2" or "kN/m2".')

    fb = float(find_value_right_of_label(ws, r"^\s*fb\b", default=165.0))
    fv = float(find_value_right_of_label(ws, r"^\s*fv\b", default=95.0))
    defl_lim = float(find_value_right_of_label(ws, r"deflection_limit", default=360.0))

    mat = Material(
        E_kN_m2=E_kN_m2,
        fb_kN_m2=nmm2_to_kN_m2(fb),
        fv_kN_m2=nmm2_to_kN_m2(fv),
        deflection_limit=defl_lim
    )

    # --- ソルバ設定 ---
    tol = float(find_value_right_of_label(ws, r"^\s*tol\b", default=1e-6))
    max_iter = int(float(find_value_right_of_label(ws, r"max_iter", default=5)))
    n_div = int(float(find_value_right_of_label(ws, r"n_div", default=2000)))
    setts = SolverSettings(tol=tol, max_iter=max_iter, n_div=n_div)

    # --- 断面候補 ---
    sections = read_sections_table(ws)

    cfg = Config(
        Lx=Lx, Ly=Ly, q=q, loads=loads, load_share_model=load_share_model,
        enable_x=enable_x, enable_y=enable_y,
        edge_beams=edge_beams, load_rule=load_rule,
        pitch_start=pitch_start, pitch_end=pitch_end, pitch_step=pitch_step,
        pitch_list=pitch_list,
        short_pitch_limit=short_pitch_limit,
        snap_tol=snap_tol
    )

    return cfg, mat, setts, sections


# -----------------------------
# Excel output helpers
# -----------------------------
def set_col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def write_header(ws, row: int, col: int, text: str) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor="DDDDDD")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# -----------------------------
# RESULT sheet
# -----------------------------
def write_solution_summary(ws, row0: int, title: str, sol: Optional[Solution]) -> int:
    ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
    headers = ["Dir", "Pitch[m]", "MaxRankUsed", "TotalWeight[kN]", "Mmax[kN*m]", "Vmax[kN]", "dmax[mm]", "utilMax", "OK/NG"]
    for j, h in enumerate(headers, start=1):
        write_header(ws, row0 + 1, j, h)
    if sol is None or not sol.ok:
        ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
        return row0 + 4
    vals = [
        sol.direction,
        sol.pitch,
        sol.max_rank_used,
        round(sol.total_weight, 6),
        round(sol.Mmax, 6),
        round(sol.Vmax, 6),
        round(sol.dmax * 1000.0, 6),
        round(sol.util_max, 6),
        "OK",
    ]
    for j, v in enumerate(vals, start=1):
        ws.cell(row=row0 + 2, column=j, value=v)
    return row0 + 4


def write_candidates_table(ws, row0: int, rows: List[CandidateRow]) -> int:
    ws.cell(row=row0, column=1, value="All candidates (Mode-A only)").font = Font(bold=True)
    row0 += 1
    headers = ["Dir", "Pitch[m]", "N_MAIN", "N_TRANS", "MaxRankUsed", "TotalWeight[kN]", "Mmax[kN*m]", "Vmax[kN]", "dmax[mm]", "utilMax", "OK/NG"]
    for j, h in enumerate(headers, start=1):
        write_header(ws, row0, j, h)
    row0 += 1

    rows_sorted = sorted(rows, key=lambda r: (not r.ok, r.total_weight if r.ok else math.inf, r.max_rank_used if r.ok else 10**9, r.direction, r.pitch))
    for r in rows_sorted:
        out = [
            r.direction, r.pitch, r.n_main, r.n_trans,
            r.max_rank_used if r.ok else "",
            round(r.total_weight, 6) if math.isfinite(r.total_weight) else "",
            round(r.Mmax, 6),
            round(r.Vmax, 6),
            round(r.dmax_mm, 6),
            round(r.util_max, 6) if math.isfinite(r.util_max) else "",
            "OK" if r.ok else f"NG: {r.ng_reason}",
        ]
        for j, v in enumerate(out, start=1):
            ws.cell(row=row0, column=j, value=v)
        row0 += 1

    return row0 + 2


def write_member_checks(ws, row0: int, title: str, sol: Optional[Solution]) -> int:
    ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
    headers = [
        "No", "BeamID", "Type", "Dir", "Pos/Fixed[m]", "Left[m]", "Right[m]", "Span[m]",
        "Rank", "Section", "w_g[kN/m]", "w_avg[kN/m]", "w_max[kN/m]",
        "Ra[kN]", "Rb[kN]", "Npoint",
        "Mmax[kN*m]", "x_Mmax[m]", "Vmax[kN]", "x_Vmax[m]", "dmax[mm]", "x_dmax[m]",
        "util_M", "util_V", "util_d", "util_max", "OK/NG"
    ]
    for j, h in enumerate(headers, start=1):
        write_header(ws, row0 + 1, j, h)
    if sol is None or not sol.ok:
        ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
        return row0 + 4

    r = row0 + 2
    # sort: MAIN then TRANS, by beam_no
    checks = sorted(sol.member_checks, key=lambda x: (0 if x.member_type == "MAIN" else 1, x.beam_no, x.beam_id))

    for mc in checks:
        out = [
            mc.beam_no,
            mc.beam_id,
            mc.member_type,
            mc.direction,
            mc.pos_or_fixed if mc.pos_or_fixed is not None else "",
            mc.left if mc.left is not None else "",
            mc.right if mc.right is not None else "",
            round(mc.span, 6),
            mc.section_rank,
            mc.section_name,
            round(mc.w_g, 6),
            round(mc.w_udl, 6),
            round(mc.w_udl_max, 6),
            round(mc.Ra, 6),
            round(mc.Rb, 6),
            mc.n_point,
            round(mc.Mmax, 6),
            round(mc.x_Mmax, 6),
            round(mc.Vmax, 6),
            round(mc.x_Vmax, 6),
            round(mc.dmax * 1000.0, 6),
            round(mc.x_dmax, 6),
            round(mc.util_M, 6),
            round(mc.util_V, 6),
            round(mc.util_d, 6),
            round(mc.util_max, 6),
            "OK" if mc.ok else "NG",
        ]
        for j, v in enumerate(out, start=1):
            ws.cell(row=r, column=j, value=v)
        r += 1

    return r + 1

def write_allocations(ws, row0: int, title: str, sol: Optional[Solution]) -> int:
    ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
    write_header(ws, row0 + 1, 1, "LoadID")
    write_header(ws, row0 + 1, 2, "SupportNo")
    write_header(ws, row0 + 1, 3, "SupportID")
    write_header(ws, row0 + 1, 4, "AllocatedP[kN]")
    if sol is None or not sol.ok:
        ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
        return row0 + 4

    no_map = {m.beam_id: m.beam_no for m in sol.member_checks}

    r = row0 + 2
    for lid, sid, p in sol.allocation_rows:
        ws.cell(row=r, column=1, value=lid)
        ws.cell(row=r, column=2, value=(0 if sid == "PERIM" else no_map.get(str(sid), "")))
        ws.cell(row=r, column=3, value=sid)
        ws.cell(row=r, column=4, value=round(p, 6))
        r += 1
    return r + 1

def write_member_list(ws, row0: int, title: str, sol: Optional[Solution]) -> int:
    ws.cell(row=row0, column=1, value=title).font = Font(bold=True)
    headers = ["MemberType", "Rank", "Section", "Count", "TotalLen[m]", "w_g[kN/m]", "Weight[kN]"]
    for j, h in enumerate(headers, start=1):
        write_header(ws, row0 + 1, j, h)
    if sol is None or not sol.ok:
        ws.cell(row=row0 + 2, column=1, value="NO FEASIBLE SOLUTION")
        return row0 + 4

    # aggregate by (type, section)
    agg: Dict[Tuple[str, int, str, float], Tuple[int, float, float]] = {}
    for mc in sol.member_checks:
        key = (mc.member_type, mc.section_rank, mc.section_name, mc.w_g)
        cnt, Ltot, W = agg.get(key, (0, 0.0, 0.0))
        agg[key] = (cnt + 1, Ltot + mc.span, W + mc.w_g * mc.span)

    r = row0 + 2
    total_w = 0.0
    for (mtype, rk, name, wg), (cnt, Ltot, W) in sorted(agg.items(), key=lambda x: (x[0][0], x[0][1], x[0][2])):
        ws.cell(row=r, column=1, value=mtype)
        ws.cell(row=r, column=2, value=rk)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=cnt)
        ws.cell(row=r, column=5, value=round(Ltot, 6))
        ws.cell(row=r, column=6, value=round(wg, 6))
        ws.cell(row=r, column=7, value=round(W, 6))
        total_w += W
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=7, value=round(total_w, 6)).font = Font(bold=True)
    return r + 2


def write_result_xlsx(
    in_path: str,
    out_path: str,
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    sections: List[Section],
    cand_rows: List[CandidateRow],
    best: Optional[Solution],
) -> None:
    wb = load_workbook(in_path)
    if "RESULT" in wb.sheetnames:
        del wb["RESULT"]
    ws = wb.create_sheet("RESULT")

    # column widths
    widths = [8, 8, 10, 10, 10, 18, 12, 12, 12, 12, 10, 28, 12, 12, 12, 10, 10, 10, 10, 10]
    for i, w in enumerate(widths, start=1):
        set_col_width(ws, i, w)

    r = 1
    ws.cell(row=r, column=1, value="Secondary Beam Optimization RESULT").font = Font(bold=True, size=14)
    r += 2
    ws.cell(row=r, column=1, value="Notes: Units are kN, m. E/fb/fv are converted from N/mm2 to kN/m2. Mode-B is removed.").font = Font(italic=True)
    r += 2

    r = write_solution_summary(ws, r, "Mode-A best (total weight minimum)", best)
    r = write_candidates_table(ws, r, cand_rows)

    r = write_member_checks(ws, r, "Beam checks (per member) - best solution", best)
    r = write_allocations(ws, r, "Load allocation to supports (includes TRANS reactions) - best solution", best)
    r = write_member_list(ws, r, "Member list - best solution", best)

    r += 1
    ws.cell(row=r, column=1, value="PLAN view is written to sheet 'LAYOUT' (cells only).").font = Font(italic=True)

    # layout sheet
    write_layout_sheet(wb, cfg, best)

    # verbose / debug sheets
    write_verbose_sheets(wb, cfg, mat, setts, sections, cand_rows, best)

    wb.save(out_path)


# -----------------------------
# LAYOUT (plan only)
# -----------------------------
def write_layout_sheet(wb, cfg: Config, best: Optional[Solution]) -> None:
    if "LAYOUT" in wb.sheetnames:
        del wb["LAYOUT"]
    ws = wb.create_sheet("LAYOUT")

    top = 1
    left = 1
    ws.cell(row=top, column=left, value="LAYOUT / PLAN VIEW (cells only)").font = Font(bold=True, size=14)

    top += 2
    if best and best.ok:
        top = draw_plan(ws, top, left, cfg, best, title="Mode-A best (total weight minimum)")
        top += 2
    else:
        ws.cell(row=top, column=left, value="NO FEASIBLE SOLUTION").font = Font(bold=True)


def draw_plan(ws, top: int, left: int, cfg: Config, sol: Solution, title: str) -> int:
    """
    Plan-only sheet (no tables). Square cells are applied only to the grid area.
    - MAIN beams: light blue
    - TRANS beams: yellow
    - Worst member: red fill
    - Point load: red ●ID at actual position
    - Allocation marker: blue ▲ID on receiving support line
    """
    # grid size (m per cell)
    grid = 0.25
    nx = int(round(cfg.Lx / grid)) + 1
    ny = int(round(cfg.Ly / grid)) + 1

    # styles
    fill_main = PatternFill("solid", fgColor="DDEBF7")
    fill_trans = PatternFill("solid", fgColor="FFF2CC")
    fill_worst = PatternFill("solid", fgColor="FF0000")  # red
    fill_blank = PatternFill("solid", fgColor="FFFFFF")

    font_small = Font(size=9)
    font_bold = Font(bold=True)

    border_thick = Border(left=Side(style="thick"), right=Side(style="thick"),
                          top=Side(style="thick"), bottom=Side(style="thick"))
    border_grid = Border(left=Side(style="thin", color="D9D9D9"),
                         right=Side(style="thin", color="D9D9D9"),
                         top=Side(style="thin", color="D9D9D9"),
                         bottom=Side(style="thin", color="D9D9D9"))

    # coordinate mapping
    grid_left = left + 2
    grid_top = top + 2

    def x_to_c(xm: float) -> int:
        return grid_left + int(round(xm / grid))

    def y_to_r(ym: float) -> int:
        return grid_top + int(round(ym / grid))

    # Title lines
    ws.cell(row=top, column=left, value=title).font = Font(bold=True)
    ws.cell(row=top + 1, column=left, value=f"PLAN: {sol.direction}-dir, pitch={sol.pitch} m, totalW={sol.total_weight:.3f} kN, maxR={sol.max_rank_used}").font = font_small

    # Format square cells only for grid area (+ axis label margins)
    col_w = 2.0   # about square with default row height
    row_h = 12.0
    for c in range(grid_left, grid_left + nx):
        ws.column_dimensions[get_column_letter(c)].width = col_w
    for r in range(grid_top, grid_top + ny):
        ws.row_dimensions[r].height = row_h

    # clear and grid
    for r in range(grid_top, grid_top + ny):
        for c in range(grid_left, grid_left + nx):
            cell = ws.cell(row=r, column=c, value="")
            cell.fill = fill_blank
            cell.border = border_grid
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # thick boundary
    for c in range(grid_left, grid_left + nx):
        ws.cell(row=grid_top, column=c).border = border_thick
        ws.cell(row=grid_top + ny - 1, column=c).border = border_thick
    for r in range(grid_top, grid_top + ny):
        ws.cell(row=r, column=grid_left).border = border_thick
        ws.cell(row=r, column=grid_left + nx - 1).border = border_thick

    # axis labels (integer meters only)
    ws.cell(row=grid_top - 1, column=grid_left, value="y[m]→x[m]").font = font_small
    for xm in range(0, int(math.floor(cfg.Lx)) + 1):
        c = x_to_c(float(xm))
        if grid_left <= c < grid_left + nx:
            ws.cell(row=grid_top - 1, column=c, value=str(xm)).font = font_small

    for ym in range(0, int(math.floor(cfg.Ly)) + 1):
        r0 = y_to_r(float(ym))
        if grid_top <= r0 < grid_top + ny:
            ws.cell(row=r0, column=grid_left - 1, value=str(ym)).font = font_small

    # build lookup of checks by id
    chk = {m.beam_id: m for m in sol.member_checks}

    # Draw MAIN beams
    # MAIN beams stored as member_checks with member_type=MAIN => need their direction and pos
    for mc in sol.member_checks:
        if mc.member_type != "MAIN":
            continue
        pos = float(mc.pos_or_fixed) if mc.pos_or_fixed is not None else 0.0
        is_worst = (mc.beam_id == sol.worst_member_id)

        if sol.direction == "Y":
            # MAIN beams are vertical (span Y), placed at x=pos
            c = x_to_c(pos)
            for r in range(grid_top, grid_top + ny):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill_worst if is_worst else fill_main
        else:
            # MAIN beams are horizontal (span X), placed at y=pos
            r0 = y_to_r(pos)
            for c in range(grid_left, grid_left + nx):
                cell = ws.cell(row=r0, column=c)
                cell.fill = fill_worst if is_worst else fill_main

        # label (merged cell to avoid vertical text)
        label = f"{mc.beam_no}"
        if sol.direction == "Y":
            c = x_to_c(pos)
            c0 = max(grid_left, c - 1)
            c1 = min(grid_left + nx - 1, c + 1)
            ws.merge_cells(start_row=grid_top - 2, start_column=c0, end_row=grid_top - 2, end_column=c1)
            lc = ws.cell(row=grid_top - 2, column=c0, value=label)
        else:
            r0 = y_to_r(pos)
            # label to the left of the beam row
            lc = ws.cell(row=r0, column=grid_left - 2, value=label)
        lc.font = font_small
        lc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Draw TRANS beams (span_dir is pitch axis; fixed_coord is perpendicular)
    for td in sol.transfer_defs:
        mc = chk.get(td.tb_id)
        if mc is None:
            continue
        is_worst = (td.tb_id == sol.worst_member_id)

        if td.span_dir == "X":
            # horizontal trans at y=fixed_coord, x from left_pos to right_pos
            r0 = y_to_r(td.fixed_coord)
            c0 = x_to_c(td.left_pos)
            c1 = x_to_c(td.right_pos)
            for c in range(min(c0, c1), max(c0, c1) + 1):
                cell = ws.cell(row=r0, column=c)
                cell.fill = fill_worst if is_worst else fill_trans
        else:
            # vertical trans at x=fixed_coord
            c0 = x_to_c(td.fixed_coord)
            r0 = y_to_r(td.left_pos)
            r1 = y_to_r(td.right_pos)
            for r in range(min(r0, r1), max(r0, r1) + 1):
                cell = ws.cell(row=r, column=c0)
                cell.fill = fill_worst if is_worst else fill_trans

        # label near mid span
        midx = 0.5 * (td.left_pos + td.right_pos) if td.span_dir == "X" else td.fixed_coord
        midy = td.fixed_coord if td.span_dir == "X" else 0.5 * (td.left_pos + td.right_pos)
        u = chk.get(td.tb_id).util_max if td.tb_id in chk else None
        label = f"{mc.beam_no}"

        # place label with a small merged area to improve legibility
        if td.span_dir == "X":
            rr = max(grid_top + 1, min(grid_top + ny, y_to_r(midy) - 1))
            cc0 = x_to_c(midx)
            c1 = max(grid_left + 1, cc0 - 1)
            c2 = min(grid_left + nx, cc0 + 1)
            if c2 > c1:
                ws.merge_cells(start_row=rr, start_column=c1, end_row=rr, end_column=c2)
            lc = ws.cell(row=rr, column=c1, value=label)
        else:
            cc = max(grid_left + 1, min(grid_left + nx, x_to_c(midx) + 1))
            rr0 = y_to_r(midy)
            r1 = max(grid_top + 1, rr0 - 1)
            r2 = min(grid_top + ny, rr0 + 1)
            if r2 > r1:
                ws.merge_cells(start_row=r1, start_column=cc, end_row=r2, end_column=cc)
            lc = ws.cell(row=r1, column=cc, value=label)

        lc.font = Font(size=8, bold=True)
        lc.alignment = Alignment(horizontal="center", vertical="center")

    # Point loads and allocation markers
    # Point load marker ● at actual coordinate
    # Allocation marker ▲ on receiving supports (MAIN beam or PERIM)
    # We'll use allocation_rows + original load coordinates
    load_map = {pl.load_id: pl for pl in cfg.loads}

        # mark actual load and allocation marker
    # - If actual (●) and allocation (▲) are mapped to the same cell, merge into one cell, e.g. "●▲P4".
    marks: Dict[Tuple[int, int], Dict[str, set]] = {}

    def add_mark(rr: int, cc: int, lid: str, kind: str) -> None:
        if rr < grid_top + 1 or rr > grid_top + ny:
            return
        if cc < grid_left + 1 or cc > grid_left + nx:
            return
        by_id = marks.setdefault((rr, cc), {})
        kinds = by_id.setdefault(lid, set())
        kinds.add(kind)

    # (1) actual point load location
    for pl in cfg.loads:
        rr = y_to_r(pl.y)
        cc = x_to_c(pl.x)
        add_mark(rr, cc, str(pl.load_id), "C")  # Circle

    # (2) allocation markers (▲) on receiving supports (including TRANS supports)
    for lid, sup_id, _P in sol.allocation_rows:
        if sup_id == "PERIM":
            # outside boundary beam; show allocation marker at the actual load cell instead
            pl = next((x for x in cfg.loads if str(x.load_id) == str(lid)), None)
            if pl is not None:
                rr = y_to_r(pl.y)
                cc = x_to_c(pl.x)
                add_mark(rr, cc, str(lid), "T")
            continue

        sup_mc = chk.get(sup_id)
        if sup_mc is None:
            continue

        pl = next((x for x in cfg.loads if str(x.load_id) == str(lid)), None)
        if pl is None:
            continue

        if sup_mc.direction == "Y" and sup_mc.pos_or_fixed is not None:
            rr = y_to_r(pl.y)
            cc = x_to_c(float(sup_mc.pos_or_fixed))
            add_mark(rr, cc, str(lid), "T")
        elif sup_mc.direction == "X" and sup_mc.pos_or_fixed is not None:
            rr = y_to_r(float(sup_mc.pos_or_fixed))
            cc = x_to_c(pl.x)
            add_mark(rr, cc, str(lid), "T")

    # write marks to sheet
    for (rr, cc), by_id in marks.items():
        parts: List[str] = []
        has_circle = False
        for lid, kinds in sorted(by_id.items(), key=lambda x: str(x[0])):
            if "C" in kinds and "T" in kinds:
                parts.append(f"◎{lid}")
                has_circle = True
            elif "C" in kinds:
                parts.append(f"●{lid}")
                has_circle = True
            else:
                parts.append(f"▲{lid}")
        val = "/".join(parts)
        cell = ws.cell(row=rr, column=cc, value=val)
        cell.font = Font(color=("C00000" if has_circle else "1F4E79"), bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center")
# Legend (below grid)
    legend_r = grid_top + ny + 1
    ws.cell(row=legend_r, column=left, value="Legend:").font = Font(bold=True)
    ws.cell(row=legend_r + 1, column=left, value="MAIN beam: light-blue fill").font = font_small
    ws.cell(row=legend_r + 2, column=left, value="TRANS beam: yellow fill").font = font_small
    ws.cell(row=legend_r + 3, column=left, value="Worst utilization member: RED fill").font = font_small
    ws.cell(row=legend_r + 4, column=left, value="Point load: red ●ID, Allocation marker: blue ▲ID (◎ID if both overlap)").font = font_small

    return legend_r + 6


# -----------------------------
# Main
# -----------------------------

# -----------------------------
# Verbose / debug sheets writer
# -----------------------------
def _safe_json(obj: object) -> str:
    try:
        return json.dumps(obj, ensure_ascii=False, default=str)
    except Exception:
        return str(obj)


def write_verbose_sheets(
    wb,
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    sections: List[Section],
    cand_rows: List["CandidateRow"],
    best: Optional["Solution"],
) -> None:
    """
    Writes detailed intermediate outputs per calculation flow.
    Sheets:
      - TRACE : chronological trace (phase/message + JSON payload)
      - DEBUG_MEMBER_TRIALS : per-member per-section trial results
      - DEBUG_MAIN_GEOMS : main beam geometry per candidate
      - DEBUG_TRANS_DEFS : transfer beam definitions per candidate
      - DEBUG_MEMBER_FINAL : selected member checks per candidate
      - DEBUG_ALLOC_FINAL : allocations per candidate (incl. TRANS reactions)
      - SAMPLE : arrays (x, V, M, y) for worst member of best solution
    """
    write_trace_sheet(wb)
    write_debug_member_trials_sheet(wb)
    write_debug_main_geoms_sheet(wb)
    write_debug_trans_defs_sheet(wb)
    write_debug_member_final_sheet(wb)
    write_debug_alloc_final_sheet(wb)
    write_sample_sheet(wb, cfg, mat, setts, sections, best)


def write_trace_sheet(wb) -> None:
    if "TRACE" in wb.sheetnames:
        del wb["TRACE"]
    ws = wb.create_sheet("TRACE")

    headers = ["seq", "time", "cand_id", "direction", "pitch", "member_id", "member_type", "phase", "message", "data_json"]
    widths = [6, 20, 8, 8, 10, 12, 10, 18, 40, 80]
    for i, w in enumerate(widths, start=1):
        set_col_width(ws, i, w)

    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)

    r = 2
    for e in _TRACE:
        ws.cell(row=r, column=1, value=e.get("seq"))
        ws.cell(row=r, column=2, value=e.get("time"))
        ws.cell(row=r, column=3, value=e.get("cand_id"))
        ws.cell(row=r, column=4, value=e.get("direction"))
        ws.cell(row=r, column=5, value=e.get("pitch"))
        ws.cell(row=r, column=6, value=e.get("member_id"))
        ws.cell(row=r, column=7, value=e.get("member_type"))
        ws.cell(row=r, column=8, value=e.get("phase"))
        ws.cell(row=r, column=9, value=e.get("message"))
        ws.cell(row=r, column=10, value=_safe_json(e.get("data", {})))
        r += 1


def write_debug_member_trials_sheet(wb) -> None:
    if "DEBUG_MEMBER_TRIALS" in wb.sheetnames:
        del wb["DEBUG_MEMBER_TRIALS"]
    ws = wb.create_sheet("DEBUG_MEMBER_TRIALS")

    headers = [
        "cand_id", "direction", "pitch", "member_id", "member_type",
        "sec_rank", "sec_name", "w_g[kN/m]",
        "span[m]", "w_avg[kN/m]", "w_max[kN/m]",
        "Ra[kN]", "Rb[kN]",
        "Mmax[kN*m]", "x_Mmax[m]",
        "Vmax[kN]", "x_Vmax[m]",
        "dmax[mm]", "x_dmax[m]",
        "util_M", "util_V", "util_d", "util_max",
        "ok", "note",
    ]
    widths = [8, 8, 10, 12, 10, 8, 20, 10, 10, 12, 12, 10, 10, 12, 10, 10, 10, 10, 10, 10, 10, 10, 10, 8, 40]
    for i, w in enumerate(widths, start=1):
        set_col_width(ws, i, w)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)

    r = 2
    for d in _DBG_MEMBER_TRIALS:
        ws.cell(row=r, column=1, value=d.get("cand_id"))
        ws.cell(row=r, column=2, value=d.get("direction"))
        ws.cell(row=r, column=3, value=d.get("pitch"))
        ws.cell(row=r, column=4, value=d.get("member_id"))
        ws.cell(row=r, column=5, value=d.get("member_type"))
        ws.cell(row=r, column=6, value=d.get("sec_rank"))
        ws.cell(row=r, column=7, value=d.get("sec_name"))
        ws.cell(row=r, column=8, value=d.get("w_g"))
        ws.cell(row=r, column=9, value=d.get("span"))
        ws.cell(row=r, column=10, value=d.get("w_udl"))
        ws.cell(row=r, column=11, value=d.get("w_udl_max"))
        ws.cell(row=r, column=12, value=d.get("Ra"))
        ws.cell(row=r, column=13, value=d.get("Rb"))
        ws.cell(row=r, column=14, value=d.get("Mmax"))
        ws.cell(row=r, column=15, value=d.get("x_Mmax"))
        ws.cell(row=r, column=16, value=d.get("Vmax"))
        ws.cell(row=r, column=17, value=d.get("x_Vmax"))
        dm = d.get("dmax")
        ws.cell(row=r, column=18, value=(dm * 1000.0 if isinstance(dm, (int, float)) else dm))
        ws.cell(row=r, column=19, value=d.get("x_dmax"))
        ws.cell(row=r, column=20, value=d.get("util_M"))
        ws.cell(row=r, column=21, value=d.get("util_V"))
        ws.cell(row=r, column=22, value=d.get("util_d"))
        ws.cell(row=r, column=23, value=d.get("util_max"))
        ws.cell(row=r, column=24, value=d.get("ok"))
        ws.cell(row=r, column=25, value=d.get("note"))
        r += 1


def write_debug_main_geoms_sheet(wb) -> None:
    if "DEBUG_MAIN_GEOMS" in wb.sheetnames:
        del wb["DEBUG_MAIN_GEOMS"]
    ws = wb.create_sheet("DEBUG_MAIN_GEOMS")
    headers = ["cand_id", "direction", "pitch", "beam_id", "pos", "span", "trib_width", "trib_left", "trib_right", "n_point"]
    widths = [8, 8, 10, 10, 10, 10, 12, 12, 12, 8]
    for i, w in enumerate(widths, start=1):
        set_col_width(ws, i, w)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    r = 2
    for d in _DBG_MAIN_GEOMS:
        for i, key in enumerate(["cand_id","direction","pitch","beam_id","pos","span","trib_width","trib_left","trib_right","n_point"], start=1):
            ws.cell(row=r, column=i, value=d.get(key))
        r += 1


def write_debug_trans_defs_sheet(wb) -> None:
    if "DEBUG_TRANS_DEFS" in wb.sheetnames:
        del wb["DEBUG_TRANS_DEFS"]
    ws = wb.create_sheet("DEBUG_TRANS_DEFS")
    headers = ["cand_id", "direction", "pitch", "tb_id", "span_dir", "fixed_coord", "left_support", "right_support", "left_pos", "right_pos", "a_tb", "a_on_support", "load_id", "P"]
    widths = [8, 8, 10, 18, 8, 10, 12, 12, 10, 10, 10, 12, 10, 10]
    for i, w in enumerate(widths, start=1):
        set_col_width(ws, i, w)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    r = 2
    for d in _DBG_TRANS_DEFS:
        vals = [d.get(k) for k in ["cand_id","direction","pitch","tb_id","span_dir","fixed_coord","left_support","right_support","left_pos","right_pos","a_tb","a_on_support","load_id","P"]]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)
        r += 1


def write_debug_member_final_sheet(wb) -> None:
    if "DEBUG_MEMBER_FINAL" in wb.sheetnames:
        del wb["DEBUG_MEMBER_FINAL"]
    ws = wb.create_sheet("DEBUG_MEMBER_FINAL")
    headers = [
        "cand_id", "direction", "pitch", "beam_id", "beam_no", "member_type", "span_dir",
        "pos_or_fixed", "left", "right", "span",
        "sec_rank", "sec_name", "w_g", "w_avg", "w_max", "Ra", "Rb", "n_point",
        "Mmax", "x_Mmax", "Vmax", "x_Vmax", "dmax[mm]", "x_dmax", "util_max", "ok"
    ]
    widths = [8, 8, 10, 14, 8, 10, 8, 10, 10, 10, 10, 8, 20, 10, 10, 10, 10, 10, 8, 12, 10, 10, 10, 10, 10, 10, 8]
    for i, w in enumerate(widths, start=1):
        set_col_width(ws, i, w)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    r = 2
    for d in _DBG_MEMBER_FINAL:
        dm = d.get("dmax")
        vals = [
            d.get("cand_id"), d.get("direction"), d.get("pitch"), d.get("beam_id"), d.get("beam_no"), d.get("member_type"), d.get("span_dir"),
            d.get("pos_or_fixed"), d.get("left"), d.get("right"), d.get("span"),
            d.get("sec_rank"), d.get("sec_name"), d.get("w_g"), d.get("w_udl"), d.get("w_udl_max"), d.get("Ra"), d.get("Rb"), d.get("n_point"),
            d.get("Mmax"), d.get("x_Mmax"), d.get("Vmax"), d.get("x_Vmax"), (dm * 1000.0 if isinstance(dm, (int,float)) else dm), d.get("x_dmax"), d.get("util_max"), d.get("ok")
        ]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)
        r += 1


def write_debug_alloc_final_sheet(wb) -> None:
    if "DEBUG_ALLOC_FINAL" in wb.sheetnames:
        del wb["DEBUG_ALLOC_FINAL"]
    ws = wb.create_sheet("DEBUG_ALLOC_FINAL")
    headers = ["cand_id", "direction", "pitch", "load_id", "support", "P_alloc[kN]"]
    widths = [8, 8, 10, 10, 12, 12]
    for i, w in enumerate(widths, start=1):
        set_col_width(ws, i, w)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h).font = Font(bold=True)
    r = 2
    for d in _DBG_ALLOC_FINAL:
        vals = [d.get("cand_id"), d.get("direction"), d.get("pitch"), d.get("load_id"), d.get("support"), d.get("P_alloc")]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)
        r += 1


def write_sample_sheet(
    wb,
    cfg: Config,
    mat: Material,
    setts: SolverSettings,
    sections: List[Section],
    best: Optional["Solution"],
) -> None:
    if "SAMPLE" in wb.sheetnames:
        del wb["SAMPLE"]
    ws = wb.create_sheet("SAMPLE")

    ws.cell(row=1, column=1, value="SAMPLE arrays for worst member of best solution").font = Font(bold=True, size=14)

    if best is None or (not best.ok):
        ws.cell(row=3, column=1, value="NO FEASIBLE SOLUTION").font = Font(bold=True)
        return

    worst_id = best.worst_member_id
    mc = next((m for m in best.member_checks if m.beam_id == worst_id), None)
    if mc is None:
        ws.cell(row=3, column=1, value="Worst member not found").font = Font(bold=True)
        return

    sec = next((s for s in sections if s.name == mc.section_name and s.rank == mc.section_rank), None)
    if sec is None:
        sec = next((s for s in sections if s.name == mc.section_name), None)
    if sec is None:
        ws.cell(row=3, column=1, value="Section not found for worst member").font = Font(bold=True)
        return

    ws.cell(row=3, column=1, value=f"Best: dir={best.direction}, pitch={best.pitch} m").font = Font(bold=True)
    ws.cell(row=4, column=1, value=f"Worst member: {worst_id} (No={mc.beam_no}, type={mc.member_type})").font = Font(bold=True)
    ws.cell(row=5, column=1, value=f"Section: rank={sec.rank}, {sec.name}").font = Font(bold=True)
    ws.cell(row=6, column=1, value=f"util_max={mc.util_max:.3f}, Mmax={mc.Mmax:.3f} kN*m, Vmax={mc.Vmax:.3f} kN, dmax={mc.dmax*1000.0:.3f} mm").font = Font(bold=True)

    # rebuild arrays by re-running check_member with return_arrays=True
    if mc.member_type == "MAIN":
        bg = next((b for b in best.main_geoms if b.beam_id == worst_id), None)
        if bg is None:
            ws.cell(row=8, column=1, value="BeamGeom not found").font = Font(bold=True)
            return
        bc, ana = check_member(
            member_type="MAIN",
            span_dir=bg.direction,
            span=bg.span,
            pos_or_fixed=bg.pos,
            left=None,
            right=None,
            trib_width=bg.trib_width,
            trib_left=bg.trib_left,
            trib_right=bg.trib_right,
            load_share_model=cfg.load_share_model,
            q=cfg.q,
            point_loads=bg.point_along,
            section=sec,
            mat=mat,
            setts=setts,
            return_arrays=True,
        )
    else:
        td = next((t for t in best.transfer_defs if t.tb_id == worst_id), None)
        if td is None:
            ws.cell(row=8, column=1, value="TransferDef not found").font = Font(bold=True)
            return
        Ltb = td.right_pos - td.left_pos
        bc, ana = check_member(
            member_type="TRANS",
            span_dir=td.span_dir,
            span=Ltb,
            pos_or_fixed=td.fixed_coord,
            left=td.left_pos,
            right=td.right_pos,
            trib_width=0.0,
            trib_left=0.0,
            trib_right=0.0,
            load_share_model="ONEWAY",
            q=cfg.q,
            point_loads=[(td.P, td.a_tb)],
            section=sec,
            mat=mat,
            setts=setts,
            return_arrays=True,
        )

    if ana is None:
        ws.cell(row=8, column=1, value="No arrays returned").font = Font(bold=True)
        return

    xs = ana.get("xs", [])
    V = ana.get("V", [])
    M = ana.get("M", [])
    y = ana.get("y", [])

    start_r = 9
    headers = ["x[m]", "V[kN]", "M[kN*m]", "y[mm]"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_r, column=c, value=h).font = Font(bold=True)

    r = start_r + 1
    for i in range(len(xs)):
        ws.cell(row=r, column=1, value=float(xs[i]))
        ws.cell(row=r, column=2, value=float(V[i]) if i < len(V) else None)
        ws.cell(row=r, column=3, value=float(M[i]) if i < len(M) else None)
        ws.cell(row=r, column=4, value=float(y[i]) * 1000.0 if i < len(y) else None)
        r += 1


def main(argv: List[str]) -> int:
    if len(argv) < 2:
        print("Usage: python beam_optimizer_b11.py input.xlsx output.xlsx [max_workers]")
        return 2

    in_path = argv[1]
    out_path = argv[2] if len(argv) >= 3 else (in_path.replace(".xlsx", "") + "_out.xlsx")
    max_workers = int(argv[3]) if len(argv) >= 4 else None
    run(in_path, out_path, max_workers=max_workers)

    print(f"Done. Output written to: {out_path}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main(sys.argv))
    except PermissionError as e:
        print(f"[ERROR] Permission denied. Close the target Excel file and retry: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)
